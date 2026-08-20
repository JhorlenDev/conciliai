import csv
import io
import os
import shutil
import re
import subprocess
import tempfile
import textwrap
import zipfile
import logging
from collections import defaultdict
from datetime import date
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4
from xml.etree import ElementTree

import fitz
from openpyxl import load_workbook
from fastapi import APIRouter, Depends, File, HTTPException, Response as FastAPIResponse, UploadFile
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.config import UPLOAD_DIR
from app.core.database import get_db
from app.models import Arquivo, Cliente, Comprovante, ComprovanteRfb, ComprovanteRfbItem, Conciliacao, ContaBancaria, Correspondencia, DocumentoImportante, LancamentoContabil, MovimentoExtrato, NotaFiscal, ProcessoConciliacao, RegraContabil, RegraContabilExcecao
from app.services.normalization import accounting_nature, is_statement_debit, normalize_name, normalize_rule_accounting_nature, normalize_statement_nature
from app.services.matching import names_similar
from app.services.parsers import deduplicate_statement_records, extract_basa_pdfplumber_pages, extract_getnet_pdfplumber_pages, extract_invoices, extract_loan_receipts, extract_receipts, extract_santander_pdfplumber_statement, extract_statement_pages, parse_brl
from app.services.getnet_adjustments import GETNET_ADJUSTMENT_COMPONENT, GETNET_ADJUSTMENT_DESCRIPTION, is_getnet_adjustment_movement, is_santander_getnet_credit, sync_getnet_anticipation_adjustments
from app.services.rfb import belongs_to_selected_bank, extract_competence, parse_rfb_page
from app.services.rule_source import choose_rule_source

router = APIRouter()
PRIMARY_BANKS = ["Banco do Brasil", "Santander", "BASA", "Bradesco", "Caixa", "Conta Caixa", "Notas", "Apropriações", "Empréstimos/Financiamentos"]
LEGACY_BANKS = ["Vendas com Cartão", "Comissões Getnet", "Empréstimos/Financeiro"]
BANKS = [*PRIMARY_BANKS, *LEGACY_BANKS]
MACHINE_STATEMENT_DOCUMENT_TYPES = {"maquininha_extrato", "getnet_extrato", "getnet_vendas", "getnet_comissoes"}
LOAN_DOCUMENT_TYPE = "emprestimo"
INVOICE_DOCUMENT_TYPE = "nota"
RECEIPT_DOCUMENT_TYPES = {"comprovante", LOAN_DOCUMENT_TYPE, *MACHINE_STATEMENT_DOCUMENT_TYPES}
DOCUMENT_TYPES = {"extrato", *RECEIPT_DOCUMENT_TYPES, INVOICE_DOCUMENT_TYPE, "rfb"}
logger = logging.getLogger(__name__)
SCANNED_PDF_OCR_MESSAGE = "PDF escaneado sem texto pesquisável. Instale o OCR Tesseract (tesseract-ocr e tesseract-ocr-por) para extrair este documento."


def extract_basa_initial_balance(text: str) -> Decimal | None:
    match = re.search(r"Saldo Dispon[ií]vel Inicial:\s*([+-]?\d{1,3}(?:\.\d{3})*,\d{2})", text, re.I)
    if match:
        return parse_brl(match.group(1))

    lines = [line.strip() for line in text.splitlines()]
    for index, line in enumerate(lines):
        if "SALDO DISPONIVEL INICIAL" not in normalize_name(line):
            continue
        window = " ".join(lines[index : index + 4])
        amount_match = re.search(r"([+-]?\d{1,3}(?:\.\d{3})*,\d{2})", window)
        if amount_match:
            return parse_brl(amount_match.group(1))
    return None


def statement_initial_balance(reconciliation: Conciliacao, db: Session) -> Decimal:
    if reconciliation.banco != "BASA":
        return Decimal("0.00")
    files = db.query(Arquivo).filter_by(conciliacao_id=reconciliation.id, tipo_documento="extrato", ativo=True).all()
    text = "\n".join(item.texto_bruto or "" for item in files)
    return extract_basa_initial_balance(text) or Decimal("0.00")


@router.get("/arquivos/{arquivo_id}/visualizar")
def view_file(arquivo_id: str, db: Session = Depends(get_db)):
    record = db.get(Arquivo, arquivo_id)
    if not record or not Path(record.caminho).is_file():
        raise HTTPException(404, "Documento original não encontrado")
    return FileResponse(Path(record.caminho), media_type="application/pdf", headers={"Content-Disposition": f'inline; filename="{record.nome_original}"'})


@router.delete("/arquivos/{arquivo_id}", status_code=204)
def delete_file(arquivo_id: str, db: Session = Depends(get_db)):
    record = db.get(Arquivo, arquivo_id)
    if not record:
        raise HTTPException(404, "Arquivo não encontrado")
    reset_reconciliation(record.conciliacao_id, db)
    if record.tipo_documento == "extrato":
        db.query(MovimentoExtrato).filter_by(arquivo_id=record.id).delete()
    elif record.tipo_documento in RECEIPT_DOCUMENT_TYPES:
        db.query(Comprovante).filter_by(arquivo_id=record.id).delete()
    elif record.tipo_documento == INVOICE_DOCUMENT_TYPE:
        db.query(NotaFiscal).filter_by(arquivo_id=record.id).delete()
    else:
        receipts = db.query(ComprovanteRfb).filter_by(arquivo_id=record.id).all()
        for receipt in receipts:
            db.query(ComprovanteRfbItem).filter_by(comprovante_rfb_id=receipt.id).delete()
        db.query(ComprovanteRfb).filter_by(arquivo_id=record.id).delete()
    Path(record.caminho).unlink(missing_ok=True)
    db.delete(record)
    reconciliation = db.get(Conciliacao, record.conciliacao_id)
    if reconciliation:
        sync_getnet_anticipation_adjustments(reconciliation, db)
    db.commit()


class ClienteInput(BaseModel):
    nome: str
    documento: str | None = None


class ClienteUpdate(BaseModel):
    nome: str
    documento: str | None = None


class ConciliacaoInput(BaseModel):
    cliente_id: str
    banco: str
    data_inicio: date
    data_fim: date


class ProcessoConciliacaoInput(BaseModel):
    cliente_id: str
    data_inicio: date
    data_fim: date
    banco: str | None = None


class ProcessoBancoInput(BaseModel):
    banco: str


class RegraContabilInput(BaseModel):
    gatilho: str
    gatilho_comprovante: str = ""
    texto_exclusao: str = ""
    natureza: str
    conta_debito: str
    conta_credito: str
    historico: str
    complemento: str = ""
    tipo_componente: str = ""
    escopo: str = "global"


class RegraContabilPreviaInput(BaseModel):
    gatilho: str = ""
    gatilho_comprovante: str = ""
    texto_exclusao: str = ""
    natureza: str
    tipo_componente: str = ""
    regra_id: str = ""


class RegraFonteInput(BaseModel):
    gatilho: str
    conta_debito: str
    conta_credito: str
    historico: str
    complemento: str = ""
    escopo: str = "global"


class ContaBancariaInput(BaseModel):
    conta_contabil: str


class ContaBancariaClienteInput(BaseModel):
    agencia: str = ""
    conta: str
    titular: str


class LancamentoItemInput(BaseModel):
    id: str = ""
    componente: str
    valor: Decimal
    efeito_no_total: str = "SOMA"
    conta_debito: str
    conta_credito: str
    historico: str
    complemento: str = ""
    descricao: str = ""
    tributo: str = ""
    codigo_receita: str = ""


class LancamentosInput(BaseModel):
    itens: list[LancamentoItemInput]


class MovimentoUsoInput(BaseModel):
    usar: bool = True


@router.get("/bancos")
def banks():
    return PRIMARY_BANKS


def important_document_payload(record: DocumentoImportante) -> dict:
    return {"id": record.id, "tipo": record.tipo, "nome": record.nome_original, "extensao": record.extensao, "criado_em": record.created_at.isoformat(), "itens_extraidos": len(record.catalogo.get("contas" if record.tipo == "plano_contas" else "historicos", []))}


def extract_important_catalog(path: Path, extension: str, tipo: str) -> dict:
    if extension == ".pdf":
        with fitz.open(path) as document:
            rows = [line.strip() for page in document for line in page.get_text().splitlines() if line.strip()]
    else:
        rows = []
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            for sheet in workbook.worksheets:
                account_columns = (0, 1)
                for row in sheet.iter_rows(values_only=True):
                    raw_values = [str(value).strip() if value is not None else "" for value in row]
                    values = [value for value in raw_values if value]
                    if tipo == "plano_contas":
                        headers = [value.lower().replace("ó", "o") for value in raw_values]
                        if "codigo" in headers and "nome" in headers:
                            account_columns = (headers.index("codigo"), headers.index("nome"))
                            continue
                        values = [raw_values[index] for index in account_columns if index < len(raw_values) and raw_values[index]]
                    if values:
                        rows.append(" - ".join(values))
        except Exception:
            # Some exports contain invalid styles. Extract cell values directly.
            namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            with zipfile.ZipFile(path) as workbook:
                shared = []
                if "xl/sharedStrings.xml" in workbook.namelist():
                    root = ElementTree.fromstring(workbook.read("xl/sharedStrings.xml"))
                    shared = ["".join(item.itertext()) for item in root.findall("x:si", namespace)]
                sheets = sorted(item for item in workbook.namelist() if item.startswith("xl/worksheets/") and item.endswith(".xml"))
                for name in sheets:
                    root = ElementTree.fromstring(workbook.read(name))
                    account_columns = (0, 1)
                    for row in root.findall(".//x:row", namespace):
                        values = []
                        for cell in row.findall("x:c", namespace):
                            value = cell.findtext("x:v", default="", namespaces=namespace)
                            if cell.get("t") == "s" and value.isdigit():
                                value = shared[int(value)] if int(value) < len(shared) else ""
                            elif cell.get("t") == "inlineStr":
                                value = "".join(cell.itertext())
                            if value.strip():
                                values.append(value.strip())
                        if tipo == "plano_contas":
                            headers = [value.lower().replace("ó", "o") for value in values]
                            if "codigo" in headers and "nome" in headers:
                                account_columns = (headers.index("codigo"), headers.index("nome"))
                                continue
                            values = [values[index] for index in account_columns if index < len(values)]
                        if values:
                            rows.append(" - ".join(values))
    # Keep meaningful rows only; account plans commonly begin with a numeric code.
    values = []
    for row in rows:
        text = re.sub(r"\s+", " ", row).strip(" -")
        if len(text) < 2:
            continue
        if tipo == "plano_contas":
            match = re.match(r"^((?:\d+[.\-]?)+)\s*[-: ]\s*(.+)$", text)
            if not match:
                continue
            text = f"{match.group(1)} - {match.group(2)}"
        values.append(text)
    key = "contas" if tipo == "plano_contas" else "historicos"
    return {key: list(dict.fromkeys(values)), "formato": "catalogo_v3"}


def extract_pdf_pages(path: Path, bank: str, document_type: str) -> list[str]:
    if document_type == "extrato" and bank == "Santander":
        pages = extract_santander_pdfplumber_pages(path)
        if pages:
            return pages
    if document_type == "extrato" and bank == "BASA":
        pages = extract_basa_pdfplumber_pages(path)
        if pages:
            return pages
    if document_type in MACHINE_STATEMENT_DOCUMENT_TYPES:
        pages = extract_getnet_pdfplumber_pages(path)
        if pages:
            return pages
    with fitz.open(path) as document:
        pages = [page.get_text() for page in document]
    if document_type == LOAN_DOCUMENT_TYPE and not any(page.strip() for page in pages):
        pages = extract_scanned_pdf_pages(path)
    if document_type == LOAN_DOCUMENT_TYPE and any("Cronograma Reposicao Exigivel" in page or "Cronograma Reposição Exigível" in page for page in pages):
        return ["\n".join(pages)]
    if document_type == "comprovante" and bank == "Caixa" and any("Comprovante de Pagamento de Boleto" in page for page in pages):
        return ["\n".join(pages)]
    return pages


def available_tesseract_languages(tesseract: str) -> set[str]:
    try:
        result = subprocess.run([tesseract, "--list-langs"], capture_output=True, text=True, timeout=10, check=False, env=tesseract_environment())
    except (OSError, subprocess.SubprocessError):
        return set()
    return {line.strip() for line in result.stdout.splitlines() if line.strip() and not line.lower().startswith("list of")}


def tesseract_language_argument(languages: set[str]) -> str:
    selected = [language for language in ("por", "eng") if language in languages]
    return "+".join(selected) if selected else "por+eng"


def local_tesseract_path() -> str | None:
    local = Path.cwd() / ".local" / "ocr" / "root" / "usr" / "bin" / "tesseract"
    return str(local) if local.is_file() else None


def tesseract_environment() -> dict[str, str]:
    env = dict(os.environ)
    local_root = Path.cwd() / ".local" / "ocr" / "root"
    tessdata = local_root / "usr" / "share" / "tesseract-ocr" / "5" / "tessdata"
    local_lib = local_root / "usr" / "lib" / "x86_64-linux-gnu"
    if tessdata.is_dir():
        env["TESSDATA_PREFIX"] = str(tessdata)
    if local_lib.is_dir():
        current = env.get("LD_LIBRARY_PATH", "")
        env["LD_LIBRARY_PATH"] = f"{local_lib}:{current}" if current else str(local_lib)
    return env


def extract_scanned_pdf_pages(path: Path) -> list[str]:
    tesseract = shutil.which("tesseract") or local_tesseract_path()
    if not tesseract:
        raise ValueError(SCANNED_PDF_OCR_MESSAGE)
    languages = available_tesseract_languages(tesseract)
    language = tesseract_language_argument(languages)
    env = tesseract_environment()
    pages: list[str] = []
    with fitz.open(path) as document, tempfile.TemporaryDirectory() as tmpdir:
        for number, page in enumerate(document, 1):
            image_path = Path(tmpdir) / f"page-{number}.png"
            pixmap = page.get_pixmap(matrix=fitz.Matrix(3, 3), alpha=False)
            pixmap.save(image_path)
            result = subprocess.run(
                [tesseract, str(image_path), "stdout", "-l", language, "--psm", "6"],
                capture_output=True,
                text=True,
                timeout=60,
                check=False,
                env=env,
            )
            if result.returncode != 0:
                raise ValueError(f"OCR falhou na página {number}: {result.stderr.strip() or 'erro desconhecido'}")
            pages.append(result.stdout)
    if not any(page.strip() for page in pages):
        raise ValueError("OCR executado, mas nenhum texto foi reconhecido no PDF.")
    return pages


def extract_santander_pdfplumber_pages(path: Path) -> list[str]:
    try:
        import pdfplumber
    except ImportError:
        logger.warning("pdfplumber não instalado; usando PyMuPDF para extrato Santander")
        return []

    try:
        with pdfplumber.open(path) as document:
            pages = [
                page.extract_text(x_tolerance=1, y_tolerance=3, layout=True) or ""
                for page in document.pages
            ]
    except Exception as error:
        logger.warning("Falha ao ler Santander com pdfplumber: %s", error)
        return []

    if any("EXTRATO CONSOLIDADO INTELIGENTE" in page and "Movimentos" in page for page in pages):
        return pages
    return []


def extract_statement_document(path: Path, bank: str, document_type: str) -> tuple[list[str], list]:
    if document_type == "extrato" and bank == "Santander":
        santander = extract_santander_pdfplumber_statement(path)
        if santander:
            return santander.pages, santander.records
    pages = extract_pdf_pages(path, bank, document_type)
    return pages, extract_statement_pages(pages, bank)


@router.get("/documentos-importantes")
def list_important_documents(db: Session = Depends(get_db)):
    return [important_document_payload(item) for item in db.query(DocumentoImportante).order_by(DocumentoImportante.created_at.desc())]


@router.get("/documentos-importantes/catalogo")
def important_document_catalog(db: Session = Depends(get_db)):
    accounts, histories = [], []
    changed = False
    for item in db.query(DocumentoImportante).all():
        if item.catalogo.get("formato") != "catalogo_v3" and Path(item.caminho).is_file():
            try:
                item.catalogo = extract_important_catalog(Path(item.caminho), item.extensao, item.tipo)
                changed = True
            except Exception:
                pass
        for account in item.catalogo.get("contas", []):
            match = re.match(r"^((?:\d+[.\-]?)+)\s*[-: ]\s*(.+)$", re.sub(r"\s+", " ", account).strip(" -"))
            if match:
                accounts.append(f"{match.group(1)} - {match.group(2)}")
        histories.extend(item.catalogo.get("historicos", []))
    if changed:
        db.commit()
    return {"contas": list(dict.fromkeys(accounts)), "historicos": list(dict.fromkeys(histories))}


@router.post("/documentos-importantes")
def upload_important_document(tipo: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    extension = Path(file.filename or "").suffix.lower()
    if tipo not in {"plano_contas", "historicos"}:
        raise HTTPException(422, "Tipo de documento inválido")
    if extension == ".xls":
        raise HTTPException(422, "Arquivos .xls não são suportados. Salve o arquivo como .xlsx e tente novamente.")
    if extension not in {".pdf", ".xlsx"}:
        raise HTTPException(422, "Envie um PDF ou arquivo XLSX")
    destination = UPLOAD_DIR / "documentos_importantes"
    destination.mkdir(parents=True, exist_ok=True)
    path = destination / f"{uuid4()}{extension}"
    with path.open("wb") as output:
        shutil.copyfileobj(file.file, output)
    try:
        catalog = extract_important_catalog(path, extension, tipo)
    except Exception as error:
        path.unlink(missing_ok=True)
        raise HTTPException(422, f"Não foi possível extrair o documento: {error}") from error
    record = DocumentoImportante(tipo=tipo, nome_original=file.filename, caminho=str(path), extensao=extension, catalogo=catalog)
    db.add(record); db.commit(); db.refresh(record)
    return important_document_payload(record)


@router.delete("/documentos-importantes/{documento_id}", status_code=204)
def delete_important_document(documento_id: str, db: Session = Depends(get_db)):
    record = db.get(DocumentoImportante, documento_id)
    if not record:
        raise HTTPException(404, "Documento não encontrado")
    Path(record.caminho).unlink(missing_ok=True)
    db.delete(record); db.commit()


@router.get("/clientes")
def list_clients(db: Session = Depends(get_db)):
    return [{"id": item.id, "nome": item.nome, "documento": item.documento} for item in db.query(Cliente).order_by(Cliente.nome)]


@router.post("/clientes")
def create_client(payload: ClienteInput, db: Session = Depends(get_db)):
    client = Cliente(**payload.model_dump())
    db.add(client); db.commit(); db.refresh(client)
    return {"id": client.id, "nome": client.nome, "documento": client.documento}


@router.patch("/clientes/{cliente_id}")
def update_client(cliente_id: str, payload: ClienteUpdate, db: Session = Depends(get_db)):
    client = db.get(Cliente, cliente_id)
    if not client:
        raise HTTPException(404, "Cliente não encontrado")
    client.nome = payload.nome
    client.documento = payload.documento
    db.commit(); db.refresh(client)
    return {"id": client.id, "nome": client.nome, "documento": client.documento}


@router.delete("/clientes/{cliente_id}", status_code=204)
def delete_client(cliente_id: str, db: Session = Depends(get_db)):
    client = db.get(Cliente, cliente_id)
    if not client:
        raise HTTPException(404, "Cliente não encontrado")
    if db.query(Conciliacao).filter_by(cliente_id=cliente_id).first():
        raise HTTPException(409, "Não é possível excluir cliente com conciliações vinculadas")
    db.delete(client); db.commit()


def client_bank_account_payload(account: ContaBancaria) -> dict:
    return {"id": account.id, "banco": account.banco, "agencia": account.agencia, "conta": account.conta, "titular": account.titular, "conta_contabil": account.conta_contabil}


@router.get("/clientes/{cliente_id}/contas-bancarias")
def client_bank_accounts(cliente_id: str, db: Session = Depends(get_db)):
    if not db.get(Cliente, cliente_id):
        raise HTTPException(404, "Cliente não encontrado")
    accounts = db.query(ContaBancaria).filter_by(cliente_id=cliente_id).order_by(ContaBancaria.banco).all()
    return [client_bank_account_payload(account) for account in accounts]


@router.put("/clientes/{cliente_id}/contas-bancarias/{banco}")
def save_client_bank_account(cliente_id: str, banco: str, payload: ContaBancariaClienteInput, db: Session = Depends(get_db)):
    client = db.get(Cliente, cliente_id)
    if not client:
        raise HTTPException(404, "Cliente não encontrado")
    if banco not in BANKS:
        raise HTTPException(422, "Banco inválido")
    if not payload.conta.strip() or not payload.titular.strip():
        raise HTTPException(422, "Informe conta e titular")
    account = db.query(ContaBancaria).filter_by(cliente_id=cliente_id, banco=banco).first()
    if not account:
        account = ContaBancaria(cliente_id=cliente_id, banco=banco)
        db.add(account)
    account.agencia = payload.agencia.strip()
    account.conta = payload.conta.strip()
    account.titular = payload.titular.strip()
    db.commit(); db.refresh(account)
    return client_bank_account_payload(account)


@router.delete("/clientes/{cliente_id}/contas-bancarias/{banco}", status_code=204)
def delete_client_bank_account(cliente_id: str, banco: str, db: Session = Depends(get_db)):
    account = db.query(ContaBancaria).filter_by(cliente_id=cliente_id, banco=banco).first()
    if not account:
        raise HTTPException(404, "Conta bancária não encontrada")
    db.delete(account); db.commit()


def reconciliation_progress(item: Conciliacao, db: Session) -> dict:
    total = (
        db.query(MovimentoExtrato.id)
        .filter(
            MovimentoExtrato.conciliacao_id == item.id,
            MovimentoExtrato.ativo == True,  # noqa: E712
            MovimentoExtrato.ignorado_no_periodo == False,  # noqa: E712
            MovimentoExtrato.data >= item.data_inicio,
            MovimentoExtrato.data <= item.data_fim,
        )
        .count()
    )
    covered = (
        db.query(Correspondencia.movimento_extrato_id)
        .join(MovimentoExtrato, Correspondencia.movimento_extrato_id == MovimentoExtrato.id)
        .join(LancamentoContabil, LancamentoContabil.correspondencia_id == Correspondencia.id)
        .filter(
            Correspondencia.conciliacao_id == item.id,
            MovimentoExtrato.ativo == True,  # noqa: E712
            MovimentoExtrato.ignorado_no_periodo == False,  # noqa: E712
            MovimentoExtrato.data >= item.data_inicio,
            MovimentoExtrato.data <= item.data_fim,
            LancamentoContabil.status == "aplicado_por_regra",
            LancamentoContabil.regra_contabil_id.isnot(None),
        )
        .distinct()
        .count()
    )
    return {"total": total, "cobertos": covered, "percentual": round((covered / total) * 100) if total else 0}


def reconciliation_payload(item: Conciliacao, db: Session | None = None) -> dict:
    payload = {"id": item.id, "processo_id": item.processo_id, "banco": item.banco, "status": item.status, "data_inicio": item.data_inicio.isoformat(), "data_fim": item.data_fim.isoformat()}
    if db:
        payload["progresso_regras"] = reconciliation_progress(item, db)
    return payload


def process_payload(item: ProcessoConciliacao, db: Session) -> dict:
    client = db.get(Cliente, item.cliente_id)
    reconciliations = db.query(Conciliacao).filter_by(processo_id=item.id).order_by(Conciliacao.created_at).all()
    return {"id": item.id, "cliente_id": item.cliente_id, "cliente_nome": client.nome if client else "Cliente removido", "data_inicio": item.data_inicio.isoformat(), "data_fim": item.data_fim.isoformat(), "criado_em": item.created_at.isoformat(), "status": item.status, "bancos": [reconciliation_payload(reconciliation, db) for reconciliation in reconciliations]}


@router.get("/processos-conciliacao")
def list_reconciliation_processes(cliente_id: str | None = None, db: Session = Depends(get_db)):
    query = db.query(ProcessoConciliacao)
    if cliente_id:
        query = query.filter_by(cliente_id=cliente_id)
    return [process_payload(item, db) for item in query.order_by(ProcessoConciliacao.updated_at.desc(), ProcessoConciliacao.created_at.desc())]


@router.post("/processos-conciliacao")
def create_reconciliation_process(payload: ProcessoConciliacaoInput, db: Session = Depends(get_db)):
    if not db.get(Cliente, payload.cliente_id) or payload.data_inicio > payload.data_fim:
        raise HTTPException(422, "Cliente ou período inválido")
    if payload.banco and payload.banco not in BANKS:
        raise HTTPException(422, "Banco inválido")
    process = ProcessoConciliacao(cliente_id=payload.cliente_id, data_inicio=payload.data_inicio, data_fim=payload.data_fim)
    db.add(process); db.flush()
    if payload.banco:
        db.add(Conciliacao(cliente_id=process.cliente_id, processo_id=process.id, banco=payload.banco, data_inicio=process.data_inicio, data_fim=process.data_fim))
    db.commit(); db.refresh(process)
    return process_payload(process, db)


@router.get("/processos-conciliacao/{processo_id}")
def get_reconciliation_process(processo_id: str, db: Session = Depends(get_db)):
    process = db.get(ProcessoConciliacao, processo_id)
    if not process:
        raise HTTPException(404, "Processo de conciliação não encontrado")
    return process_payload(process, db)


@router.delete("/processos-conciliacao/{processo_id}", status_code=204)
def delete_reconciliation_process(processo_id: str, db: Session = Depends(get_db)):
    process = db.get(ProcessoConciliacao, processo_id)
    if not process:
        raise HTTPException(404, "Processo de conciliação não encontrado")
    reconciliations = db.query(Conciliacao).filter_by(processo_id=process.id).all()
    for reconciliation in reconciliations:
        rules = db.query(RegraContabil).filter_by(conciliacao_id=reconciliation.id).all()
        local_rule_ids = [rule.id for rule in rules if rule.escopo == "periodo"]
        if local_rule_ids:
            db.query(RegraContabilExcecao).filter(RegraContabilExcecao.regra_contabil_id.in_(local_rule_ids)).delete(synchronize_session=False)
        for rule in rules:
            if rule.escopo == "global":
                rule.conciliacao_id = None
            else:
                rule.ativo = False
                rule.conciliacao_id = None
        db.query(RegraContabilExcecao).filter_by(conciliacao_id=reconciliation.id).delete(synchronize_session=False)
        matches = db.query(Correspondencia).filter_by(conciliacao_id=reconciliation.id).all()
        for match in matches:
            db.query(LancamentoContabil).filter_by(correspondencia_id=match.id).delete()
        db.query(Correspondencia).filter_by(conciliacao_id=reconciliation.id).delete()
        receipts = db.query(ComprovanteRfb).filter_by(conciliacao_id=reconciliation.id).all()
        for receipt in receipts:
            db.query(ComprovanteRfbItem).filter_by(comprovante_rfb_id=receipt.id).delete()
        db.query(ComprovanteRfb).filter_by(conciliacao_id=reconciliation.id).delete()
        db.query(MovimentoExtrato).filter_by(conciliacao_id=reconciliation.id).delete()
        db.query(Comprovante).filter_by(conciliacao_id=reconciliation.id).delete()
        db.query(NotaFiscal).filter_by(conciliacao_id=reconciliation.id).delete()
        files = db.query(Arquivo).filter_by(conciliacao_id=reconciliation.id).all()
        for file in files:
            Path(file.caminho).unlink(missing_ok=True)
            db.delete(file)
        db.delete(reconciliation)
    db.delete(process)
    db.commit()


@router.post("/processos-conciliacao/{processo_id}/bancos")
def resume_process_bank(processo_id: str, payload: ProcessoBancoInput, db: Session = Depends(get_db)):
    process = db.get(ProcessoConciliacao, processo_id)
    if not process:
        raise HTTPException(404, "Processo de conciliação não encontrado")
    if payload.banco not in BANKS:
        raise HTTPException(422, "Banco inválido")
    reconciliation = db.query(Conciliacao).filter_by(processo_id=process.id, banco=payload.banco).first()
    if not reconciliation:
        reconciliation = Conciliacao(cliente_id=process.cliente_id, processo_id=process.id, banco=payload.banco, data_inicio=process.data_inicio, data_fim=process.data_fim)
        process.status = "em_andamento"
        db.add(reconciliation); db.commit(); db.refresh(reconciliation)
    return reconciliation_payload(reconciliation)


@router.post("/conciliacoes")
def create_reconciliation(payload: ConciliacaoInput, db: Session = Depends(get_db)):
    if payload.banco not in BANKS or not db.get(Cliente, payload.cliente_id):
        raise HTTPException(422, "Cliente ou banco inválido")
    process = ProcessoConciliacao(cliente_id=payload.cliente_id, data_inicio=payload.data_inicio, data_fim=payload.data_fim)
    db.add(process); db.flush()
    item = Conciliacao(**payload.model_dump(), processo_id=process.id)
    db.add(item); db.commit(); db.refresh(item)
    return reconciliation_payload(item)


def trigger_matches(source: str, trigger: str) -> bool:
    normalized_trigger, normalized_source = normalize_name(trigger), normalize_name(source)
    if not normalized_trigger:
        return True
    if normalized_trigger in normalized_source:
        return True
    trigger_tokens = set(normalized_trigger.split())
    source_tokens = set(normalized_source.split())
    if trigger_tokens and trigger_tokens <= source_tokens:
        return True
    trigger_digits = re.sub(r"\D", "", trigger)
    source_digits = re.sub(r"\D", "", source)
    if not trigger_digits or re.search(r"[A-Z]", normalized_trigger):
        return False
    if trigger_digits in source_digits:
        return True
    stripped = trigger_digits.lstrip("0")
    return bool(len(stripped) >= 6 and stripped in source_digits)


def receipt_trigger_text(receipt: Comprovante | None, rfb: ComprovanteRfb | None = None, rfb_items: list[ComprovanteRfbItem] | None = None) -> str:
    bank_text = " ".join(item for item in [receipt.favorecido, receipt.beneficiario, receipt.beneficiario_final, receipt.nome_fantasia, receipt.numero_documento] if item) if receipt else ""
    rfb_text = " ".join(item for item in [rfb.tipo, rfb.razao_social, rfb.numero_documento, rfb.competencia, rfb.periodo_apuracao] if item) if rfb else ""
    item_text = " ".join(item for item in [f"{item.codigo} {item.descricao}" for item in (rfb_items or [])] if item)
    return " ".join(item for item in [bank_text, rfb_text, item_text] if item)


def receipt_description(receipt: Comprovante | None) -> str:
    if not receipt:
        return ""
    party = next((value.strip() for value in [receipt.beneficiario_final, receipt.beneficiario, receipt.favorecido] if value and value.strip()), "")
    operation = (receipt.tipo_operacao or "").strip()
    document = f"Documento: {receipt.numero_documento}" if receipt.numero_documento else ""
    if operation and party:
        return " ".join(item for item in [operation, party, document] if item)
    if party:
        return " ".join(item for item in [party, document] if item)
    fallback = next((value for value in [receipt.dados_normalizados.get("descricao", ""), receipt.dados_normalizados.get("historico", ""), receipt.texto_original] if isinstance(value, str) and value.strip()), "")
    return " ".join(fallback.split())


def is_loan_receipt(receipt: Comprovante | None) -> bool:
    original = receipt.dados_originais if receipt and isinstance(receipt.dados_originais, dict) else {}
    return bool(receipt and (original.get("tipo_documento") == LOAN_DOCUMENT_TYPE or "EMPR" in normalize_name(receipt.tipo_operacao) or "FINANCI" in normalize_name(receipt.tipo_operacao)))


def tax_complement(entry: LancamentoContabil, match: Correspondencia, db: Session) -> tuple[str, str, str, str]:
    rfb = db.get(ComprovanteRfb, match.comprovante_rfb_id) if match.comprovante_rfb_id else None
    receipt = db.get(Comprovante, match.comprovante_id) if match.comprovante_id else None
    movement = db.get(MovimentoExtrato, match.movimento_extrato_id)
    rfb_items = db.query(ComprovanteRfbItem).filter_by(comprovante_rfb_id=rfb.id).all() if rfb else []
    rule = db.get(RegraContabil, entry.regra_contabil_id) if entry.regra_contabil_id else None
    source = " ".join(str(value or "") for value in [entry.componente, entry.tributo, entry.codigo_receita, entry.descricao, entry.historico, movement.historico if movement else "", receipt.texto_original if receipt else "", rfb.tipo if rfb else "", rfb.razao_social if rfb else "", " ".join(f"{item.codigo} {item.descricao}" for item in rfb_items), rule.tipo_componente if rule else "", rule.historico if rule else ""])
    normalized = normalize_name(source)
    if "SIMPLES" in normalized:
        tax = "SIMPLES NACIONAL"
    elif entry.componente == "IRRF" or entry.codigo_receita.lstrip("0") in {"156", "561"} or "IRRF" in normalized:
        tax = "IRRF"
    elif entry.componente == "INSS" or "INSS" in normalized or "PREVIDENCI" in normalized:
        tax = "INSS"
    elif entry.componente == "FGTS" or "FGTS" in normalized:
        tax = "FGTS"
    else:
        return "", "", "", ""
    competence = extract_competence(rfb.texto_original, rfb.competencia, rfb.periodo_apuracao) if rfb else ""
    origin = "Comprovante RFB" if rfb else "Comprovante bancário" if receipt else ""
    complement = competence
    return tax, competence, complement, origin


def normalized_trigger_key(value: str) -> str:
    """Makes equivalent keyword sets stable regardless of order and punctuation."""
    return " ".join(sorted(normalize_name(value).split()))


def rule_identity(cliente_id: str | None, banco: str, natureza: str, componente: str, gatilho: str, gatilho_comprovante: str, texto_exclusao: str, conta_debito: str, conta_credito: str, historico: str, complemento: str) -> tuple[str, ...]:
    return (
        cliente_id or "",
        normalize_name(banco),
        normalize_rule_accounting_nature(natureza),
        normalize_name(componente or "PRINCIPAL"),
        normalized_trigger_key(gatilho),
        normalized_trigger_key(gatilho_comprovante),
        normalized_trigger_key(texto_exclusao),
        normalize_name(conta_debito),
        normalize_name(conta_credito),
        normalize_name(historico),
        normalize_name(complemento),
    )


def saved_rule_identity(rule: RegraContabil) -> tuple[str, ...]:
    return rule_identity(rule.cliente_id, rule.banco, rule.tipo_operacao, rule.tipo_componente, rule.favorecido_normalizado, rule.gatilho_comprovante_normalizado, rule.texto_exclusao_normalizado, rule.conta_debito, rule.conta_credito, rule.historico, rule.complemento)


def legacy_trigger_matches_receipt(rule: RegraContabil, history: str, receipt: Comprovante | None, rfb: ComprovanteRfb | None = None, rfb_items: list[ComprovanteRfbItem] | None = None) -> bool:
    """Supports rules saved before receipt triggers had their own field."""
    if rule.gatilho_comprovante_normalizado or not receipt:
        return False
    trigger_tokens = set(normalize_name(rule.favorecido_normalizado).split())
    statement_tokens = set(normalize_name(history).split())
    receipt_tokens = set(normalize_name(receipt_trigger_text(receipt, rfb, rfb_items)).split())
    if not trigger_tokens or not trigger_tokens <= statement_tokens | receipt_tokens:
        return False
    # A legacy combined trigger must still identify a party from the receipt.
    return bool(trigger_tokens & receipt_tokens)


def rule_match_history(movement: MovimentoExtrato, receipt: Comprovante | None) -> str:
    history = " ".join(item for item in [movement.historico, movement.nome_encontrado] if item)
    if receipt and receipt.beneficiario_final:
        if receipt.beneficiario and receipt.beneficiario.lower() in history.lower():
            return re.sub(re.escape(receipt.beneficiario), receipt.beneficiario_final, history, flags=re.I)
        return f"{movement.historico} {receipt.beneficiario_final}".strip()
    return history


def rule_component_key(component: str) -> str:
    return "PRINCIPAL" if component in {"", "PRINCIPAL", "VALOR_COBRADO"} else component


def rule_component_trigger_label(component: str) -> str:
    return {
        "DESCONTO_ABATIMENTO": "Desconto",
        "DESCONTO": "Desconto",
        "ABATIMENTO": "Abatimento",
        "JUROS": "Juros",
        "MULTA": "Multa",
        GETNET_ADJUSTMENT_COMPONENT: "Getnet",
    }.get((component or "").upper(), "")


def preferred_accounting_entry(existing: LancamentoContabil, candidate: LancamentoContabil) -> LancamentoContabil:
    if rule_component_key(existing.componente) != "PRINCIPAL" or rule_component_key(candidate.componente) != "PRINCIPAL":
        return existing
    if candidate.origem == "comprovante" and existing.origem != "comprovante":
        return candidate
    if candidate.componente == "VALOR_COBRADO" and existing.componente == "PRINCIPAL":
        return candidate
    return existing


def accounting_entry_unique_key(entry: LancamentoContabil) -> str:
    component = rule_component_key(entry.componente)
    if entry.status == "editado_manual" and entry.origem == "manual" and component != "PRINCIPAL":
        return f"manual:{entry.id}"
    return component


def unique_accounting_entries(entries: list[LancamentoContabil]) -> list[LancamentoContabil]:
    selected: dict[str, LancamentoContabil] = {}
    for entry in sorted(entries, key=accounting_entry_order):
        key = accounting_entry_unique_key(entry)
        selected[key] = preferred_accounting_entry(selected[key], entry) if key in selected else entry
    return sorted(selected.values(), key=accounting_entry_order)


def prune_automatic_duplicate_components(entries: list[LancamentoContabil], db: Session) -> list[LancamentoContabil]:
    """Remove stale automatic aliases like PRINCIPAL once VALOR_COBRADO exists."""
    selected: dict[str, LancamentoContabil] = {}
    automatic = [entry for entry in entries if entry.status != "editado_manual"]
    for entry in sorted(automatic, key=accounting_entry_order):
        key = rule_component_key(entry.componente)
        selected[key] = preferred_accounting_entry(selected[key], entry) if key in selected else entry
    keep_ids = {entry.id for entry in selected.values()}
    for entry in automatic:
        if entry.id not in keep_ids:
            db.delete(entry)
    return [entry for entry in entries if entry.status == "editado_manual" or entry.id in keep_ids]


def rule_matches_movement(rule: RegraContabil, movement: MovimentoExtrato, receipt: Comprovante | None = None, component: str = "", rfb: ComprovanteRfb | None = None, rfb_items: list[ComprovanteRfbItem] | None = None) -> bool:
    trigger = rule.favorecido_normalizado
    history = rule_match_history(movement, receipt)
    component_matches = not component or (rule_component_key(rule.tipo_componente) == rule_component_key(component) if rule.tipo_componente else rule_component_key(component) == "PRINCIPAL")
    component_source = " ".join(item for item in [history, rule_component_trigger_label(component), component] if item)
    receipt_source = receipt_trigger_text(receipt, rfb, rfb_items)
    receipt_trigger = normalize_name(rule.gatilho_comprovante_normalizado or "")
    receipt_tokens_match = not receipt_trigger or receipt_trigger in normalize_name(receipt_source) or set(receipt_trigger.split()) <= set(normalize_name(receipt_source).split())
    statement_matches = trigger_matches(component_source, trigger) or legacy_trigger_matches_receipt(rule, history, receipt, rfb, rfb_items)
    exclusion = normalize_name(rule.texto_exclusao_normalizado or "")
    exclusion_matches = bool(exclusion and trigger_matches(component_source, exclusion))
    return bool(statement_matches and not exclusion_matches and receipt_tokens_match and (not rule.tipo_operacao or normalize_rule_accounting_nature(rule.tipo_operacao) == accounting_nature(movement.natureza)) and component_matches)


def in_reconciliation_period(reconciliation: Conciliacao, movement: MovimentoExtrato) -> bool:
    return bool(movement.data and reconciliation.data_inicio <= movement.data <= reconciliation.data_fim)


def cef_alias_matches(left: str, right: str) -> bool:
    left_normalized = normalize_name(left)
    right_normalized = normalize_name(right)
    left_tokens = set(left_normalized.split())
    right_tokens = set(right_normalized.split())
    left_is_cef = "CEF" in left_tokens or {"CAIXA", "ECONOMICA", "FEDERAL"} <= left_tokens
    right_is_cef = "CEF" in right_tokens or {"CAIXA", "ECONOMICA", "FEDERAL"} <= right_tokens
    return left_is_cef and right_is_cef


def receipt_matches_movement(movement_text: str, beneficiary: str) -> bool:
    text_tokens = normalize_name(movement_text).split()
    beneficiary_tokens = normalize_name(beneficiary).split()
    if not text_tokens or not beneficiary_tokens:
        return False
    if cef_alias_matches(movement_text, beneficiary):
        return True
    if names_similar(movement_text, beneficiary, allow_truncated_terminal=True):
        return True
    # A statement history begins with the operation name, while the beneficiary appears later.
    compatible = lambda token, candidate: token.startswith(candidate) or candidate.startswith(token)
    if len(beneficiary_tokens) >= 2 and all(any(compatible(token, candidate) for token in text_tokens) for candidate in beneficiary_tokens):
        return True
    matched_tokens = [candidate for candidate in beneficiary_tokens if any(compatible(token, candidate) for token in text_tokens)]
    strong_matches = [candidate for candidate in matched_tokens if len(candidate) >= 4]
    return len(matched_tokens) >= 2 and bool(strong_matches)


def receipt_match_criterion(movement_text: str, receipt: Comprovante, movement_document: str = "") -> str:
    if receipt.numero_documento and trigger_matches(movement_text, receipt.numero_documento):
        return "número do documento"
    if movement_document and caixa_operation_codes_match_from_document(movement_document, receipt):
        return "código Caixa sem prefixo do dia"
    if movement_document and receipt.numero_documento and document_numbers_overlap(movement_document, receipt.numero_documento):
        return "número do documento"
    for field, label in ((receipt.beneficiario or receipt.favorecido, "beneficiário"), (receipt.beneficiario_final, "beneficiário final"), (receipt.nome_fantasia, "nome fantasia")):
        if field and receipt_matches_movement(movement_text, field):
            return label
    return ""


def movement_document_number(movement: MovimentoExtrato) -> str:
    original = movement.dados_originais if isinstance(movement.dados_originais, dict) else {}
    return str(original.get("numero_documento") or "")


def document_numbers_overlap(left: str, right: str) -> bool:
    left_digits = re.sub(r"\D", "", left or "").lstrip("0")
    right_digits = re.sub(r"\D", "", right or "").lstrip("0")
    if len(left_digits) < 4 or len(right_digits) < 4:
        return False
    return left_digits in right_digits or right_digits in left_digits


def caixa_receipt_operation_code_without_day(receipt: Comprovante) -> str:
    if not receipt.data:
        return ""
    digits = re.sub(r"\D", "", receipt.numero_documento or "")
    day_prefix = f"0{receipt.data.day:02d}"
    if len(digits) > len(day_prefix) and digits.startswith(day_prefix):
        return digits[len(day_prefix):]
    return ""


def caixa_operation_codes_match(movement: MovimentoExtrato, receipt: Comprovante) -> bool:
    return caixa_operation_codes_match_from_document(movement_document_number(movement), receipt)


def caixa_operation_codes_match_from_document(movement_document: str, receipt: Comprovante) -> bool:
    movement_digits = re.sub(r"\D", "", movement_document or "").lstrip("0")
    receipt_without_day = caixa_receipt_operation_code_without_day(receipt).lstrip("0")
    return bool(len(movement_digits) >= 4 and receipt_without_day and movement_digits == receipt_without_day)


def boleto_without_counterparty(movement: MovimentoExtrato) -> bool:
    history_tokens = set(normalize_name(movement.historico).split())
    if "BOLETO" not in history_tokens:
        return False
    if normalize_name(movement.nome_encontrado):
        return False
    generic_tokens = {"PAG", "PG", "PAGTO", "PAGAMENTO", "BOLETO", "BLOQTO", "COBRANCA", "COBRANÇA", "ELETRON", "ELETRONICO"}
    return history_tokens <= generic_tokens


def receipt_operation_matches(movement: MovimentoExtrato, receipt: Comprovante) -> bool:
    history, operation = normalize_name(movement.historico), normalize_name(receipt.tipo_operacao)
    if not operation:
        return True
    if "PIX" in history:
        return "PIX" in operation
    if any(term in history for term in ("TRANSFERENCIA", "TED", "DOC")):
        return any(term in operation for term in ("TRANSFERENCIA", "TED", "DOC"))
    return True


def receipt_time_matches(movement: MovimentoExtrato, receipt: Comprovante) -> bool:
    if not movement.hora or not receipt.hora or movement.hora[:5] == receipt.hora[:5]:
        return True
    if "AGENDAMENTO" not in normalize_name(movement.historico):
        return False
    def minutes(value: str) -> int | None:
        match = re.match(r"^(\d{2}):(\d{2})", value or "")
        return int(match.group(1)) * 60 + int(match.group(2)) if match else None
    movement_minutes = minutes(movement.hora)
    receipt_minutes = minutes(receipt.hora)
    return movement_minutes is not None and receipt_minutes is not None and abs(movement_minutes - receipt_minutes) <= 10


def receipt_tariff_date_matches(movement: MovimentoExtrato, receipt: Comprovante) -> bool:
    occurrence = re.search(r"ocorr[êe]ncia\s+(\d{2}/\d{2}/\d{4})", movement.nome_encontrado, re.I)
    if occurrence:
        return bool(receipt.data and receipt.data.strftime("%d/%m/%Y") == occurrence.group(1))
    return movement.data == receipt.data


def is_transfer_without_counterparty(movement: MovimentoExtrato) -> bool:
    history = normalize_name(movement.historico)
    name = normalize_name(movement.nome_encontrado)
    return bool(any(term in history for term in ("TRANSFERENCIA", "TED", "DOC")) and (not name or re.fullmatch(r"\d{2} \d{2}", name)))


def fallback_boleto_receipt_candidate(movement: MovimentoExtrato, receipts: list[Comprovante], used_receipts: set[str]) -> tuple[Comprovante | None, str]:
    if not boleto_without_counterparty(movement):
        return None, ""
    candidates = [
        item
        for item in receipts
        if item.id not in used_receipts
        and item.valor_pago == movement.valor
        and item.data == movement.data
        and "BOLETO" in normalize_name(item.tipo_operacao)
        and receipt_time_matches(movement, item)
    ]
    movement_document = movement_document_number(movement)
    caixa_document_matches = [item for item in candidates if caixa_operation_codes_match(movement, item)]
    if len(caixa_document_matches) == 1:
        return caixa_document_matches[0], "código Caixa sem prefixo do dia"
    document_matches = [item for item in candidates if document_numbers_overlap(movement_document, item.numero_documento)]
    if len(document_matches) == 1:
        return document_matches[0], "número do documento parcial"
    if len(candidates) == 1:
        return candidates[0], "data, valor e tipo boleto"
    return None, ""


def rule_payload(rule: RegraContabil, movements: list[MovimentoExtrato], receipts: dict[str, Comprovante], covered_entries: list[tuple[str, LancamentoContabil]]) -> dict:
    movement_by_id = {movement.id: movement for movement in movements}
    covered = [(movement_by_id[movement_id], entry) for movement_id, entry in covered_entries if movement_id in movement_by_id]
    return {"id": rule.id, "gatilho": rule.favorecido_normalizado, "gatilho_comprovante": rule.gatilho_comprovante_normalizado, "texto_exclusao": rule.texto_exclusao_normalizado, "natureza": normalize_rule_accounting_nature(rule.tipo_operacao), "tipo_componente": rule.tipo_componente, "conta_debito": rule.conta_debito, "conta_credito": rule.conta_credito, "historico": rule.historico, "complemento": rule.complemento, "escopo": rule.escopo, "banco_origem": rule.banco, "criada_em": rule.created_at.isoformat() if rule.created_at else "", "cobertos": len(covered), "movimentos": [{"data": movement.data.strftime("%d/%m/%Y") if movement.data else "—", "historico": " ".join(part for part in [movement.historico, movement.nome_encontrado] if part), "texto_extrato": " ".join(part for part in [movement.historico, movement.nome_encontrado] if part), "texto_comprovante": receipt_description(receipts.get(movement.id)), "tem_comprovante": bool(receipts.get(movement.id)), "valor": str(entry.valor or 0), "tipo_componente": entry.componente, "natureza": normalize_statement_nature(movement.natureza), "natureza_contabil": accounting_nature(movement.natureza)} for movement, entry in covered]}


def scoped_rules(reconciliation: Conciliacao, db: Session) -> list[RegraContabil]:
    rules = db.query(RegraContabil).filter_by(cliente_id=reconciliation.cliente_id, tipo_fonte="extrato", ativo=True).order_by(RegraContabil.created_at.desc()).all()
    return [rule for rule in rules if rule.escopo == "global" or rule.conciliacao_id == reconciliation.id]


def scoped_reconciliations(reconciliation: Conciliacao, db: Session) -> list[Conciliacao]:
    return db.query(Conciliacao).filter_by(cliente_id=reconciliation.cliente_id).all()


def current_bank_rules(reconciliation: Conciliacao, db: Session) -> list[RegraContabil]:
    ignored_rule_ids = {item.regra_contabil_id for item in db.query(RegraContabilExcecao).filter_by(conciliacao_id=reconciliation.id)}
    return [rule for rule in scoped_rules(reconciliation, db) if rule.banco == reconciliation.banco and rule.id not in ignored_rule_ids]


def rule_period_exception(rule_id: str, reconciliation: Conciliacao, db: Session) -> RegraContabilExcecao | None:
    return db.query(RegraContabilExcecao).filter_by(regra_contabil_id=rule_id, conciliacao_id=reconciliation.id).first()


def complete_accounting_entry(entry: LancamentoContabil) -> bool:
    return bool(entry.status in {"aplicado_por_regra", "editado_manual"} and entry.valor and entry.valor > 0 and entry.conta_debito.strip() and entry.conta_credito.strip() and entry.historico.strip())


def count_rule_entries_in_reconciliation(rule_id: str, reconciliation: Conciliacao, db: Session) -> int:
    return (
        db.query(LancamentoContabil)
        .join(Correspondencia, LancamentoContabil.correspondencia_id == Correspondencia.id)
        .filter(
            Correspondencia.conciliacao_id == reconciliation.id,
            LancamentoContabil.regra_contabil_id == rule_id,
            LancamentoContabil.status == "aplicado_por_regra",
        )
        .count()
    )


DISCOUNT_COMPONENTS = {"DESCONTO", "ABATIMENTO", "DESCONTO_ABATIMENTO"}


def is_discount_component(component: str | None) -> bool:
    return (component or "") in DISCOUNT_COMPONENTS


def is_balanced_other_component(component: str | None) -> bool:
    return is_discount_component(component) or component == GETNET_ADJUSTMENT_COMPONENT


def is_other_accounting_entry(entry: LancamentoContabil) -> bool:
    return entry.efeito_no_total == "OUTROS" or is_balanced_other_component(entry.componente)


def accounting_entry_order(entry: LancamentoContabil) -> tuple[int, int, str]:
    component_order = {"PRINCIPAL": 1, "VALOR_COBRADO": 1, "MULTA": 2, "JUROS": 3, "ENCARGOS": 4, "DESCONTO": 5, "ABATIMENTO": 6, "DESCONTO_ABATIMENTO": 7}
    fallback = component_order.get(entry.componente, 99)
    return (entry.ordem or fallback, fallback, entry.id)


def accounting_export_order(entry: LancamentoContabil) -> tuple[int, int, str]:
    component_order = {"PRINCIPAL": 1, "VALOR_COBRADO": 1, "MULTA": 2, "JUROS": 3, "ENCARGOS": 4, "DESCONTO": 5, "ABATIMENTO": 6, "DESCONTO_ABATIMENTO": 7}
    fallback = component_order.get(entry.componente, 99)
    return (fallback, entry.ordem or fallback, entry.id)


def statement_effect_value(value: Decimal, effect: str) -> Decimal:
    return value if effect == "SOMA" else -value if effect == "SUBTRAI" else Decimal("0.00")


def active_accounting_entry(entry: LancamentoContabil, active_rule_ids: set[str]) -> bool:
    return entry.status == "editado_manual" or entry.regra_contabil_id in active_rule_ids


def movement_used_in_period(movement: MovimentoExtrato | None) -> bool:
    return bool(movement and not getattr(movement, "ignorado_no_periodo", False))


def movement_eligible_for_accounting(movement: MovimentoExtrato | None) -> bool:
    return bool(movement and movement_used_in_period(movement) and (movement.ativo or is_getnet_adjustment_movement(movement)))


def accounting_rule_movements(reconciliation: Conciliacao, db: Session) -> list[MovimentoExtrato]:
    movements = [
        item
        for item in db.query(MovimentoExtrato).filter_by(conciliacao_id=reconciliation.id).all()
        if movement_eligible_for_accounting(item) and in_reconciliation_period(reconciliation, item)
    ]
    movements.sort(key=movement_statement_order)
    return movements


def accounting_integrity(reconciliation: Conciliacao, db: Session) -> dict:
    matches = db.query(Correspondencia).filter_by(conciliacao_id=reconciliation.id).all()
    active_rule_ids = {rule.id for rule in current_bank_rules(reconciliation, db)}
    entries_by_match = {match.id: [] for match in matches}
    if matches:
        for entry in db.query(LancamentoContabil).filter(LancamentoContabil.correspondencia_id.in_(entries_by_match)).all():
            entries_by_match[entry.correspondencia_id].append(entry)
    movements = {movement.id: movement for movement in db.query(MovimentoExtrato).filter(MovimentoExtrato.id.in_([match.movimento_extrato_id for match in matches])).all()} if matches else {}
    valid_entries = []
    bank_debit = Decimal("0.00")
    bank_credit = Decimal("0.00")
    other_debit = Decimal("0.00")
    other_credit = Decimal("0.00")
    incomplete = []
    for match in matches:
        entries = entries_by_match[match.id]
        active_entries = [entry for entry in entries if active_accounting_entry(entry, active_rule_ids)]
        completed = unique_accounting_entries([entry for entry in active_entries if complete_accounting_entry(entry)])
        invalid_applied = [entry for entry in entries if active_accounting_entry(entry, active_rule_ids) and entry.status in {"aplicado_por_regra", "editado_manual"} and not complete_accounting_entry(entry)]
        movement = movements.get(match.movimento_extrato_id)
        if not movement_eligible_for_accounting(movement):
            continue
        if invalid_applied or (match.regra_contabil_id in active_rule_ids and not any(entry.regra_contabil_id == match.regra_contabil_id for entry in completed)):
            incomplete.append({"movimento_id": match.movimento_extrato_id, "data": movement.data.strftime("%d/%m/%Y") if movement and movement.data else "—", "historico": " ".join(part for part in [movement.historico, movement.nome_encontrado] if part) if movement else "Movimento não encontrado"})
        valid_entries.extend(completed)
        for entry in completed:
            if is_other_accounting_entry(entry):
                if is_balanced_other_component(entry.componente):
                    other_debit += entry.valor
                    other_credit += entry.valor
                elif accounting_nature(movement.natureza) == "Débito":
                    other_debit += entry.valor
                else:
                    other_credit += entry.valor
            elif accounting_nature(movement.natureza) == "Débito":
                bank_debit += entry.valor
            else:
                bank_credit += entry.valor
    other_entries = [entry for entry in valid_entries if is_other_accounting_entry(entry)]
    accounting_debit = sum((entry.valor for entry in valid_entries), Decimal("0.00"))
    accounting_credit = sum((entry.valor for entry in valid_entries), Decimal("0.00"))
    difference = accounting_debit - accounting_credit
    other = other_debit - other_credit
    return {"debito": bank_debit, "credito": bank_credit, "outros": other, "outros_debito": other_debit, "outros_credito": other_credit, "diferenca": difference, "movimentos_incompletos": incomplete, "csv_permitido": not incomplete and abs(difference) <= Decimal("0.01"), "lancamentos_validos": valid_entries, "lancamentos_outros": other_entries}


def apply_accounting_rules(reconciliation: Conciliacao, db: Session) -> int:
    """Create or update accounting entries without changing document matching links."""
    rules = current_bank_rules(reconciliation, db)
    applied = 0
    movements = accounting_rule_movements(reconciliation, db)
    correspondences = db.query(Correspondencia).filter_by(conciliacao_id=reconciliation.id).all()
    match_by_movement = {item.movimento_extrato_id: item for item in correspondences}
    match_ids = [item.id for item in correspondences]
    receipt_ids = [item.comprovante_id for item in correspondences if item.comprovante_id]
    rfb_ids = [item.comprovante_rfb_id for item in correspondences if item.comprovante_rfb_id]
    receipts = {item.id: item for item in db.query(Comprovante).filter(Comprovante.id.in_(receipt_ids)).all()} if receipt_ids else {}
    rfb_receipts = {item.id: item for item in db.query(ComprovanteRfb).filter(ComprovanteRfb.id.in_(rfb_ids)).all()} if rfb_ids else {}
    rfb_items_by_receipt: dict[str, list[ComprovanteRfbItem]] = defaultdict(list)
    if rfb_ids:
        for item in db.query(ComprovanteRfbItem).filter(ComprovanteRfbItem.comprovante_rfb_id.in_(rfb_ids)).all():
            rfb_items_by_receipt[item.comprovante_rfb_id].append(item)
    entries_by_match: dict[str, list[LancamentoContabil]] = defaultdict(list)
    if match_ids:
        for entry in db.query(LancamentoContabil).filter(LancamentoContabil.correspondencia_id.in_(match_ids)).order_by(LancamentoContabil.ordem, LancamentoContabil.id).all():
            entries_by_match[entry.correspondencia_id].append(entry)
    for movement in movements:
        match = match_by_movement.get(movement.id)
        receipt = receipts.get(match.comprovante_id) if match and match.comprovante_id else None
        rfb = rfb_receipts.get(match.comprovante_rfb_id) if match and match.comprovante_rfb_id else None
        rfb_items = rfb_items_by_receipt.get(rfb.id, []) if rfb else []
        if not match:
            match = Correspondencia(conciliacao_id=reconciliation.id, movimento_extrato_id=movement.id)
            db.add(match)
            db.flush()
            match_by_movement[movement.id] = match
        entries = prune_automatic_duplicate_components(entries_by_match.get(match.id, []), db)
        entries = unique_accounting_entries(entries)
        # A document can have distinct rules for principal, discount, and taxes.
        # Manual entries are deliberately never overwritten by a rule refresh.
        source_entries = unique_accounting_entries([entry for entry in entries if entry.status != "editado_manual"])
        if not source_entries:
            source_entries = [LancamentoContabil(correspondencia_id=match.id, componente="PRINCIPAL", categoria="PRINCIPAL", valor=movement.valor or 0, origem="extrato", ordem=1)]
            db.add(source_entries[0])
        applied_rule = None
        for entry in source_entries:
            rule = next((item for item in rules if rule_matches_movement(item, movement, receipt, entry.componente, rfb, rfb_items)), None)
            if not rule:
                continue
            entry.regra_contabil_id = rule.id
            entry.conta_debito = rule.conta_debito
            entry.conta_credito = rule.conta_credito
            entry.historico = rule.historico
            entry.status = "aplicado_por_regra"
            applied_rule = rule
            applied += 1
        match.regra_contabil_id = applied_rule.id if applied_rule else None
    return applied


def reconciliation_or_404(conciliacao_id: str, db: Session) -> Conciliacao:
    reconciliation = db.get(Conciliacao, conciliacao_id)
    if not reconciliation:
        raise HTTPException(404, "Conciliação não encontrada")
    return reconciliation


def movement_statement_order(item: MovimentoExtrato) -> tuple:
    original = item.dados_originais if isinstance(item.dados_originais, dict) else {}
    raw_order = original.get("ordem_extrato")
    try:
        statement_order = int(raw_order)
    except (TypeError, ValueError):
        statement_order = 10**9
    return (item.data or date.max, item.hora or "", item.pagina_numero or 0, statement_order, item.id)


@router.get("/conciliacoes/{conciliacao_id}/conta-bancaria")
def bank_account(conciliacao_id: str, db: Session = Depends(get_db)):
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    account = db.query(ContaBancaria).filter_by(cliente_id=reconciliation.cliente_id, banco=reconciliation.banco).first()
    return {"conta_contabil": account.conta_contabil if account else ""}


@router.put("/conciliacoes/{conciliacao_id}/conta-bancaria")
def save_bank_account(conciliacao_id: str, payload: ContaBancariaInput, db: Session = Depends(get_db)):
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    account = db.query(ContaBancaria).filter_by(cliente_id=reconciliation.cliente_id, banco=reconciliation.banco).first()
    if not account:
        account = ContaBancaria(cliente_id=reconciliation.cliente_id, banco=reconciliation.banco)
        db.add(account)
    account.conta_contabil = payload.conta_contabil.strip()
    db.commit()
    return {"conta_contabil": account.conta_contabil}


SOURCE_RULES = {"maquininha", "nota"}


def source_rule_bank(reconciliation: Conciliacao, source: str) -> str:
    return "Notas" if source == "nota" else reconciliation.banco


def source_rule_text(row) -> str:
    return " ".join(str(value or "") for value in [row.texto, row.documento, getattr(row, "forma_pagamento", ""), getattr(row, "servico", "")]).strip()


def independent_source_rows(reconciliation: Conciliacao, source: str, db: Session) -> list[SimpleNamespace]:
    if source == "maquininha":
        files = {item.id: item for item in db.query(Arquivo).filter_by(conciliacao_id=reconciliation.id, ativo=True).all() if item.tipo_documento in MACHINE_STATEMENT_DOCUMENT_TYPES}
        records = [item for item in db.query(Comprovante).filter_by(conciliacao_id=reconciliation.id, ativo=True).all() if item.arquivo_id in files]
        rows = [
            SimpleNamespace(
                id=item.id,
                data=item.data,
                texto=item.beneficiario or item.favorecido,
                documento=item.numero_documento,
                valor=item.valor_original or item.valor_pago,
                arquivo_id=item.arquivo_id,
                pagina=item.pagina_numero,
                fonte="maquininha",
            )
            for item in records
            if item.data and reconciliation.data_inicio <= item.data <= reconciliation.data_fim and (item.valor_original or item.valor_pago)
        ]
    elif source == "nota":
        records = db.query(NotaFiscal).filter_by(conciliacao_id=reconciliation.id, ativo=True).all()
        rows = [
            SimpleNamespace(
                id=item.id,
                data=item.data_emissao,
                texto=item.fornecedor,
                documento=item.numero_nota,
                valor=item.valor_total,
                arquivo_id=item.arquivo_id,
                pagina=item.pagina_numero,
                forma_pagamento=(item.dados_originais or {}).get("forma_pagamento", "") if isinstance(item.dados_originais, dict) else "",
                servico=(item.dados_originais or {}).get("servico", "") if isinstance(item.dados_originais, dict) else "",
                fonte="nota",
            )
            for item in records
            if item.valor_total and (not item.data_emissao or reconciliation.data_inicio <= item.data_emissao <= reconciliation.data_fim)
        ]
    else:
        raise HTTPException(404, "Fonte de regras não encontrada")
    return sorted(rows, key=lambda item: (item.data or date.max, item.texto or "", item.id))


def source_rules_for(reconciliation: Conciliacao, source: str, db: Session) -> list[RegraContabil]:
    bank = source_rule_bank(reconciliation, source)
    return [
        rule
        for rule in db.query(RegraContabil).filter_by(cliente_id=reconciliation.cliente_id, banco=bank, tipo_fonte=source, ativo=True).order_by(RegraContabil.created_at.desc()).all()
        if rule.escopo == "global" or rule.conciliacao_id == reconciliation.id
    ]


def source_rule_matches(rule: RegraContabil, row) -> bool:
    return trigger_matches(source_rule_text(row), rule.favorecido_normalizado)


def source_rule_payload(rule: RegraContabil, rows: list[SimpleNamespace]) -> dict:
    covered = [row for row in rows if source_rule_matches(rule, row)]
    return {
        "id": rule.id,
        "gatilho": rule.favorecido_normalizado,
        "conta_debito": rule.conta_debito,
        "conta_credito": rule.conta_credito,
        "historico": rule.historico,
        "complemento": rule.complemento,
        "escopo": rule.escopo,
        "criada_em": rule.created_at.isoformat() if rule.created_at else "",
        "cobertos": len(covered),
    }


def source_rule_match_payload(row, rule: RegraContabil | None = None) -> dict:
    return {
        "id": row.id,
        "data": row.data.strftime("%d/%m/%Y") if row.data else "—",
        "texto": row.texto or "—",
        "documento": row.documento or "—",
        "forma_pagamento": getattr(row, "forma_pagamento", "") or "—",
        "valor": str(row.valor or 0),
        "arquivo_id": row.arquivo_id,
        "pagina": row.pagina,
        "regra_id": rule.id if rule else "",
        "conta_debito": rule.conta_debito if rule else "",
        "conta_credito": rule.conta_credito if rule else "",
        "historico_contabil": rule.historico if rule else "",
        "complemento": rule.complemento if rule else "",
    }


@router.get("/conciliacoes/{conciliacao_id}/regras-fonte/{fonte}")
def source_accounting_rules(conciliacao_id: str, fonte: str, db: Session = Depends(get_db)):
    if fonte not in SOURCE_RULES:
        raise HTTPException(404, "Fonte de regras não encontrada")
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    rows = independent_source_rows(reconciliation, fonte, db)
    rules = source_rules_for(reconciliation, fonte, db)
    matched_rule_by_row = {}
    for row in rows:
        matched_rule_by_row[row.id] = next((rule for rule in rules if source_rule_matches(rule, row)), None)
    pending = [source_rule_match_payload(row) for row in rows if not matched_rule_by_row[row.id]]
    classified = [source_rule_match_payload(row, matched_rule_by_row[row.id]) for row in rows if matched_rule_by_row[row.id]]
    return {
        "pendentes": pending,
        "classificados": classified,
        "salvas": [source_rule_payload(rule, rows) for rule in rules],
        "resumo": {"total": len(rows), "classificados": len(classified), "pendentes": len(pending)},
    }


@router.post("/conciliacoes/{conciliacao_id}/regras-fonte/{fonte}/previa")
def preview_source_accounting_rule(conciliacao_id: str, fonte: str, payload: RegraFonteInput, db: Session = Depends(get_db)):
    if fonte not in SOURCE_RULES:
        raise HTTPException(404, "Fonte de regras não encontrada")
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    rule = RegraContabil(cliente_id=reconciliation.cliente_id, conciliacao_id=reconciliation.id, banco=source_rule_bank(reconciliation, fonte), tipo_fonte=fonte, tipo_operacao="Débito", tipo_componente="PRINCIPAL", favorecido_normalizado=normalize_name(payload.gatilho))
    matches = [source_rule_match_payload(row) for row in independent_source_rows(reconciliation, fonte, db) if source_rule_matches(rule, row)]
    return {"quantidade": len(matches), "lancamentos": matches, "motivo": "Nenhum registro corresponde ao gatilho informado." if not matches else ""}


@router.post("/conciliacoes/{conciliacao_id}/regras-fonte/{fonte}")
def create_source_accounting_rule(conciliacao_id: str, fonte: str, payload: RegraFonteInput, db: Session = Depends(get_db)):
    if fonte not in SOURCE_RULES:
        raise HTTPException(404, "Fonte de regras não encontrada")
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    if not all([payload.gatilho.strip(), payload.conta_debito.strip(), payload.conta_credito.strip(), payload.historico.strip()]):
        raise HTTPException(422, "Preencha gatilho, débito, crédito e histórico")
    rows = independent_source_rows(reconciliation, fonte, db)
    bank = source_rule_bank(reconciliation, fonte)
    normalized_trigger = normalize_name(payload.gatilho)
    if any(rule.favorecido_normalizado == normalized_trigger for rule in source_rules_for(reconciliation, fonte, db)):
        raise HTTPException(409, "Já existe uma regra equivalente nesta aba")
    rule = RegraContabil(
        cliente_id=reconciliation.cliente_id,
        conciliacao_id=reconciliation.id,
        banco=bank,
        tipo_fonte=fonte,
        tipo_operacao="Débito",
        tipo_componente="PRINCIPAL",
        favorecido_normalizado=normalized_trigger,
        conta_debito=payload.conta_debito.strip(),
        conta_credito=payload.conta_credito.strip(),
        historico=payload.historico.strip(),
        complemento=payload.complemento.strip(),
        escopo="global" if payload.escopo == "global" else "periodo",
    )
    if not any(source_rule_matches(rule, row) for row in rows):
        raise HTTPException(422, "O gatilho não cobre nenhum registro desta aba")
    db.add(rule); db.commit(); db.refresh(rule)
    return {"id": rule.id, "regras": source_accounting_rules(conciliacao_id, fonte, db)}


@router.delete("/conciliacoes/{conciliacao_id}/regras-fonte/{fonte}/{regra_id}")
def delete_source_accounting_rule(conciliacao_id: str, fonte: str, regra_id: str, db: Session = Depends(get_db)):
    if fonte not in SOURCE_RULES:
        raise HTTPException(404, "Fonte de regras não encontrada")
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    rule = db.get(RegraContabil, regra_id)
    if not rule or rule.cliente_id != reconciliation.cliente_id or rule.banco != source_rule_bank(reconciliation, fonte) or rule.tipo_fonte != fonte:
        raise HTTPException(404, "Regra não encontrada")
    rule.ativo = False
    db.commit()
    return {"message": "Regra excluída.", "regras": source_accounting_rules(conciliacao_id, fonte, db)}


@router.get("/conciliacoes/{conciliacao_id}/regras-fonte/{fonte}/csv")
def source_accounting_csv(conciliacao_id: str, fonte: str, db: Session = Depends(get_db)):
    if fonte not in SOURCE_RULES:
        raise HTTPException(404, "Fonte de regras não encontrada")
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    rows = independent_source_rows(reconciliation, fonte, db)
    rules = source_rules_for(reconciliation, fonte, db)
    output = io.StringIO(newline="")
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow(["Data", "Debito", "Credito", "Historico", "Valor", "Complemento"])
    for row in rows:
        rule = next((item for item in rules if source_rule_matches(item, row)), None)
        if not rule:
            continue
        writer.writerow([
            row.data.strftime("%d/%m/%Y") if row.data else "",
            accounting_code(rule.conta_debito),
            accounting_code(rule.conta_credito),
            accounting_code(rule.historico),
            f"{Decimal(row.valor or 0):.2f}",
            rule.complemento,
        ])
    period = reconciliation.data_inicio.strftime("%m%y")
    filename = f"{period}_{fonte}.csv"
    return Response(output.getvalue().encode("utf-8-sig"), media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/conciliacoes/{conciliacao_id}/regras-contabeis")
def accounting_rules(conciliacao_id: str, db: Session = Depends(get_db), response: FastAPIResponse = None, auto_hide_zero_covered: bool = True):
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    movements = [item for item in db.query(MovimentoExtrato).filter_by(conciliacao_id=conciliacao_id, ativo=True) if movement_used_in_period(item) and in_reconciliation_period(reconciliation, item)]
    if reconciliation.banco == "Santander" and any(is_santander_getnet_credit(item) for item in movements):
        sync_getnet_anticipation_adjustments(reconciliation, db)
        movements = accounting_rule_movements(reconciliation, db)
    else:
        movements.sort(key=movement_statement_order)
    movement_by_id = {item.id: item for item in movements}
    if reconciliation.banco == "Banco do Brasil":
        for movement in movements:
            logger.info("Total do extrato BB: data=%s historico=%s valor=%s natureza=%s", movement.data, movement.historico, movement.valor, movement.natureza)
    same_bank_rules = current_bank_rules(reconciliation, db)
    matches = {item.movimento_extrato_id: item for item in db.query(Correspondencia).filter_by(conciliacao_id=conciliacao_id).all()}
    match_ids = [item.id for item in matches.values()]
    receipt_ids = [item.comprovante_id for item in matches.values() if item.comprovante_id]
    rfb_ids = [item.comprovante_rfb_id for item in matches.values() if item.comprovante_rfb_id]
    receipts_by_id = {item.id: item for item in db.query(Comprovante).filter(Comprovante.id.in_(receipt_ids)).all()} if receipt_ids else {}
    rfb_by_id = {item.id: item for item in db.query(ComprovanteRfb).filter(ComprovanteRfb.id.in_(rfb_ids)).all()} if rfb_ids else {}
    rfb_items_by_receipt: dict[str, list[ComprovanteRfbItem]] = defaultdict(list)
    if rfb_ids:
        for item in db.query(ComprovanteRfbItem).filter(ComprovanteRfbItem.comprovante_rfb_id.in_(rfb_ids)).all():
            rfb_items_by_receipt[item.comprovante_rfb_id].append(item)
    entries_by_match: dict[str, list[LancamentoContabil]] = defaultdict(list)
    if match_ids:
        for entry in db.query(LancamentoContabil).filter(LancamentoContabil.correspondencia_id.in_(match_ids)).all():
            entries_by_match[entry.correspondencia_id].append(entry)
    receipts = {movement_id: receipts_by_id[match.comprovante_id] for movement_id, match in matches.items() if match.comprovante_id in receipts_by_id}
    entries_by_movement = {movement_id: unique_accounting_entries(entries_by_match.get(match.id, [])) for movement_id, match in matches.items()}
    tariff_receipt_ids = {
        match.comprovante_id
        for match in matches.values()
        if match.comprovante_id and (movement := movement_by_id.get(match.movimento_extrato_id)) and "TARIFA PIX" in normalize_name(movement.historico)
    }
    covered_by_rule = {rule.id: [] for rule in same_bank_rules}
    covered_components = set()
    for movement in movements:
        movement_id = movement.id
        entries = entries_by_movement.get(movement_id, [])
        for entry in entries:
            if complete_accounting_entry(entry) and entry.regra_contabil_id in covered_by_rule:
                covered_by_rule[entry.regra_contabil_id].append((movement_id, entry))
                covered_components.add((movement_id, rule_component_key(entry.componente)))
    integrity = accounting_integrity(reconciliation, db)
    def pending_payload(item: MovimentoExtrato, component: str = "PRINCIPAL", value: Decimal | None = None, document_components: list[str] | None = None, covered_document_components: list[dict] | None = None):
        match = matches.get(item.id)
        receipt = receipts_by_id.get(match.comprovante_id) if match and match.comprovante_id else None
        rfb = rfb_by_id.get(match.comprovante_rfb_id) if match and match.comprovante_rfb_id else None
        rfb_items = rfb_items_by_receipt.get(rfb.id, []) if rfb else []
        is_getnet_adjustment = is_getnet_adjustment_movement(item)
        original = item.dados_originais if isinstance(item.dados_originais, dict) else {}
        history = rule_match_history(item, receipt)
        tariff_in_statement = bool(receipt and receipt.valor_tarifa and "TARIFA PIX" not in normalize_name(item.historico) and receipt.id in tariff_receipt_ids)
        tariff_reference = bool(receipt and "TARIFA PIX" in normalize_name(item.historico))
        bank_receipt_words = list(dict.fromkeys(([receipt.numero_documento] if receipt and receipt.numero_documento else []) + normalize_name(receipt_trigger_text(receipt)).split()))
        rfb_words = list(dict.fromkeys(normalize_name(receipt_trigger_text(None, rfb, rfb_items)).split()))
        receipt_words = list(dict.fromkeys([*bank_receipt_words, *rfb_words]))
        composition = ""
        if is_getnet_adjustment:
            receipt_words = ["JUROS", "ANTECIPACOES", "GETNET", "SANTANDER"]
            composition = "\n".join([
                "Ajuste Getnet/Santander",
                f"Getnet líquido: R$ {Decimal(original.get('total_getnet') or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                f"Santander recebido: R$ {Decimal(original.get('total_santander') or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                f"Diferença a lançar: R$ {Decimal(original.get('diferenca') or item.valor or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
            ])
        if rfb and (rfb.tipo.upper() == "DAS" or any("SIMPLES NACIONAL" in tax.descricao.upper() for tax in rfb_items)):
            taxes = rfb_items
            money = lambda amount: f"R$ {amount:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            lines = [f"{tax.descricao.split(' - ')[0]}: {money(tax.valor_principal or 0)}" for tax in taxes if tax.valor_principal]
            if rfb.valor_multa: lines.append(f"Multa: {money(rfb.valor_multa)}")
            if rfb.valor_juros: lines.append(f"Juros: {money(rfb.valor_juros)}")
            composition = "Composição:\n- " + "\n- ".join(lines) + f"\nTotal do documento: {money(rfb.valor_total or 0)}"
        receipt_type = "emprestimo" if is_loan_receipt(receipt) else "comprovante" if receipt else ""
        if receipt_type == "emprestimo":
            money = lambda amount: f"R$ {amount:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            lines = []
            if receipt.valor_original:
                lines.append(f"Principal/Capital: {money(receipt.valor_original)}")
            if receipt.valor_juros:
                lines.append(f"Juros: {money(receipt.valor_juros)}")
            if receipt.valor_encargos:
                lines.append(f"Encargos: {money(receipt.valor_encargos)}")
            composition = "Comprovante de empréstimo/financiamento\n- " + "\n- ".join(lines) + f"\nTotal do extrato: {money(item.valor or 0)}"
        document_components = document_components or [component]
        return {"id": f"{item.id}:{component}", "movimento_id": item.id, "usado_no_periodo": movement_used_in_period(item), "data": item.data.strftime("%d/%m/%y") if item.data else "—", "historico": history, "valor": str(value if value is not None else item.valor or 0), "natureza": normalize_statement_nature(item.natureza), "natureza_contabil": accounting_nature(item.natureza), "tipo_componente": component, "movimento_composto": len(document_components) > 1, "componentes_documento": document_components, "componentes_cobertos": covered_document_components or [], "palavras_comprovante": receipt_words, "palavras_comprovante_banco": bank_receipt_words, "palavras_comprovante_rfb": rfb_words, "valor_documento": str(receipt.valor_original) if receipt and receipt.valor_original else "", "composicao_simples": composition, "tarifa_no_extrato": tariff_in_statement, "tarifa_referente_ao_comprovante": tariff_reference, "tarifa_referencia_nome": (receipt.beneficiario or receipt.favorecido) if tariff_reference else "", "tarifa_referencia_valor": str(receipt.valor_pago or 0) if tariff_reference else "", "tarifa_referencia_data": receipt.data.strftime("%d/%m/%Y") if tariff_reference and receipt.data else "", "pagina": item.pagina_numero, "arquivo_id": item.arquivo_id, "comprovante_tipo": receipt_type, "comprovante_arquivo_id": receipt.arquivo_id if receipt else None, "comprovante_pagina": receipt.pagina_numero if receipt else None, "comprovante_rfb_arquivo_id": rfb.arquivo_id if rfb else None, "comprovante_rfb_pagina": rfb.pagina_numero if rfb else None, "comprovante_confere": bool((receipt or rfb) and match and match.status.startswith("Conciliado")), "ajuste_getnet": is_getnet_adjustment, "gatilho_sugerido": "JUROS ANTECIPACOES GETNET" if is_getnet_adjustment else "", "complemento_sugerido": "DIFERENÇA ENTRE GETNET E RECEBIMENTOS NO SANTANDER" if is_getnet_adjustment else ""}
    pending = []
    for movement in movements:
        match = matches.get(movement.id)
        items = entries_by_movement.get(movement.id, [])
        candidates = [(entry.componente, entry.valor) for entry in items if entry.valor and entry.valor > 0] or [("PRINCIPAL", movement.valor)]
        document_components = list(dict.fromkeys(component for component, _ in candidates))
        covered_document_components = [{"componente": entry.componente, "valor": str(entry.valor)} for entry in items if complete_accounting_entry(entry) and entry.regra_contabil_id in covered_by_rule]
        pending.extend(
            pending_payload(movement, component, value, document_components, covered_document_components)
            for component, value in candidates
            if (movement.id, rule_component_key(component)) not in covered_components
        )
    if response:
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    saved_payloads = [rule_payload(rule, movements, receipts, covered_by_rule[rule.id]) for rule in same_bank_rules]
    initial_balance = statement_initial_balance(reconciliation, db)
    def saved_rule_order(item: dict):
        entry_order = min((accounting_entry_order(entry) for _, entry in covered_by_rule[item["id"]]), default=(10**9, 10**9, ""))
        return ((item.get("criada_em") or "")[:19], -entry_order[0], -entry_order[1])
    saved_payloads.sort(key=saved_rule_order, reverse=True)
    if auto_hide_zero_covered:
        zero_rule_ids = [item["id"] for item in saved_payloads if item["cobertos"] == 0]
        if zero_rule_ids:
            for rule_id in zero_rule_ids:
                if not db.query(RegraContabilExcecao).filter_by(regra_contabil_id=rule_id, conciliacao_id=reconciliation.id).first():
                    db.add(RegraContabilExcecao(regra_contabil_id=rule_id, conciliacao_id=reconciliation.id))
            release_rule_entries(zero_rule_ids, db, reconciliation.id)
            db.commit()
            saved_payloads = [item for item in saved_payloads if item["id"] not in zero_rule_ids]
    exceptions = db.query(RegraContabilExcecao).filter_by(conciliacao_id=reconciliation.id).all()
    exception_rule_ids = [item.regra_contabil_id for item in exceptions]
    rules_by_id = {item.id: item for item in db.query(RegraContabil).filter(RegraContabil.id.in_(exception_rule_ids)).all()} if exception_rule_ids else {}
    ignored = []
    for exception in exceptions:
        rule = rules_by_id.get(exception.regra_contabil_id)
        if rule and rule.ativo and rule.cliente_id == reconciliation.cliente_id and rule.banco == reconciliation.banco:
            ignored.append({"id": rule.id, "gatilho": rule.favorecido_normalizado, "gatilho_comprovante": rule.gatilho_comprovante_normalizado, "texto_exclusao": rule.texto_exclusao_normalizado, "tipo_componente": rule.tipo_componente, "historico": rule.historico})
    return {
        "pendentes": pending,
        "salvas": saved_payloads,
        "ignoradas": ignored,
        "resumo": {
            "extrato": {
                "saldo_anterior": str(initial_balance),
                "debito": str(sum((item.valor or 0 for item in movements if item.ativo and is_statement_debit(item.natureza)), 0)),
                "credito": str(sum((item.valor or 0 for item in movements if item.ativo and not is_statement_debit(item.natureza)), 0)),
                "outros": "0.00",
                "outros_debito": "0.00",
                "outros_credito": "0.00",
            },
            "razao": {"debito": str(integrity["debito"]), "credito": str(integrity["credito"]), "outros": str(integrity["outros"]), "outros_debito": str(integrity["outros_debito"]), "outros_credito": str(integrity["outros_credito"])},
        },
        "integridade": {"csv_permitido": integrity["csv_permitido"], "diferenca": str(integrity["diferenca"]), "movimentos_incompletos": integrity["movimentos_incompletos"]},
    }


def rule_preview(rule: RegraContabil, reconciliation: Conciliacao, db: Session, current_rule_id: str | None = None) -> list[dict]:
    matches = []
    movements = accounting_rule_movements(reconciliation, db)
    correspondences = db.query(Correspondencia).filter_by(conciliacao_id=reconciliation.id).all()
    match_by_movement = {item.movimento_extrato_id: item for item in correspondences}
    match_ids = [item.id for item in correspondences]
    receipt_ids = [item.comprovante_id for item in correspondences if item.comprovante_id]
    rfb_ids = [item.comprovante_rfb_id for item in correspondences if item.comprovante_rfb_id]
    receipts = {item.id: item for item in db.query(Comprovante).filter(Comprovante.id.in_(receipt_ids)).all()} if receipt_ids else {}
    rfb_receipts = {item.id: item for item in db.query(ComprovanteRfb).filter(ComprovanteRfb.id.in_(rfb_ids)).all()} if rfb_ids else {}
    entries_by_match: dict[str, list[LancamentoContabil]] = {item_id: [] for item_id in match_ids}
    if match_ids:
        for entry in db.query(LancamentoContabil).filter(LancamentoContabil.correspondencia_id.in_(match_ids)).all():
            entries_by_match[entry.correspondencia_id].append(entry)
    has_complete_entries = any(complete_accounting_entry(entry) for entries in entries_by_match.values() for entry in entries)
    active_rule_ids = {item.id for item in current_bank_rules(reconciliation, db)} if has_complete_entries else set()
    rfb_items_by_receipt: dict[str, list[ComprovanteRfbItem]] = {item_id: [] for item_id in rfb_ids}
    if rfb_ids:
        for item in db.query(ComprovanteRfbItem).filter(ComprovanteRfbItem.comprovante_rfb_id.in_(rfb_ids)).all():
            rfb_items_by_receipt[item.comprovante_rfb_id].append(item)
    for movement in movements:
        match = match_by_movement.get(movement.id)
        receipt = receipts.get(match.comprovante_id) if match and match.comprovante_id else None
        rfb = rfb_receipts.get(match.comprovante_rfb_id) if match and match.comprovante_rfb_id else None
        rfb_items = rfb_items_by_receipt.get(rfb.id, []) if rfb else []
        entries = entries_by_match.get(match.id, []) if match else []
        candidates = entries or [SimpleNamespace(componente="PRINCIPAL", status="pendente", valor=Decimal("0"), conta_debito="", conta_credito="", historico="")]
        for entry in candidates:
            if complete_accounting_entry(entry):
                if entry.status == "editado_manual":
                    continue
                if entry.regra_contabil_id and entry.regra_contabil_id != current_rule_id and entry.regra_contabil_id in active_rule_ids:
                    continue
            if rule_matches_movement(rule, movement, receipt, entry.componente, rfb, rfb_items):
                original_history = " ".join(item for item in [movement.historico, movement.nome_encontrado] if item)
                source = "Beneficiário final" if receipt and receipt.beneficiario_final and not trigger_matches(original_history, rule.favorecido_normalizado) else "Extrato"
                matches.append({"movimento_id": movement.id, "data": movement.data.strftime("%d/%m/%Y") if movement.data else "—", "historico": rule_match_history(movement, receipt), "componente": entry.componente, "fonte": source})
    return matches


def rule_has_eligible_movement(rule: RegraContabil, reconciliation: Conciliacao, db: Session) -> bool:
    return bool(rule_preview(rule, reconciliation, db))


@router.post("/conciliacoes/{conciliacao_id}/regras-contabeis/previa")
def preview_accounting_rule(conciliacao_id: str, payload: RegraContabilPreviaInput, db: Session = Depends(get_db)):
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    current_rule_id = payload.regra_id.strip() or None
    if current_rule_id:
        current_rule = db.get(RegraContabil, current_rule_id)
        if not current_rule or current_rule.cliente_id != reconciliation.cliente_id or current_rule.banco != reconciliation.banco or not current_rule.ativo:
            raise HTTPException(404, "Regra não encontrada")
    rule = RegraContabil(cliente_id=reconciliation.cliente_id, conciliacao_id=reconciliation.id, banco=reconciliation.banco, tipo_fonte="extrato", tipo_operacao=payload.natureza, tipo_componente=payload.tipo_componente.strip().upper(), favorecido_normalizado=normalize_name(payload.gatilho), gatilho_comprovante_normalizado=normalize_name(payload.gatilho_comprovante), texto_exclusao_normalizado=normalize_name(payload.texto_exclusao))
    matches = rule_preview(rule, reconciliation, db, current_rule_id)
    return {"quantidade": len(matches), "lancamentos": matches, "motivo": "Nenhum lançamento elegível corresponde ao gatilho informado." if not matches else ""}


@router.post("/conciliacoes/{conciliacao_id}/regras-contabeis")
def create_accounting_rule(conciliacao_id: str, payload: RegraContabilInput, db: Session = Depends(get_db)):
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    if not (payload.gatilho.strip() or payload.gatilho_comprovante.strip()) or not all([payload.conta_debito.strip(), payload.conta_credito.strip(), payload.historico.strip()]):
        raise HTTPException(422, "Preencha ao menos um gatilho, débito, crédito e histórico")
    candidate_identity = rule_identity(reconciliation.cliente_id, reconciliation.banco, payload.natureza, payload.tipo_componente, payload.gatilho, payload.gatilho_comprovante, payload.texto_exclusao, payload.conta_debito, payload.conta_credito, payload.historico, payload.complemento)
    existing_rule = next((item for item in scoped_rules(reconciliation, db) if saved_rule_identity(item) == candidate_identity), None)
    if existing_rule:
        exception = rule_period_exception(existing_rule.id, reconciliation, db)
        if not exception:
            raise HTTPException(409, "Já existe uma regra equivalente neste banco")
        try:
            db.delete(exception)
            targets = [item for item in scoped_reconciliations(reconciliation, db) if item.banco == reconciliation.banco] if existing_rule.escopo == "global" else [reconciliation]
            for item in targets:
                apply_accounting_rules(item, db)
            applied = count_rule_entries_in_reconciliation(existing_rule.id, reconciliation, db)
            db.commit(); db.refresh(existing_rule)
        except Exception:
            db.rollback()
            raise
        return {"id": existing_rule.id, "movimentos_aplicados": applied, "reativada": True, "regras": accounting_rules(conciliacao_id, db)}
    scope = "global" if payload.escopo == "global" else "periodo"
    rule = RegraContabil(cliente_id=reconciliation.cliente_id, conciliacao_id=reconciliation.id, banco=reconciliation.banco, tipo_fonte="extrato", tipo_operacao=payload.natureza, tipo_componente=payload.tipo_componente.strip().upper(), favorecido_normalizado=normalize_name(payload.gatilho), gatilho_comprovante_normalizado=normalize_name(payload.gatilho_comprovante), texto_exclusao_normalizado=normalize_name(payload.texto_exclusao), conta_debito=payload.conta_debito.strip(), conta_credito=payload.conta_credito.strip(), historico=payload.historico.strip(), complemento=payload.complemento.strip(), escopo=scope)
    if not rule_has_eligible_movement(rule, reconciliation, db):
        raise HTTPException(422, "O gatilho não cobre nenhum lançamento elegível nesta conciliação")
    try:
        db.add(rule); db.flush()
        targets = [item for item in scoped_reconciliations(reconciliation, db) if item.banco == reconciliation.banco] if scope == "global" else [reconciliation]
        for item in targets:
            apply_accounting_rules(item, db)
        applied = count_rule_entries_in_reconciliation(rule.id, reconciliation, db)
        db.commit(); db.refresh(rule)
    except Exception:
        db.rollback()
        raise
    return {"id": rule.id, "movimentos_aplicados": applied, "regras": accounting_rules(conciliacao_id, db)}


def release_rule_entries(rule_ids: list[str], db: Session, conciliacao_id: str | None = None) -> None:
    """Returns source components to the pending list without losing their composition."""
    if not rule_ids:
        return
    entries = db.query(LancamentoContabil).filter(LancamentoContabil.regra_contabil_id.in_(rule_ids))
    if conciliacao_id:
        entries = entries.join(Correspondencia, LancamentoContabil.correspondencia_id == Correspondencia.id).filter(Correspondencia.conciliacao_id == conciliacao_id)
    for entry in entries:
        entry.regra_contabil_id = None
        entry.conta_debito = ""
        entry.conta_credito = ""
        entry.historico = entry.descricao or ""
        entry.status = "pendente_regra" if entry.origem in {"comprovante", "rfb"} else "pendente"


@router.patch("/conciliacoes/{conciliacao_id}/regras-contabeis/{regra_id}")
def update_accounting_rule(conciliacao_id: str, regra_id: str, payload: RegraContabilInput, db: Session = Depends(get_db)):
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    rule = db.get(RegraContabil, regra_id)
    if not rule or rule.cliente_id != reconciliation.cliente_id or rule.banco != reconciliation.banco or not rule.ativo:
        raise HTTPException(404, "Regra não encontrada")
    if not (payload.gatilho.strip() or payload.gatilho_comprovante.strip()) or not all([payload.conta_debito.strip(), payload.conta_credito.strip(), payload.historico.strip()]):
        raise HTTPException(422, "Preencha ao menos um gatilho, débito, crédito e histórico")
    candidate_identity = rule_identity(reconciliation.cliente_id, reconciliation.banco, payload.natureza, payload.tipo_componente, payload.gatilho, payload.gatilho_comprovante, payload.texto_exclusao, payload.conta_debito, payload.conta_credito, payload.historico, payload.complemento)
    if any(item.id != rule.id and saved_rule_identity(item) == candidate_identity for item in scoped_rules(reconciliation, db)):
        raise HTTPException(409, "Já existe uma regra equivalente neste banco")
    try:
        rule.tipo_operacao = payload.natureza
        rule.tipo_componente = payload.tipo_componente.strip().upper()
        rule.favorecido_normalizado = normalize_name(payload.gatilho)
        rule.gatilho_comprovante_normalizado = normalize_name(payload.gatilho_comprovante)
        rule.texto_exclusao_normalizado = normalize_name(payload.texto_exclusao)
        rule.conta_debito = payload.conta_debito.strip()
        rule.conta_credito = payload.conta_credito.strip()
        rule.historico = payload.historico.strip()
        rule.complemento = payload.complemento.strip()
        release_rule_entries([rule.id], db, reconciliation.id if rule.escopo == "periodo" else None)
        if not rule_has_eligible_movement(rule, reconciliation, db):
            raise HTTPException(422, "O gatilho não cobre nenhum lançamento elegível nesta conciliação")
        targets = [item for item in scoped_reconciliations(reconciliation, db) if item.banco == reconciliation.banco] if rule.escopo == "global" else [reconciliation]
        for item in targets:
            apply_accounting_rules(item, db)
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"id": rule.id, "regras": accounting_rules(conciliacao_id, db)}


def delete_rule_globally(rule: RegraContabil, db: Session) -> int:
    rule.ativo = False
    db.query(RegraContabilExcecao).filter_by(regra_contabil_id=rule.id).delete(synchronize_session=False)
    release_rule_entries([rule.id], db)
    for match in db.query(Correspondencia).filter_by(regra_contabil_id=rule.id):
        match.regra_contabil_id = None
    reconciliations = db.query(Conciliacao).filter_by(cliente_id=rule.cliente_id, banco=rule.banco).all()
    for item in reconciliations:
        apply_accounting_rules(item, db)
    return len(reconciliations)


@router.delete("/conciliacoes/{conciliacao_id}/regras-contabeis/{regra_id}/periodo")
def ignore_rule_in_period(conciliacao_id: str, regra_id: str, db: Session = Depends(get_db)):
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    rule = db.get(RegraContabil, regra_id)
    if not rule or rule.cliente_id != reconciliation.cliente_id or rule.banco != reconciliation.banco or not rule.ativo:
        raise HTTPException(404, "Regra não encontrada para este cliente e banco")
    try:
        if not db.query(RegraContabilExcecao).filter_by(regra_contabil_id=rule.id, conciliacao_id=reconciliation.id).first():
            db.add(RegraContabilExcecao(regra_contabil_id=rule.id, conciliacao_id=reconciliation.id))
        release_rule_entries([rule.id], db, reconciliation.id)
        for match in db.query(Correspondencia).filter_by(conciliacao_id=reconciliation.id, regra_contabil_id=rule.id):
            match.regra_contabil_id = None
        apply_accounting_rules(reconciliation, db)
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"message": f"Regra removida somente deste período ({reconciliation.data_inicio.strftime('%m/%Y')}).", "regras": accounting_rules(conciliacao_id, db)}


@router.delete("/conciliacoes/{conciliacao_id}/regras-contabeis/{regra_id}/periodo/excecao")
def restore_rule_in_period(conciliacao_id: str, regra_id: str, db: Session = Depends(get_db)):
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    rule = db.get(RegraContabil, regra_id)
    exception = db.query(RegraContabilExcecao).filter_by(regra_contabil_id=regra_id, conciliacao_id=conciliacao_id).first()
    if not rule or rule.cliente_id != reconciliation.cliente_id or rule.banco != reconciliation.banco or not exception:
        raise HTTPException(404, "Regra ignorada não encontrada neste período")
    try:
        db.delete(exception)
        apply_accounting_rules(reconciliation, db)
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"message": f"Regra restaurada neste período ({reconciliation.data_inicio.strftime('%m/%Y')}).", "regras": accounting_rules(conciliacao_id, db)}


@router.post("/conciliacoes/{conciliacao_id}/regras-contabeis/ocultas/restaurar-com-cobertura")
def restore_covered_hidden_accounting_rules(conciliacao_id: str, db: Session = Depends(get_db)):
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    exceptions = db.query(RegraContabilExcecao).filter_by(conciliacao_id=reconciliation.id).all()
    rule_ids = [item.regra_contabil_id for item in exceptions]
    rules_by_id = {item.id: item for item in db.query(RegraContabil).filter(RegraContabil.id.in_(rule_ids)).all()} if rule_ids else {}
    restored = []
    try:
        for exception in exceptions:
            rule = rules_by_id.get(exception.regra_contabil_id)
            if not rule or not rule.ativo or rule.cliente_id != reconciliation.cliente_id or rule.banco != reconciliation.banco:
                continue
            if rule_preview(rule, reconciliation, db):
                db.delete(exception)
                restored.append(rule.id)
        if restored:
            apply_accounting_rules(reconciliation, db)
        db.commit()
    except Exception:
        db.rollback()
        raise
    count = len(restored)
    message = f"{count} regra(s) existente(s) restaurada(s) e aplicada(s)." if count else "Nenhuma regra oculta com cobertura encontrada neste período."
    return {"message": message, "quantidade": count, "regras": accounting_rules(conciliacao_id, db)}


@router.delete("/conciliacoes/{conciliacao_id}/regras-contabeis/sem-cobertura")
def delete_zero_covered_accounting_rules(conciliacao_id: str, db: Session = Depends(get_db)):
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    try:
        zero_rule_ids = [
            item["id"]
            for item in accounting_rules(conciliacao_id, db, auto_hide_zero_covered=False)["salvas"]
            if item["cobertos"] == 0
        ]
        for rule_id in zero_rule_ids:
            if not db.query(RegraContabilExcecao).filter_by(regra_contabil_id=rule_id, conciliacao_id=reconciliation.id).first():
                db.add(RegraContabilExcecao(regra_contabil_id=rule_id, conciliacao_id=reconciliation.id))
        if zero_rule_ids:
            release_rule_entries(zero_rule_ids, db, reconciliation.id)
            for match in db.query(Correspondencia).filter_by(conciliacao_id=reconciliation.id):
                if match.regra_contabil_id in zero_rule_ids:
                    match.regra_contabil_id = None
            apply_accounting_rules(reconciliation, db)
        db.commit()
    except Exception:
        db.rollback()
        raise
    count = len(zero_rule_ids)
    message = f"{count} regra(s) sem cobertura ocultada(s) somente deste período." if count else "Não havia regras sem cobertura neste período."
    return {"message": message, "quantidade": count, "regras": accounting_rules(conciliacao_id, db)}


@router.delete("/conciliacoes/{conciliacao_id}/regras-contabeis/{regra_id}")
def delete_accounting_rule_for_reconciliation(conciliacao_id: str, regra_id: str, db: Session = Depends(get_db)):
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    rule = db.get(RegraContabil, regra_id)
    if not rule or rule.cliente_id != reconciliation.cliente_id or rule.banco != reconciliation.banco or not rule.ativo:
        raise HTTPException(404, "Regra não encontrada para este cliente e banco")
    try:
        if rule.escopo == "periodo":
            rule.ativo = False
            db.query(RegraContabilExcecao).filter_by(regra_contabil_id=rule.id, conciliacao_id=reconciliation.id).delete(synchronize_session=False)
            release_rule_entries([rule.id], db, reconciliation.id)
            affected = 1
            message = f"Regra excluída deste período ({reconciliation.data_inicio.strftime('%m/%Y')})."
        else:
            affected = delete_rule_globally(rule, db)
            message = f"Regra excluída de {affected} período{'s' if affected != 1 else ''}."
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"message": message, "regras": accounting_rules(conciliacao_id, db), "periodos_afetados": affected}


@router.delete("/regras-contabeis/{regra_id}")
def delete_accounting_rule(regra_id: str, db: Session = Depends(get_db)):
    rule = db.get(RegraContabil, regra_id)
    if not rule:
        raise HTTPException(404, "Regra não encontrada")
    try:
        affected = delete_rule_globally(rule, db)
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"periodos_afetados": affected}


@router.delete("/conciliacoes/{conciliacao_id}/regras-contabeis")
def delete_all_accounting_rules(conciliacao_id: str, db: Session = Depends(get_db)):
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    try:
        rules = [rule for rule in db.query(RegraContabil).filter_by(cliente_id=reconciliation.cliente_id, banco=reconciliation.banco, tipo_fonte="extrato").all() if rule.escopo == "global" or rule.conciliacao_id == reconciliation.id]
        local_rules = [rule for rule in rules if rule.escopo == "periodo"]
        global_rules = [rule for rule in rules if rule.escopo == "global" and rule.ativo]
        inactive_rules = [rule for rule in rules if not rule.ativo]
        for rule in local_rules:
            rule.ativo = False
        if local_rules:
            release_rule_entries([rule.id for rule in local_rules], db, reconciliation.id)
        if inactive_rules:
            release_rule_entries([rule.id for rule in inactive_rules], db, reconciliation.id)
        for rule in global_rules:
            if not db.query(RegraContabilExcecao).filter_by(regra_contabil_id=rule.id, conciliacao_id=reconciliation.id).first():
                db.add(RegraContabilExcecao(regra_contabil_id=rule.id, conciliacao_id=reconciliation.id))
            release_rule_entries([rule.id], db, reconciliation.id)
        for match in db.query(Correspondencia).filter_by(conciliacao_id=reconciliation.id):
            if match.regra_contabil_id in {rule.id for rule in rules}:
                match.regra_contabil_id = None
        apply_accounting_rules(reconciliation, db)
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"message": "Regras removidas somente deste período.", "regras": accounting_rules(conciliacao_id, db)}


def accounting_csv_response(conciliacao_id: str, db: Session):
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    sync_getnet_anticipation_adjustments(reconciliation, db)
    db.commit()
    account = db.query(ContaBancaria).filter_by(cliente_id=reconciliation.cliente_id, banco=reconciliation.banco).first()
    if not account or not account.conta_contabil.strip():
        raise HTTPException(422, "Informe a conta deste banco no plano de contas antes de gerar o CSV")
    integrity = accounting_integrity(reconciliation, db)
    difference = integrity["diferenca"]
    if abs(difference) > Decimal("0.01"):
        raise HTTPException(422, f"Não foi possível gerar o CSV. O Razão possui uma diferença de R$ {abs(difference):.2f}. Revise os lançamentos pendentes ou incompletos.")
    if integrity["movimentos_incompletos"]:
        references = ", ".join(item["data"] for item in integrity["movimentos_incompletos"][:3])
        raise HTTPException(422, f"Não foi possível gerar o CSV. Existem lançamentos incompletos em {references}.")
    output = io.StringIO(newline="")
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow(["Data", "Debito", "Credito", "Historico", "Valor", "Complemento"])
    entries = integrity["lancamentos_validos"]
    rows = []
    for entry in entries:
        match = db.get(Correspondencia, entry.correspondencia_id)
        movement = db.get(MovimentoExtrato, match.movimento_extrato_id) if match else None
        rule = db.get(RegraContabil, entry.regra_contabil_id) if entry.regra_contabil_id else None
        rows.append((movement.data if movement and movement.data else date.max, accounting_export_order(entry), movement, entry, rule))
    for _, _, movement, entry, rule in sorted(rows):
        writer.writerow([
            movement.data.strftime("%d/%m/%Y") if movement and movement.data else "",
            accounting_code(entry.conta_debito),
            accounting_code(entry.conta_credito),
            accounting_code(entry.historico),
            f"{entry.valor:.2f}",
            entry.complemento or (rule.complemento if rule else ""),
        ])
    period = reconciliation.data_inicio.strftime("%m%y")
    bank_account = re.sub(r"\D", "", accounting_code(account.conta_contabil)) or "conta_bancaria"
    filename = f"{period}{bank_account}.csv"
    return Response(output.getvalue().encode("utf-8-sig"), media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


def accounting_code(value: str) -> str:
    """Returns the leading code stored as `codigo - descricao` in catalogs."""
    match = re.match(r"^\s*(\d+(?:[.\-]\d+)*)\s*(?:[-:]\s*|\s+)", value or "")
    return match.group(1) if match else (value or "").strip()


@router.get("/conciliacoes/{conciliacao_id}/lancamentos-contabeis.csv")
def accounting_csv(conciliacao_id: str, db: Session = Depends(get_db)):
    return accounting_csv_response(conciliacao_id, db)


@router.get("/conciliacoes/{conciliacao_id}/lancamentos-contabeis.pdf")
def accounting_pdf(conciliacao_id: str, db: Session = Depends(get_db)):
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    sync_getnet_anticipation_adjustments(reconciliation, db)
    db.commit()
    integrity = accounting_integrity(reconciliation, db)
    if abs(integrity["diferenca"]) > Decimal("0.01"):
        raise HTTPException(422, f"Não foi possível gerar o PDF. O Razão possui uma diferença de R$ {abs(integrity['diferenca']):.2f}.")
    if integrity["movimentos_incompletos"]:
        raise HTTPException(422, "Não foi possível gerar o PDF. Existem lançamentos incompletos.")
    months = ("Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro")
    rows = []
    for entry in integrity["lancamentos_validos"]:
        match = db.get(Correspondencia, entry.correspondencia_id)
        movement = db.get(MovimentoExtrato, match.movimento_extrato_id) if match else None
        if movement and movement.data:
            rows.append((movement.data, entry))
    rows.sort(key=lambda item: (item[0], accounting_export_order(item[1])))
    document = fitz.open()
    page = document.new_page(width=842, height=595)
    y = 48

    def line(text: str, size: float = 8, bold: bool = False):
        nonlocal page, y
        if y > 555:
            page = document.new_page(width=842, height=595)
            y = 48
        font = "hebo" if bold else "helv"
        page.insert_text((42, y), text, fontsize=size, fontname=font)
        y += size + 4

    columns = [(42, "Data", 11), (96, "Debito", 15), (170, "Credito", 15), (244, "Historico", 18), (332, "Valor", 13), (402, "Complemento", 62)]

    def table_header():
        nonlocal y
        if y > 535:
            line("", 3)
        for x, label, _ in columns:
            page.insert_text((x, y), label, fontsize=7, fontname="hebo")
        y += 12

    def table_row(values: list[str], italic: bool = False):
        nonlocal page, y
        fragments = [textwrap.wrap(value, width=width) or [""] for value, (_, _, width) in zip(values, columns)]
        height = max(len(parts) for parts in fragments) * 10 + 3
        if y + height > 555:
            page = document.new_page(width=842, height=595)
            y = 48
            table_header()
        for index, parts in enumerate(fragments):
            x = columns[index][0]
            for row, part in enumerate(parts):
                page.insert_text((x, y + row * 10), part, fontsize=7, fontname="helvi" if italic else "helv")
        y += height

    line("Relatório de lançamentos contábeis", 15, True)
    line(f"Banco: {reconciliation.banco} | Período: {reconciliation.data_inicio.strftime('%d/%m/%Y')} a {reconciliation.data_fim.strftime('%d/%m/%Y')}", 9)
    line("", 4)
    table_header()
    current_month = None
    bank_total = Decimal("0.00")
    other_total = Decimal("0.00")
    month_bank = Decimal("0.00")
    month_other = Decimal("0.00")
    for movement_date, entry in rows:
        month_key = (movement_date.year, movement_date.month)
        if current_month and month_key != current_month:
            line(f"Subtotal {months[current_month[1] - 1]}: Bancários R$ {month_bank:,.2f} | Outros R$ {month_other:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."), 8, True)
            line("", 3)
            month_bank = Decimal("0.00")
            month_other = Decimal("0.00")
        if month_key != current_month:
            line(f"{months[movement_date.month - 1]} de {movement_date.year}", 11, True)
            current_month = month_key
        is_other = is_other_accounting_entry(entry)
        if is_other:
            month_other += entry.valor
            other_total += entry.valor
        else:
            month_bank += entry.valor
            bank_total += entry.valor
        value = f"R$ {entry.valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        rule = db.get(RegraContabil, entry.regra_contabil_id) if entry.regra_contabil_id else None
        table_row([
            movement_date.strftime("%d/%m/%Y"),
            accounting_code(entry.conta_debito),
            accounting_code(entry.conta_credito),
            accounting_code(entry.historico),
            value,
            entry.complemento or (rule.complemento if rule else ""),
        ], italic=is_other)
    if current_month:
        line(f"Subtotal {months[current_month[1] - 1]}: Bancários R$ {month_bank:,.2f} | Outros R$ {month_other:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."), 8, True)
    line("", 5)
    line(f"Total do período: Bancários R$ {bank_total:,.2f} | Outros R$ {other_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."), 10, True)
    return Response(document.tobytes(), media_type="application/pdf", headers={"Content-Disposition": 'attachment; filename="lancamentos-contabeis.pdf"'})


@router.post("/conciliacoes/{conciliacao_id}/arquivos")
def upload(conciliacao_id: str, tipo_documento: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    if tipo_documento not in DOCUMENT_TYPES or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(422, "Envie um PDF com tipo de documento válido")
    reconciliation = db.get(Conciliacao, conciliacao_id)
    if not reconciliation:
        raise HTTPException(404, "Conciliação não encontrada")
    reset_reconciliation(conciliacao_id, db)
    destination = UPLOAD_DIR / conciliacao_id
    destination.mkdir(parents=True, exist_ok=True)
    path = destination / f"{uuid4()}.pdf"
    with path.open("wb") as output:
        shutil.copyfileobj(file.file, output)
    record = Arquivo(conciliacao_id=conciliacao_id, tipo_documento=tipo_documento, banco_selecionado=reconciliation.banco, nome_original=file.filename, caminho=str(path))
    db.add(record); db.commit(); db.refresh(record)
    try:
        pages, extracted = extract_statement_document(path, reconciliation.banco, tipo_documento) if tipo_documento == "extrato" else (extract_pdf_pages(path, reconciliation.banco, tipo_documento), [])
        record.texto_bruto = "\n\f\n".join(pages); record.paginas = len(pages)
        for number, page_text in enumerate(pages, 1):
            parsed_rfb = parse_rfb_page(page_text, number) if tipo_documento == "rfb" else None
            if tipo_documento != "extrato":
                extracted.extend([parsed_rfb] if parsed_rfb else [] if tipo_documento == "rfb" else extract_invoices(page_text, number) if tipo_documento == INVOICE_DOCUMENT_TYPE else extract_loan_receipts(page_text, number) if tipo_documento == LOAN_DOCUMENT_TYPE else extract_receipts(page_text, number))
        if tipo_documento == "extrato" and reconciliation.banco == "Banco do Brasil":
            extracted = deduplicate_statement_records(extracted)
        for position, item in enumerate(extracted, 1):
            if tipo_documento == "rfb":
                receipt = ComprovanteRfb(conciliacao_id=conciliacao_id, arquivo_id=record.id, pagina_numero=item.pagina_numero, tipo=item.tipo, cnpj=item.cnpj, razao_social=item.razao_social, competencia=item.competencia, periodo_apuracao=item.periodo_apuracao, data_vencimento=item.data_vencimento, data_arrecadacao=item.data_arrecadacao, numero_documento=item.numero_documento, codigo_banco=item.codigo_banco, nome_banco=item.nome_banco, agencia=item.agencia, valor_principal=item.valor_principal, valor_multa=item.valor_multa, valor_juros=item.valor_juros, valor_total=item.valor_total, texto_original=item.texto_original, status="composição divergente" if item.composicao_divergente else "pronto")
                db.add(receipt); db.flush()
                for tax in item.itens:
                    db.add(ComprovanteRfbItem(comprovante_rfb_id=receipt.id, codigo=tax.codigo, descricao=tax.descricao, valor_principal=tax.valor_principal, valor_multa=tax.valor_multa, valor_juros=tax.valor_juros, valor_total=tax.valor_total))
                continue
            if tipo_documento == INVOICE_DOCUMENT_TYPE:
                db.add(NotaFiscal(conciliacao_id=conciliacao_id, arquivo_id=record.id, pagina_numero=item.pagina_numero, texto_original=item.texto_original, dados_originais={"tipo_documento": tipo_documento, **item.dados}, dados_normalizados={"nome": normalize_name(item.fornecedor)}, data_emissao=item.data_emissao, fornecedor=item.fornecedor, cpf_cnpj=item.cpf_cnpj, numero_nota=item.numero_nota, valor_total=item.valor_total))
                continue
            original_data = {"origem_nome": item.origem_nome, "tipo_documento": tipo_documento} if tipo_documento in RECEIPT_DOCUMENT_TYPES else {}
            if tipo_documento == "extrato":
                original_data["ordem_extrato"] = position
                if getattr(item, "numero_documento", ""):
                    original_data["numero_documento"] = item.numero_documento
            common = dict(conciliacao_id=conciliacao_id, arquivo_id=record.id, pagina_numero=item.pagina_numero, texto_original=item.texto_original, dados_originais=original_data, dados_normalizados={"nome": normalize_name(item.favorecido if tipo_documento in RECEIPT_DOCUMENT_TYPES else item.nome)})
            if tipo_documento in RECEIPT_DOCUMENT_TYPES:
                common.update(beneficiario=item.beneficiario or item.favorecido, nome_fantasia=item.nome_fantasia, beneficiario_final=item.beneficiario_final, pagador=item.pagador, cnpj_beneficiario=item.cnpj_beneficiario, cnpj_beneficiario_final=item.cnpj_beneficiario_final, numero_documento=item.numero_documento)
            db.add(Comprovante(**common, data=item.data, hora=item.hora, favorecido=item.favorecido, valor=item.financeiros.valor_pago, valor_original=item.financeiros.valor_original, valor_desconto=item.financeiros.valor_desconto, valor_abatimento=item.financeiros.valor_abatimento, valor_desconto_abatimento=item.financeiros.valor_desconto_abatimento, valor_juros=item.financeiros.valor_juros, valor_multa=item.financeiros.valor_multa, valor_encargos=item.financeiros.valor_encargos, valor_tarifa=item.financeiros.valor_tarifa, valor_pago=item.financeiros.valor_pago, detalhes_financeiros=item.financeiros.detalhes, status_revisao="revisao" if item.financeiros.composicao_divergente else "valido", tipo_operacao=item.tipo_operacao) if tipo_documento in RECEIPT_DOCUMENT_TYPES else MovimentoExtrato(**common, data=item.data, hora=item.hora, historico=item.historico, nome_encontrado=item.nome, valor=item.valor, natureza=item.natureza, data_origem=item.data_origem))
        if tipo_documento == "extrato" and any(not rule.gatilho_comprovante_normalizado for rule in current_bank_rules(reconciliation, db)):
            apply_accounting_rules(reconciliation, db)
        sync_getnet_anticipation_adjustments(reconciliation, db)
        record.status_processamento = "concluido"
    except Exception as error:
        record.status_processamento = "erro"; record.mensagem_erro = str(error)
    db.commit()
    return {"id": record.id, "status": record.status_processamento}


def clear_reconciliation_results(conciliacao_id: str, db: Session) -> None:
    matches = db.query(Correspondencia).filter_by(conciliacao_id=conciliacao_id).all()
    for match in matches:
        db.query(LancamentoContabil).filter_by(correspondencia_id=match.id).delete()
    db.query(Correspondencia).filter_by(conciliacao_id=conciliacao_id).delete()


def sync_document_items(match: Correspondencia, lines, db: Session) -> None:
    """Refresh source-derived components without replacing user edits."""
    existing = db.query(LancamentoContabil).filter_by(correspondencia_id=match.id).all()
    automatic = [item for item in existing if item.status != "editado_manual"]
    manual_keys = {(item.origem, item.componente, item.codigo_receita, item.descricao) for item in existing if item.status == "editado_manual"}
    by_key = {(item.origem, item.componente, item.codigo_receita, item.descricao): item for item in automatic}
    expected = set()
    for order, line in enumerate(lines, 1):
        key = (line.origem, line.componente, line.codigo_receita, line.descricao)
        expected.add(key)
        if key in manual_keys:
            continue
        item = by_key.get(key)
        if not item:
            item = LancamentoContabil(correspondencia_id=match.id, origem=line.origem)
            db.add(item)
        item.componente = line.componente
        item.categoria = line.componente
        item.codigo_receita = line.codigo_receita
        item.descricao = line.descricao
        item.tributo = line.descricao if line.origem == "rfb" else ""
        item.valor = line.valor
        item.efeito_no_total = line.efeito_no_total
        item.ordem = order
        item.historico = item.historico or line.descricao
        _, _, complement, _ = tax_complement(item, match, db)
        if complement:
            item.complemento = complement
        item.status = "pendente_regra"
    for item in automatic:
        key = (item.origem, item.componente, item.codigo_receita, item.descricao)
        if key not in expected:
            db.delete(item)


def reset_reconciliation(conciliacao_id: str, db: Session) -> None:
    clear_reconciliation_results(conciliacao_id, db)
    reconciliation = db.get(Conciliacao, conciliacao_id)
    if not reconciliation:
        return
    reconciliation.status = "rascunho"
    if reconciliation.processo_id:
        process = db.get(ProcessoConciliacao, reconciliation.processo_id)
        if process:
            process.status = "em_andamento"


@router.post("/arquivos/{arquivo_id}/reprocessar")
def reprocess_document(arquivo_id: str, db: Session = Depends(get_db)):
    record = db.get(Arquivo, arquivo_id)
    if not record or record.tipo_documento not in DOCUMENT_TYPES:
        raise HTTPException(404, "Documento reprocessável não encontrado")
    reset_reconciliation(record.conciliacao_id, db)
    model = MovimentoExtrato if record.tipo_documento == "extrato" else Comprovante if record.tipo_documento in RECEIPT_DOCUMENT_TYPES else NotaFiscal if record.tipo_documento == INVOICE_DOCUMENT_TYPE else ComprovanteRfb
    db.query(model).filter_by(arquivo_id=record.id).delete()
    pages = (record.texto_bruto or "").split("\n\f\n")
    if Path(record.caminho).is_file():
        try:
            if record.tipo_documento == "extrato":
                pages, extracted = extract_statement_document(Path(record.caminho), record.banco_selecionado, record.tipo_documento)
            else:
                pdf_pages = extract_pdf_pages(Path(record.caminho), record.banco_selecionado, record.tipo_documento)
                if pdf_pages:
                    pages = pdf_pages
                extracted = []
        except ValueError as error:
            record.status_processamento = "erro"
            record.mensagem_erro = str(error)
            db.commit()
            return {"registros_extraidos": 0, "status": record.status_processamento}
        except Exception:
            logger.warning("Falha ao reler PDF; usando texto bruto salvo para arquivo %s", record.id)
            extracted = extract_statement_pages(pages, record.banco_selecionado) if record.tipo_documento == "extrato" else []
    else:
        extracted = extract_statement_pages(pages, record.banco_selecionado) if record.tipo_documento == "extrato" else []
    record.texto_bruto = "\n\f\n".join(pages); record.paginas = len(pages)
    for number, page_text in enumerate(pages, 1):
        parsed_rfb = parse_rfb_page(page_text, number) if record.tipo_documento == "rfb" else None
        if record.tipo_documento != "extrato":
            extracted.extend([parsed_rfb] if parsed_rfb else [] if record.tipo_documento == "rfb" else extract_invoices(page_text, number) if record.tipo_documento == INVOICE_DOCUMENT_TYPE else extract_loan_receipts(page_text, number) if record.tipo_documento == LOAN_DOCUMENT_TYPE else extract_receipts(page_text, number))
    if record.tipo_documento == "extrato" and record.banco_selecionado == "Banco do Brasil":
        extracted = deduplicate_statement_records(extracted)
    for position, item in enumerate(extracted, 1):
        if record.tipo_documento == "rfb":
            receipt = ComprovanteRfb(conciliacao_id=record.conciliacao_id, arquivo_id=record.id, pagina_numero=item.pagina_numero, tipo=item.tipo, cnpj=item.cnpj, razao_social=item.razao_social, competencia=item.competencia, periodo_apuracao=item.periodo_apuracao, data_vencimento=item.data_vencimento, data_arrecadacao=item.data_arrecadacao, numero_documento=item.numero_documento, codigo_banco=item.codigo_banco, nome_banco=item.nome_banco, agencia=item.agencia, valor_principal=item.valor_principal, valor_multa=item.valor_multa, valor_juros=item.valor_juros, valor_total=item.valor_total, texto_original=item.texto_original, status="composição divergente" if item.composicao_divergente else "pronto")
            db.add(receipt); db.flush()
            for tax in item.itens:
                db.add(ComprovanteRfbItem(comprovante_rfb_id=receipt.id, codigo=tax.codigo, descricao=tax.descricao, valor_principal=tax.valor_principal, valor_multa=tax.valor_multa, valor_juros=tax.valor_juros, valor_total=tax.valor_total))
            continue
        if record.tipo_documento == INVOICE_DOCUMENT_TYPE:
            db.add(NotaFiscal(conciliacao_id=record.conciliacao_id, arquivo_id=record.id, pagina_numero=item.pagina_numero, texto_original=item.texto_original, dados_originais={"tipo_documento": record.tipo_documento, **item.dados}, dados_normalizados={"nome": normalize_name(item.fornecedor)}, data_emissao=item.data_emissao, fornecedor=item.fornecedor, cpf_cnpj=item.cpf_cnpj, numero_nota=item.numero_nota, valor_total=item.valor_total))
            continue
        original_data = {"origem_nome": item.origem_nome, "tipo_documento": record.tipo_documento} if record.tipo_documento in RECEIPT_DOCUMENT_TYPES else {}
        if record.tipo_documento == "extrato":
            original_data["ordem_extrato"] = position
            if getattr(item, "numero_documento", ""):
                original_data["numero_documento"] = item.numero_documento
        common = dict(conciliacao_id=record.conciliacao_id, arquivo_id=record.id, pagina_numero=item.pagina_numero, texto_original=item.texto_original, dados_originais=original_data, dados_normalizados={"nome": normalize_name(item.nome if record.tipo_documento == "extrato" else item.favorecido)})
        if record.tipo_documento in RECEIPT_DOCUMENT_TYPES:
            common.update(beneficiario=item.beneficiario or item.favorecido, nome_fantasia=item.nome_fantasia, beneficiario_final=item.beneficiario_final, pagador=item.pagador, cnpj_beneficiario=item.cnpj_beneficiario, cnpj_beneficiario_final=item.cnpj_beneficiario_final, numero_documento=item.numero_documento)
        db.add(MovimentoExtrato(**common, data=item.data, hora=item.hora, historico=item.historico, nome_encontrado=item.nome, valor=item.valor, natureza=item.natureza, data_origem=item.data_origem) if record.tipo_documento == "extrato" else Comprovante(**common, data=item.data, hora=item.hora, favorecido=item.favorecido, valor=item.financeiros.valor_pago, valor_original=item.financeiros.valor_original, valor_desconto=item.financeiros.valor_desconto, valor_abatimento=item.financeiros.valor_abatimento, valor_desconto_abatimento=item.financeiros.valor_desconto_abatimento, valor_juros=item.financeiros.valor_juros, valor_multa=item.financeiros.valor_multa, valor_encargos=item.financeiros.valor_encargos, valor_tarifa=item.financeiros.valor_tarifa, valor_pago=item.financeiros.valor_pago, detalhes_financeiros=item.financeiros.detalhes, status_revisao="revisao" if item.financeiros.composicao_divergente else "valido", tipo_operacao=item.tipo_operacao))
    reconciliation = db.get(Conciliacao, record.conciliacao_id)
    if record.tipo_documento == "extrato" and reconciliation and any(not rule.gatilho_comprovante_normalizado for rule in current_bank_rules(reconciliation, db)):
        apply_accounting_rules(reconciliation, db)
    if reconciliation:
        sync_getnet_anticipation_adjustments(reconciliation, db)
    record.status_processamento = "concluido"
    record.mensagem_erro = None
    db.commit()
    return {"registros_extraidos": len(extracted), "status": record.status_processamento}


@router.post("/conciliacoes/{conciliacao_id}/conciliar")
def reconcile(conciliacao_id: str, db: Session = Depends(get_db)):
    reconciliation = db.get(Conciliacao, conciliacao_id)
    if not reconciliation:
        raise HTTPException(404, "Conciliação não encontrada")
    receipts = db.query(Comprovante).filter_by(conciliacao_id=conciliacao_id, ativo=True).all()
    rfb_receipts = [item for item in db.query(ComprovanteRfb).filter_by(conciliacao_id=conciliacao_id) if belongs_to_selected_bank(item, reconciliation.banco)]
    existing_matches = {item.movimento_extrato_id: item for item in db.query(Correspondencia).filter_by(conciliacao_id=conciliacao_id).all()}
    used_receipts, used_rfb = set(), set()
    movements = [item for item in db.query(MovimentoExtrato).filter_by(conciliacao_id=conciliacao_id, ativo=True) if movement_used_in_period(item) and is_statement_debit(item.natureza) and in_reconciliation_period(reconciliation, item)]
    movements.sort(key=movement_statement_order)
    for movement in movements:
        # The extracted counterparty is more precise than the operation history.
        movement_text = " ".join(item for item in [movement.historico, movement.nome_encontrado] if item)
        movement_document = movement_document_number(movement)
        tariff = "TARIFA PIX" in normalize_name(movement.historico)
        receipt_candidates = [(item, "tarifa do comprovante") for item in receipts if tariff and item.valor_tarifa == movement.valor and receipt_tariff_date_matches(movement, item)]
        if not tariff:
            receipt_candidates = [(item, receipt_match_criterion(movement_text, item, movement_document)) for item in receipts if item.id not in used_receipts and item.valor_pago == movement.valor and item.data == movement.data and receipt_operation_matches(movement, item) and receipt_time_matches(movement, item)]
        receipt_candidates = [(item, criterion) for item, criterion in receipt_candidates if criterion]
        if not receipt_candidates:
            receipt, criterion = fallback_boleto_receipt_candidate(movement, receipts, used_receipts)
            receipt_candidates = [(receipt, criterion)] if receipt else []
        if not receipt_candidates and is_transfer_without_counterparty(movement):
            transfer_candidates = [item for item in receipts if item.id not in used_receipts and item.valor_pago == movement.valor and item.data == movement.data and "TRANSFER" in normalize_name(item.tipo_operacao)]
            if len(transfer_candidates) == 1:
                receipt_candidates = [(transfer_candidates[0], "data, valor e tipo transferência")]
        receipt, criterion = receipt_candidates[0] if receipt_candidates else (None, "")
        rfb = next((item for item in rfb_receipts if item.id not in used_rfb and item.valor_total == movement.valor and item.data_arrecadacao == movement.data), None)
        source_rfb = None
        if rfb:
            items = db.query(ComprovanteRfbItem).filter_by(comprovante_rfb_id=rfb.id).all()
            source_rfb = SimpleNamespace(tipo=rfb.tipo, valor_principal=rfb.valor_principal, valor_multa=rfb.valor_multa, valor_juros=rfb.valor_juros, valor_total=rfb.valor_total, itens=items)
        decision = choose_rule_source(movement.valor, receipt, source_rfb, tariff=tariff)
        match = existing_matches.get(movement.id)
        if not match:
            match = Correspondencia(conciliacao_id=conciliacao_id, movimento_extrato_id=movement.id)
            db.add(match); db.flush()
        match.comprovante_id = receipt.id if receipt else None
        match.comprovante_rfb_id = rfb.id if rfb else None
        match.fonte_regra = decision.fonte_regra
        match.confianca = "alta" if not decision.exige_revisao else "média"
        match.criterio_correspondencia = f"Correspondência pelo {criterion}" if criterion else ""
        match.status = decision.status
        sync_document_items(match, decision.linhas, db)
        if receipt and not tariff: used_receipts.add(receipt.id)
        if rfb: used_rfb.add(rfb.id)
    apply_accounting_rules(reconciliation, db)
    sync_getnet_anticipation_adjustments(reconciliation, db)
    reconciliation.status = "concluido"
    if reconciliation.processo_id:
        process = db.get(ProcessoConciliacao, reconciliation.processo_id)
        if process:
            process.status = "concluido" if all(item.status == "concluido" for item in db.query(Conciliacao).filter_by(processo_id=process.id)) else "em_andamento"
    db.commit()
    return {"conciliacoes_geradas": len(movements), "resultados": result(conciliacao_id, db)}


@router.get("/conciliacoes/{conciliacao_id}/resultado")
def result(conciliacao_id: str, db: Session = Depends(get_db), response: FastAPIResponse = None):
    if response:
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    matches = db.query(Correspondencia).filter_by(conciliacao_id=conciliacao_id).all()
    matches_by_id = {item.id: item for item in matches}
    match_ids = [item.id for item in matches]
    movement_ids = [item.movimento_extrato_id for item in matches]
    receipt_ids = [item.comprovante_id for item in matches if item.comprovante_id]
    rfb_ids = [item.comprovante_rfb_id for item in matches if item.comprovante_rfb_id]
    movements = {item.id: item for item in db.query(MovimentoExtrato).filter(MovimentoExtrato.id.in_(movement_ids)).all()} if movement_ids else {}
    receipts = {item.id: item for item in db.query(Comprovante).filter(Comprovante.id.in_(receipt_ids)).all()} if receipt_ids else {}
    rfb_receipts = {item.id: item for item in db.query(ComprovanteRfb).filter(ComprovanteRfb.id.in_(rfb_ids)).all()} if rfb_ids else {}
    rfb_items_by_receipt: dict[str, list[ComprovanteRfbItem]] = defaultdict(list)
    if rfb_ids:
        for item in db.query(ComprovanteRfbItem).filter(ComprovanteRfbItem.comprovante_rfb_id.in_(rfb_ids)).all():
            rfb_items_by_receipt[item.comprovante_rfb_id].append(item)
    entries_by_match: dict[str, list[LancamentoContabil]] = defaultdict(list)
    rule_ids = set()
    if match_ids:
        for entry in db.query(LancamentoContabil).filter(LancamentoContabil.correspondencia_id.in_(match_ids)).order_by(LancamentoContabil.ordem, LancamentoContabil.id).all():
            entries_by_match[entry.correspondencia_id].append(entry)
            if entry.regra_contabil_id:
                rule_ids.add(entry.regra_contabil_id)
    rules = {item.id: item for item in db.query(RegraContabil).filter(RegraContabil.id.in_(rule_ids)).all()} if rule_ids else {}
    rows = []
    def money(value):
        if value is None: return "—"
        integer, decimal = f"{value:.2f}".split(".")
        return f"R$ {int(integer):,}".replace(",", ".") + f",{decimal}"
    def payment_type(history):
        text = history.upper()
        if "TARIFA" in text: return "Tarifa"
        if "COBRANÇA" in text: return "Cobrança"
        if "SEG CRÉD" in text or "SEGURO" in text: return "Seguro"
        if "RENDE FÁCIL" in text or "RENDIMENTO" in text: return "Rendimento"
        if "PIX" in text: return "PIX"
        if "TED" in text or "TRANSFERÊNCIA" in text: return "Transferência"
        if "BOLETO" in text: return "Boleto"
        if "IMPOSTO" in text or "DAS" in text: return "Imposto"
        if "CARTÃO" in text: return "Cartão"
        return "Outro"
    def tax_complement_cached(entry: LancamentoContabil, match: Correspondencia) -> tuple[str, str, str, str]:
        rfb = rfb_receipts.get(match.comprovante_rfb_id) if match.comprovante_rfb_id else None
        receipt = receipts.get(match.comprovante_id) if match.comprovante_id else None
        movement = movements.get(match.movimento_extrato_id)
        rfb_items = rfb_items_by_receipt.get(rfb.id, []) if rfb else []
        rule = rules.get(entry.regra_contabil_id) if entry.regra_contabil_id else None
        source = " ".join(str(value or "") for value in [entry.componente, entry.tributo, entry.codigo_receita, entry.descricao, entry.historico, movement.historico if movement else "", receipt.texto_original if receipt else "", rfb.tipo if rfb else "", rfb.razao_social if rfb else "", " ".join(f"{item.codigo} {item.descricao}" for item in rfb_items), rule.tipo_componente if rule else "", rule.historico if rule else ""])
        normalized = normalize_name(source)
        if "SIMPLES" in normalized:
            tax = "SIMPLES NACIONAL"
        elif entry.componente == "IRRF" or entry.codigo_receita.lstrip("0") in {"156", "561"} or "IRRF" in normalized:
            tax = "IRRF"
        elif entry.componente == "INSS" or "INSS" in normalized or "PREVIDENCI" in normalized:
            tax = "INSS"
        elif entry.componente == "FGTS" or "FGTS" in normalized:
            tax = "FGTS"
        else:
            return "", "", "", ""
        competence = extract_competence(rfb.texto_original, rfb.competencia, rfb.periodo_apuracao) if rfb else ""
        origin = "Comprovante RFB" if rfb else "Comprovante bancário" if receipt else ""
        return tax, competence, competence, origin
    for match in matches:
        movement, receipt, rfb = movements.get(match.movimento_extrato_id), receipts.get(match.comprovante_id) if match.comprovante_id else None, rfb_receipts.get(match.comprovante_rfb_id) if match.comprovante_rfb_id else None
        if not movement:
            continue
        used_in_period = movement_used_in_period(movement)
        entries = [] if not used_in_period else unique_accounting_entries(entries_by_match.get(match.id, []))
        total_lines = sum((statement_effect_value(line.valor, line.efeito_no_total) for line in entries), Decimal("0.00"))
        movement_date = movement.data.strftime("%d/%m/%Y")
        receipt_detail = "—"
        if receipt:
            beneficiary = receipt.beneficiario or receipt.favorecido
            final_beneficiary = f"\nBeneficiário final: {receipt.beneficiario_final}" if receipt.beneficiario_final and receipt.beneficiario_final != beneficiary else ""
            fantasy_name = f"\nNome fantasia: {receipt.nome_fantasia}" if receipt.nome_fantasia else ""
            receipt_detail = f"Data: {receipt.data.strftime('%d/%m/%Y')}\nTipo: {receipt.tipo_operacao or '—'}\nDocumento: {receipt.numero_documento or '—'}\nBeneficiário: {beneficiary}{final_beneficiary}{fantasy_name}\nValor pago: {money(receipt.valor_pago)}"
        rfb_detail = "—" if not rfb else f"Data: {rfb.data_arrecadacao.strftime('%d/%m/%Y') if rfb.data_arrecadacao else '—'}\nTipo: {rfb.tipo}\nCompetência: {rfb.competencia or rfb.periodo_apuracao or '—'}\nBanco: {rfb.nome_banco}\nTotal: {money(rfb.valor_total)}"
        receipt_composition = None
        if receipt:
            discount = (receipt.valor_desconto or 0) + (receipt.valor_abatimento or 0) + (receipt.valor_desconto_abatimento or 0)
            additions = (receipt.valor_juros or 0) + (receipt.valor_multa or 0) + (receipt.valor_encargos or 0)
            expected = (receipt.valor_original or receipt.valor_pago or 0) - discount + additions
            receipt_composition = {"valor_documento": money(receipt.valor_original), "valor_cobrado": money(receipt.valor_pago), "desconto": money(discount), "juros": money(receipt.valor_juros), "multa": money(receipt.valor_multa), "encargos": money(receipt.valor_encargos), "valor_calculado": money(expected), "diferenca": money((receipt.valor_pago or 0) - expected), "confere": abs((receipt.valor_pago or 0) - expected) <= Decimal("0.01")}
        rows.append({"id": match.id, "movimento_id": movement.id, "usado_no_periodo": used_in_period, "data": movement_date, "tipo_pagamento": payment_type(movement.historico), "natureza": normalize_statement_nature(movement.natureza), "natureza_contabil": accounting_nature(movement.natureza), "extrato": f"Data: {movement_date}\nTexto: {' '.join(item for item in [movement.historico, movement.nome_encontrado] if item)}\nValor: {money(movement.valor)}", "comprovante_bancario": receipt_detail, "comprovante_tipo": "emprestimo" if is_loan_receipt(receipt) else "comprovante" if receipt else "", "comprovante_rfb": rfb_detail, "comprovante_composicao": receipt_composition, "extrato_arquivo_id": movement.arquivo_id, "extrato_pagina": movement.pagina_numero, "comprovante_arquivo_id": receipt.arquivo_id if receipt else None, "comprovante_pagina": receipt.pagina_numero if receipt else None, "rfb_arquivo_id": rfb.arquivo_id if rfb else None, "rfb_pagina": rfb.pagina_numero if rfb else None, "valor": money(movement.valor), "fonte_regra": "ignorado" if not used_in_period else match.fonte_regra or "—", "total_lancamentos": money(total_lines), "diferenca": money(Decimal("0.00") if not used_in_period else movement.valor - total_lines), "confianca": "—" if not used_in_period else match.confianca, "situacao": "Não usado neste período" if not used_in_period else match.status})
    for row in rows:
        if row["usado_no_periodo"] is False:
            row["lancamentos"] = []
            continue
        items = unique_accounting_entries(entries_by_match.get(row["id"], []))
        match = matches_by_id[row["id"]]
        row["lancamentos"] = []
        for item in items:
            tax, competence, _, source = tax_complement_cached(item, match)
            row["lancamentos"].append({"id": item.id, "componente": item.componente, "categoria": item.categoria, "tributo": item.tributo, "codigo_receita": item.codigo_receita, "descricao": item.descricao, "efeito_no_total": item.efeito_no_total, "valor": money(item.valor), "conta_debito": item.conta_debito, "conta_credito": item.conta_credito, "historico": item.historico, "complemento": item.complemento, "imposto": tax, "competencia": competence, "competencia_nao_identificada": bool(tax and not competence), "comprovante_origem": source, "origem": item.origem, "status": item.status, "regra_contabil_id": item.regra_contabil_id})
    return rows


@router.put("/conciliacoes/{conciliacao_id}/movimentos-extrato/{movimento_id}/uso")
def set_statement_movement_usage(conciliacao_id: str, movimento_id: str, payload: MovimentoUsoInput, db: Session = Depends(get_db)):
    reconciliation = reconciliation_or_404(conciliacao_id, db)
    movement = db.get(MovimentoExtrato, movimento_id)
    if not movement or movement.conciliacao_id != reconciliation.id:
        raise HTTPException(404, "Lançamento do extrato não encontrado")
    movement.ignorado_no_periodo = not payload.usar
    match = db.query(Correspondencia).filter_by(conciliacao_id=conciliacao_id, movimento_extrato_id=movement.id).first()
    if not payload.usar:
        if not match:
            match = Correspondencia(conciliacao_id=conciliacao_id, movimento_extrato_id=movement.id)
            db.add(match)
            db.flush()
        match.regra_contabil_id = None
        match.fonte_regra = "ignorado"
        match.confianca = "—"
        match.criterio_correspondencia = "Lançamento marcado para não usar neste período"
        match.status = "Não usado neste período"
        db.commit()
    else:
        db.commit()
        reconcile(conciliacao_id, db)
    return {"id": movement.id, "usado_no_periodo": movement_used_in_period(movement), "resultado": result(conciliacao_id, db), "regras": accounting_rules(conciliacao_id, db)}


@router.put("/conciliacoes/{conciliacao_id}/correspondencias/{correspondencia_id}/lancamentos")
def save_accounting_items(conciliacao_id: str, correspondencia_id: str, payload: LancamentosInput, db: Session = Depends(get_db)):
    match = db.get(Correspondencia, correspondencia_id)
    if not match or match.conciliacao_id != conciliacao_id:
        raise HTTPException(404, "Conciliação não encontrada")
    movement = db.get(MovimentoExtrato, match.movimento_extrato_id)
    if not movement or not payload.itens or any(item.valor <= 0 or not all((item.conta_debito.strip(), item.conta_credito.strip(), item.historico.strip())) for item in payload.itens):
        raise HTTPException(422, "Todos os itens precisam de valor positivo, débito, crédito e histórico")
    if not movement_used_in_period(movement):
        raise HTTPException(422, "Este lançamento não está sendo usado neste período.")
    total = sum((statement_effect_value(item.valor, item.efeito_no_total) for item in payload.itens), Decimal("0.00"))
    existing = {item.id: item for item in db.query(LancamentoContabil).filter_by(correspondencia_id=match.id).all()}
    for order, item in enumerate(payload.itens, 1):
        entry = existing.get(item.id)
        is_new = entry is None
        if not entry:
            entry = LancamentoContabil(correspondencia_id=match.id)
            db.add(entry)
        entry.componente = item.componente.strip().upper()
        entry.categoria = entry.componente
        entry.tributo = item.tributo
        entry.codigo_receita = item.codigo_receita
        entry.descricao = item.descricao
        entry.valor = item.valor
        entry.efeito_no_total = item.efeito_no_total
        entry.conta_debito = item.conta_debito.strip()
        entry.conta_credito = item.conta_credito.strip()
        entry.historico = item.historico.strip()
        entry.complemento = item.complemento.strip()
        if is_new:
            entry.origem = "manual"
        entry.ordem = order
        entry.status = "editado_manual"
    db.flush()
    total = sum((statement_effect_value(entry.valor, entry.efeito_no_total) for entry in db.query(LancamentoContabil).filter_by(correspondencia_id=match.id)), Decimal("0.00"))
    difference = (movement.valor or Decimal("0.00")) - total
    match.status = "Conciliado manualmente" if abs(difference) <= Decimal("0.01") else "Lançamentos pendentes de conferência"
    db.commit()
    return {"id": match.id, "itens": len(payload.itens), "total_contabil": str(total), "valor_extrato": str(movement.valor or 0), "diferenca": str(difference), "status": match.status}


@router.get("/conciliacoes/{conciliacao_id}/documentos-nao-utilizados")
def unused_documents(conciliacao_id: str, db: Session = Depends(get_db), response: FastAPIResponse = None):
    if response:
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    reconciliation = db.get(Conciliacao, conciliacao_id)
    matches = db.query(Correspondencia).filter_by(conciliacao_id=conciliacao_id).all()
    movement_ids = [item.movimento_extrato_id for item in matches]
    movements_by_id = {item.id: item for item in db.query(MovimentoExtrato).filter(MovimentoExtrato.id.in_(movement_ids)).all()} if movement_ids else {}
    usable_matches = [item for item in matches if movement_used_in_period(movements_by_id.get(item.movimento_extrato_id))]
    used_receipts = {item.comprovante_id for item in usable_matches if item.comprovante_id}
    used_rfb = {item.comprovante_rfb_id for item in usable_matches if item.comprovante_rfb_id}
    def display_date(value): return value.strftime("%d/%m/%Y") if value else "—"
    def in_period(value): return bool(value and reconciliation.data_inicio <= value <= reconciliation.data_fim)
    files = {item.id: item for item in db.query(Arquivo).filter_by(conciliacao_id=conciliacao_id).all()}
    receipts = db.query(Comprovante).filter_by(conciliacao_id=conciliacao_id, ativo=True).all()
    bank_receipts = [item for item in receipts if files.get(item.arquivo_id) and files[item.arquivo_id].tipo_documento == "comprovante"]
    loan_receipts = [item for item in receipts if files.get(item.arquivo_id) and files[item.arquivo_id].tipo_documento == LOAN_DOCUMENT_TYPE]
    rfb_receipts = [item for item in db.query(ComprovanteRfb).filter_by(conciliacao_id=conciliacao_id) if belongs_to_selected_bank(item, reconciliation.banco)]
    unused_receipts = [{"id": item.id, "data": display_date(item.data), "hora": item.hora or "—", "documento": item.numero_documento, "favorecido": item.beneficiario_final or item.beneficiario or item.favorecido, "valor_pago": str(item.valor_pago), "tipo": item.tipo_operacao, "situacao": "Sem movimento no extrato" if in_period(item.data) else "Fora do período"} for item in bank_receipts if item.id not in used_receipts]
    unused_loans = [{"id": item.id, "data": display_date(item.data), "documento": item.numero_documento, "favorecido": item.beneficiario_final or item.beneficiario or item.favorecido, "valor_pago": str(item.valor_pago), "tipo": item.tipo_operacao, "situacao": "Sem movimento no extrato" if in_period(item.data) else "Fora do período"} for item in loan_receipts if item.id not in used_receipts]
    unused_rfb = [{"id": item.id, "tipo": item.tipo, "data_arrecadacao": display_date(item.data_arrecadacao), "documento": item.numero_documento, "banco": item.nome_banco, "total": str(item.valor_total), "situacao": "Sem movimento no extrato" if in_period(item.data_arrecadacao) else "Fora do período"} for item in rfb_receipts if item.id not in used_rfb]
    def summary(items, used, date_field): return {"total": len(items), "utilizados": sum(item.id in used for item in items), "nao_utilizados": sum(item.id not in used for item in items), "fora_periodo": sum(not in_period(getattr(item, date_field)) for item in items)}
    return {"comprovantes": unused_receipts, "emprestimos": unused_loans, "rfb": unused_rfb, "resumo": {"comprovantes": summary(bank_receipts, used_receipts, "data"), "emprestimos": summary(loan_receipts, used_receipts, "data"), "rfb": summary(rfb_receipts, used_rfb, "data_arrecadacao")}}


@router.get("/conciliacoes/{conciliacao_id}/revisao")
def review(conciliacao_id: str, db: Session = Depends(get_db), response: FastAPIResponse = None):
    if response:
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    def base(record, fields):
        return {"id": record.id, "arquivo_id": record.arquivo_id, "pagina": record.pagina_numero, "revisao": getattr(record, "status_revisao", record.status if hasattr(record, "status") else ""), **fields}

    def display_date(value):
        return value.strftime("%d/%m/%Y") if value else ""
    def money(value):
        if value is None:
            return "—"
        integer, decimal = f"{value:.2f}".split(".")
        return f"R$ {int(integer):,}".replace(",", ".") + f",{decimal}"

    def display_metadata_date(value):
        try:
            return date.fromisoformat(value).strftime("%d/%m/%Y") if value else "—"
        except ValueError:
            return value or "—"

    def statement_balances():
        if not reconciliation or reconciliation.banco != "BASA":
            return {}
        text = "\n".join(item.texto_bruto or "" for item in files.values() if item.tipo_documento == "extrato")
        initial_balance = extract_basa_initial_balance(text)
        return {"saldo_anterior": money(initial_balance)} if initial_balance is not None else {}

    def adjustments(item):
        fields = (("Desconto", item.valor_desconto), ("Abatimento", item.valor_abatimento), ("Desconto/abatimento", item.valor_desconto_abatimento), ("Taxa/tarifa", item.valor_tarifa), ("Juros", item.valor_juros), ("Multa", item.valor_multa), ("Encargos", item.valor_encargos))
        values = [f"{label}: {money(value)}" for label, value in fields if value is not None and value > 0]
        suffix = " Composição de valor divergente." if item.status_revisao == "revisao" else ""
        return (" + ".join(values) if values else "Sem ajustes") + suffix

    def receipt_counterparty(item):
        beneficiary = item.beneficiario or item.favorecido
        if item.beneficiario_final and item.beneficiario_final != beneficiary:
            return f"Beneficiário: {beneficiary}\nBeneficiário final: {item.beneficiario_final}"
        return f"Beneficiário: {beneficiary}"

    reconciliation = db.get(Conciliacao, conciliacao_id)
    getnet_adjustments = sync_getnet_anticipation_adjustments(reconciliation, db) if reconciliation else []
    db.commit()
    rfb_records = [item for item in db.query(ComprovanteRfb).filter_by(conciliacao_id=conciliacao_id).order_by(ComprovanteRfb.data_arrecadacao.asc(), ComprovanteRfb.id.asc()) if reconciliation and belongs_to_selected_bank(item, reconciliation.banco)]
    files = {item.id: item for item in db.query(Arquivo).filter_by(conciliacao_id=conciliacao_id, ativo=True).all()}

    statement_records = db.query(MovimentoExtrato).filter_by(conciliacao_id=conciliacao_id, ativo=True).all()
    statement_records.sort(key=movement_statement_order)
    receipt_records = db.query(Comprovante).filter_by(conciliacao_id=conciliacao_id, ativo=True).order_by(Comprovante.data.asc(), Comprovante.hora.asc().nulls_last(), Comprovante.id.asc()).all()
    receipt_payloads = [base(x, {"data": display_date(x.data), "hora": x.hora, "documento": x.numero_documento, "favorecido": receipt_counterparty(x), "valor_original": money(x.valor_original), "ajustes": adjustments(x), "valor_pago": money(x.valor_pago), "tipo": x.tipo_operacao}) for x in receipt_records]
    receipt_type_by_id = {item.id: files[item.arquivo_id].tipo_documento for item in receipt_records if item.arquivo_id in files}
    invoice_records = db.query(NotaFiscal).filter_by(conciliacao_id=conciliacao_id, ativo=True).order_by(NotaFiscal.data_emissao.asc().nulls_last(), NotaFiscal.id.asc()).all()

    return {
        "extratos": [base(x, {"data": display_date(x.data), "hora": x.hora, "historico": " ".join(item for item in [x.historico, x.nome_encontrado] if item), "valor": str(x.valor), "natureza": normalize_statement_nature(x.natureza), "natureza_contabil": accounting_nature(x.natureza), "usado_no_periodo": movement_used_in_period(x)}) for x in statement_records],
        "comprovantes": [item for item in receipt_payloads if receipt_type_by_id.get(item["id"]) == "comprovante"],
        "maquininhas": [item for item in receipt_payloads if receipt_type_by_id.get(item["id"]) in MACHINE_STATEMENT_DOCUMENT_TYPES],
        "emprestimos": [item for item in receipt_payloads if receipt_type_by_id.get(item["id"]) == LOAN_DOCUMENT_TYPE],
        "notas": [base(x, {"data_emissao": display_date(x.data_emissao), "fornecedor": x.fornecedor or "—", "cpf_cnpj": x.cpf_cnpj or "—", "numero_nota": x.numero_nota or "—", "forma_pagamento": (x.dados_originais or {}).get("forma_pagamento", "—") if isinstance(x.dados_originais, dict) else "—", "data_pagamento": display_metadata_date((x.dados_originais or {}).get("data_pagamento", "")) if isinstance(x.dados_originais, dict) else "—", "valor_total": money(x.valor_total), "situacao": x.status_revisao}) for x in invoice_records],
        "rfb": [base(x, {"tipo": x.tipo, "competencia_apuracao": x.competencia or x.periodo_apuracao or "—", "data_arrecadacao": display_date(x.data_arrecadacao), "documento": x.numero_documento, "banco": x.nome_banco, "principal": money(x.valor_principal), "multa_juros": "Sem acréscimos" if not ((x.valor_multa or 0) + (x.valor_juros or 0)) else f"Multa: {money(x.valor_multa)} + Juros: {money(x.valor_juros)}", "total": money(x.valor_total), "situacao": x.status}) for x in rfb_records],
        "arquivos": [{"id": x.id, "nome": x.nome_original, "tipo": x.tipo_documento, "status": x.status_processamento, "erro": x.mensagem_erro} for x in files.values()],
        "ajustes_getnet": getnet_adjustments,
        "saldos": statement_balances(),
    }
