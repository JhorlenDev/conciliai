import csv
import re
import logging
import calendar
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from app.services.normalization import normalize_name

logger = logging.getLogger(__name__)

DATE_RE = r"(\d{2}[/.]\d{2}[/.]\d{4})"
TIME_RE = r"(\d{2}:\d{2}(?::\d{2})?)"
NAME_LABELS = ("BENEFICIÁRIO", "BENEFICIARIO", "FAVORECIDO", "PAGO PARA", "BENEFICIÁRIO FINAL", "BENEFICIARIO FINAL", "CONVÊNIO", "CONVENIO")
PT_MONTHS = {
    "janeiro": 1,
    "fevereiro": 2,
    "marco": 3,
    "março": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
}


@dataclass
class FinancialValues:
    valor_original: Decimal | None = None
    valor_desconto: Decimal | None = None
    valor_abatimento: Decimal | None = None
    valor_desconto_abatimento: Decimal | None = None
    valor_juros: Decimal | None = None
    valor_multa: Decimal | None = None
    valor_encargos: Decimal | None = None
    valor_tarifa: Decimal | None = None
    valor_pago: Decimal | None = None
    detalhes: dict[str, str] = field(default_factory=dict)
    composicao_divergente: bool = False


@dataclass
class ParsedReceipt:
    data: date | None
    hora: str | None
    favorecido: str
    valor: Decimal | None
    tipo_operacao: str
    texto_original: str
    pagina_numero: int
    origem_nome: str = ""
    financeiros: FinancialValues = field(default_factory=FinancialValues)
    beneficiario: str = ""
    nome_fantasia: str = ""
    beneficiario_final: str = ""
    pagador: str = ""
    cnpj_beneficiario: str = ""
    cnpj_beneficiario_final: str = ""
    numero_documento: str = ""


@dataclass
class ParsedStatement:
    data: date | None
    hora: str | None
    historico: str
    nome: str
    valor: Decimal | None
    natureza: str
    texto_original: str
    pagina_numero: int
    data_origem: str = ""
    numero_documento: str = ""


@dataclass
class PdfStatementExtraction:
    pages: list[str]
    records: list[ParsedStatement]


@dataclass
class ParsedInvoice:
    data_emissao: date | None
    fornecedor: str
    cpf_cnpj: str
    numero_nota: str
    valor_total: Decimal | None
    texto_original: str
    pagina_numero: int
    dados: dict = field(default_factory=dict)


SANTANDER_IGNORE_AUTOMATIC_INVESTMENT_MOVEMENTS = False
SANTANDER_AUTOMATIC_INVESTMENT_RE = re.compile(r"\b(?:APLICACAO|APLICAÇÃO|RESGATE)\s+CONTAMAX\b", re.I)
SANTANDER_EXPECTED_RECORDS = 98
SANTANDER_EXPECTED_CREDIT = Decimal("47224.73")
SANTANDER_EXPECTED_DEBIT = Decimal("47224.73")
SANTANDER_EXPECTED_INITIAL_BALANCE = Decimal("0.00")
SANTANDER_EXPECTED_FINAL_BALANCE = Decimal("0.00")


def parse_brl(value: str) -> Decimal | None:
    cleaned = re.sub(r"[^0-9,.-]", "", value).replace(".", "").replace(",", ".")
    try:
        return Decimal(cleaned).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def parse_date_time(value: str) -> tuple[date | None, str | None]:
    date_match = re.search(DATE_RE, value)
    time_match = re.search(TIME_RE, value)
    parsed_date = date(*map(int, reversed(re.split(r"[/.]", date_match.group(1))))) if date_match else None
    return parsed_date, time_match.group(1) if time_match else None


def amount_text(value: str) -> bool:
    return bool(re.fullmatch(r"-?\s*(?:R\$\s*)?\d{1,3}(?:\.\d{3})*,\d{2}\s*[CD]?", value.strip(), re.I))


def receipt_label_value(text: str, *labels: str, value_prefix: str = r".") -> str | None:
    """Read BB-style labels whose value can be on the same line or the next line."""
    for label in labels:
        pattern = rf"^\s*{re.escape(label)}(?:[ \t]*:[ \t]*|[ \t]+|[ \t]*\n[ \t]*)(?={value_prefix})([^\n]+)"
        match = re.search(pattern, text, re.I | re.M)
        if match and match.group(1).strip():
            return match.group(1).strip()
    return None


def extract_financial_values(text: str) -> FinancialValues:
    def amount_for(*labels: str) -> Decimal | None:
        for label in labels:
            value = receipt_label_value(text, label, value_prefix=r"(?:R?\$?\s*[\d.,-])")
            if value:
                amount = parse_brl(value)
                if amount is not None:
                    return amount
        return None

    values = FinancialValues(
        valor_original=amount_for("VALOR DO DOCUMENTO", "VALOR ORIGINAL"),
        valor_desconto=amount_for("DESCONTO"),
        valor_abatimento=amount_for("ABATIMENTO"),
        valor_desconto_abatimento=amount_for("DESCONTO/ABATIMENTO"),
        valor_juros=amount_for("JUROS/MORA", "JUROS"),
        valor_multa=amount_for("MULTA/MORA", "MULTA"),
        valor_encargos=amount_for("ENCARGOS"),
        valor_tarifa=amount_for("TARIFA"),
        valor_pago=amount_for("VALOR COBRADO", "VALOR PAGO", "VALOR RECEBIDO", "VALOR TRANSFERIDO", "VALOR DA TRANSFERENCIA", "VALOR DA TRANSFERÊNCIA", "VALOR TOTAL", "VALOR DO DEBITO", "VALOR DO DÉBITO", "VALOR"),
    )
    if values.valor_original is None:
        values.valor_original = values.valor_pago
    adjustments = [
        values.valor_desconto, values.valor_abatimento, values.valor_desconto_abatimento,
        values.valor_juros, values.valor_multa, values.valor_encargos,
    ]
    if values.valor_pago is None and values.valor_original is not None and not any(value is not None for value in adjustments):
        values.valor_pago = values.valor_original
    for key in (
        "valor_original", "valor_desconto", "valor_abatimento", "valor_desconto_abatimento",
        "valor_juros", "valor_multa", "valor_encargos", "valor_tarifa", "valor_pago",
    ):
        amount = getattr(values, key)
        if amount is not None:
            values.detalhes[key] = str(amount)
    if values.valor_original is not None and values.valor_pago is not None and any(value is not None for value in adjustments):
        expected = values.valor_original - (values.valor_desconto or Decimal()) - (values.valor_abatimento or Decimal()) - (values.valor_desconto_abatimento or Decimal()) + (values.valor_juros or Decimal()) + (values.valor_multa or Decimal()) + (values.valor_encargos or Decimal())
        values.composicao_divergente = abs(expected - values.valor_pago) > Decimal("0.01")
    return values


def find_receipt_name(text: str) -> str | None:
    found = find_receipt_name_with_origin(text)
    return found[0] if found else None


def find_receipt_name_with_origin(text: str) -> tuple[str, str] | None:
    """Returns the named counterparty and label using the required precedence."""
    for label in NAME_LABELS:
        match = re.search(rf"^\s*{re.escape(label)}\s*:?[ \t]*(?:\n[ \t]*)?([^\n]+)", text, re.I | re.M)
        if not match:
            continue
        name = match.group(1).strip()
        if label in {"BENEFICIÁRIO", "BENEFICIARIO"} and name.upper().startswith("FINAL"):
            continue
        if name and name.upper() not in {"PIX", "TED", "DOCUMENTO", "BANCO", "PAGADOR"}:
            return name, label
    return None


def label_value_match(text: str, label: str) -> re.Match[str] | None:
    suffix = r"(?!\s+FINAL\b)" if label in {"BENEFICIÁRIO", "BENEFICIARIO", "CNPJ BENEFICIÁRIO", "CNPJ BENEFICIARIO"} else ""
    return re.search(rf"^\s*{re.escape(label)}{suffix}\s*:?[ \t]*(?:\n[ \t]*)?([^\n]+)", text, re.I | re.M)


def cnpj_after_section(text: str, labels: tuple[str, ...]) -> str:
    for label in labels:
        suffix = r"(?!\s+FINAL\b)" if label in {"BENEFICIÁRIO", "BENEFICIARIO"} else ""
        stop_labels = "BENEFICI[ÁA]RIO FINAL|PAGADOR|NR\\. DOCUMENTO|NOSSO NUMERO|CONVENIO|DATA DE VENCIMENTO|DATA DO PAGAMENTO|VALOR"
        if "FINAL" in label:
            stop_labels = "PAGADOR|NR\\. DOCUMENTO|NOSSO NUMERO|CONVENIO|DATA DE VENCIMENTO|DATA DO PAGAMENTO|VALOR"
        section = re.search(rf"^\s*{re.escape(label)}{suffix}\s*:?[ \t]*(.*?)(?=^\s*(?:{stop_labels})\b|\Z)", text, re.I | re.M | re.S)
        if not section:
            continue
        match = re.search(r"\bCNPJ\s*:?[ \t]*(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})", section.group(1), re.I)
        if match:
            return match.group(1).strip()
    return ""


def receipt_participants(text: str) -> dict[str, str]:
    labels = {"beneficiario_final": ("BENEFICIÁRIO FINAL", "BENEFICIARIO FINAL"), "beneficiario": ("BENEFICIÁRIO", "BENEFICIARIO", "FAVORECIDO"), "nome_fantasia": ("NOME FANTASIA",), "pagador": ("PAGADOR", "PAGO POR"), "cnpj_beneficiario_final": ("CNPJ BENEFICIÁRIO FINAL", "CNPJ BENEFICIARIO FINAL"), "cnpj_beneficiario": ("CNPJ BENEFICIÁRIO", "CNPJ BENEFICIARIO")}
    participants = {}
    for key, names in labels.items():
        for label in names:
            match = label_value_match(text, label)
            if match and match.group(1).strip():
                participants[key] = match.group(1).strip()
                break
    participants.setdefault("cnpj_beneficiario", cnpj_after_section(text, ("BENEFICIÁRIO", "BENEFICIARIO", "FAVORECIDO")))
    participants.setdefault("cnpj_beneficiario_final", cnpj_after_section(text, ("BENEFICIÁRIO FINAL", "BENEFICIARIO FINAL")))
    return participants


def receipt_document_number(text: str) -> str:
    """Reads the document identifier printed by bank receipt layouts."""
    labels = (
        "NR. DOCUMENTO",
        "NR DOCUMENTO",
        "Nº DOCUMENTO",
        "N° DOCUMENTO",
        "NO DOCUMENTO",
        "NUMERO DOCUMENTO",
        "NÚMERO DOCUMENTO",
        "NUMERO DO DOCUMENTO",
        "NÚMERO DO DOCUMENTO",
        "DOCUMENTO",
    )
    for label in labels:
        value = receipt_label_value(text, label)
        if value:
            value = " ".join(value.split())
            if value and value.upper() not in {"PIX", "TED", "DOCUMENTO"}:
                return value
    return ""


def _clean_receipt_amount(value: str | None) -> Decimal | None:
    if not value:
        return None
    return parse_brl(value.replace("R$", "").strip())


def _lines(text: str) -> list[str]:
    return [" ".join(line.split()) for line in text.splitlines() if line.strip()]


def _line_is_label(value: str) -> bool:
    return value.strip().endswith(":")


def _line_value_near(lines: list[str], label: str, prefer_before: bool = False) -> str:
    normalized_label = normalize_name(label)
    for index, line in enumerate(lines):
        normalized_line = normalize_name(line)
        if normalized_label not in normalized_line:
            continue
        same_line = re.split(r":", line, maxsplit=1)
        if len(same_line) > 1 and same_line[1].strip():
            return same_line[1].strip()
        offsets = (-1, 1, -2, 2) if prefer_before else (1, -1, 2, -2)
        for offset in offsets:
            candidate_index = index + offset
            if 0 <= candidate_index < len(lines):
                candidate = lines[candidate_index].strip()
                if candidate and normalize_name(candidate) != normalized_label and not _line_is_label(candidate):
                    return candidate
    return ""


def _value_after_section(lines: list[str], section: str, label: str, limit: int = 12) -> str:
    normalized_section = normalize_name(section)
    normalized_label = normalize_name(label)
    for section_index, line in enumerate(lines):
        if normalized_section not in normalize_name(line):
            continue
        end = min(len(lines), section_index + limit + 1)
        for index in range(section_index + 1, end):
            normalized_line = normalize_name(lines[index])
            if normalized_label not in normalized_line:
                continue
            same_line = re.split(r":", lines[index], maxsplit=1)
            if len(same_line) > 1 and same_line[1].strip():
                return same_line[1].strip()
            for candidate in lines[index + 1:end]:
                candidate = candidate.strip()
                if candidate and not _line_is_label(candidate):
                    return candidate
    return ""


def _previous_value_for_label_context(lines: list[str], label: str, required_next: str = "", rejected_next: tuple[str, ...] = ()) -> str:
    normalized_label = normalize_name(label)
    normalized_required = normalize_name(required_next)
    normalized_rejected = tuple(normalize_name(item) for item in rejected_next)
    for index, line in enumerate(lines):
        normalized_line = normalize_name(line)
        if normalized_line != normalized_label:
            continue
        next_line = normalize_name(lines[index + 1]) if index + 1 < len(lines) else ""
        if normalized_required and normalized_required not in next_line:
            continue
        if any(item in next_line for item in normalized_rejected):
            continue
        for candidate_index in range(index - 1, max(-1, index - 4), -1):
            candidate = lines[candidate_index].strip()
            if candidate and not _line_is_label(candidate):
                return candidate
    return ""


def _bradesco_original_amount(lines: list[str]) -> str:
    for index, line in enumerate(lines):
        if normalize_name(line) != "VALOR":
            continue
        for candidate_index in range(index - 1, max(-1, index - 4), -1):
            candidate = lines[candidate_index].strip()
            if candidate and amount_text(candidate):
                return candidate
    return ""


def _receipt_financial(original: Decimal | None, paid: Decimal | None, text: str) -> FinancialValues:
    financial = extract_financial_values(text)
    if original is not None:
        financial.valor_original = original
    if paid is not None:
        financial.valor_pago = paid
    for key in ("valor_original", "valor_pago"):
        amount = getattr(financial, key)
        if amount is not None:
            financial.detalhes[key] = str(amount)
    return financial


def _extract_caixa_receipts(text: str, page_number: int) -> list[ParsedReceipt]:
    normalized = normalize_name(text)
    if "COMPROVANTE PAGAMENTO BOLETO" not in normalized or "INTERNET BANKING CAIXA" not in normalized:
        return []
    lines = _lines(text)
    paid_date, paid_time = parse_date_time(_line_value_near(lines, "Data/hora da operação") or _line_value_near(lines, "Data de Efetivação / Agendamento"))
    beneficiary = (
        _value_after_section(lines, "Beneficiário original / Cedente", "Nome/Razão Social")
        or _value_after_section(lines, "Beneficiário original / Cedente", "Nome Fantasia")
        or _line_value_near(lines, "Nome Fantasia")
        or _line_value_near(lines, "Nome/Razão Social")
    )
    fantasy = _value_after_section(lines, "Beneficiário original / Cedente", "Nome Fantasia")
    cnpj = _value_after_section(lines, "Beneficiário original / Cedente", "CPF/CNPJ") or _line_value_near(lines, "CPF/CNPJ")
    document = _line_value_near(lines, "Código da operação") or receipt_document_number(text)
    original = _clean_receipt_amount(_line_value_near(lines, "Valor Nominal do Boleto"))
    paid = _clean_receipt_amount(_line_value_near(lines, "Valor Pago")) or _clean_receipt_amount(_line_value_near(lines, "Valor Calculado"))
    if not paid_date or paid is None or not beneficiary:
        return []
    financial = _receipt_financial(original, paid, text)
    return [ParsedReceipt(
        data=paid_date,
        hora=paid_time,
        favorecido=beneficiary,
        valor=paid,
        tipo_operacao="BOLETO",
        texto_original=text.strip(),
        pagina_numero=page_number,
        origem_nome="CAIXA BOLETO",
        financeiros=financial,
        beneficiario=beneficiary,
        nome_fantasia=fantasy,
        pagador=_line_value_near(lines, "Pagador Final - Correntista") or _line_value_near(lines, "Nome"),
        cnpj_beneficiario=cnpj,
        numero_documento=document,
    )]


def _extract_bradesco_receipts(text: str, page_number: int) -> list[ParsedReceipt]:
    normalized = normalize_name(text)
    if "COMPROVANTE TRANSACAO BANCARIA" not in normalized or "BOLETO COBRANCA" not in normalized:
        return []
    blocks = [text.strip()]
    results: list[ParsedReceipt] = []
    for block in blocks:
        lines = _lines(block)
        date_value = _line_value_near(lines, "Data da operação") or _line_value_near(lines, "Data de débito")
        paid_date, paid_time = parse_date_time(date_value)
        document = _line_value_near(lines, "Documento", prefer_before=True)
        beneficiary = (
            _previous_value_for_label_context(lines, "Razão Social", "Beneficiário", ("Final",))
            or _previous_value_for_label_context(lines, "Nome Fantasia", "Beneficiário", ("Final",))
        )
        fantasy = _previous_value_for_label_context(lines, "Nome Fantasia", "Beneficiário", ("Final",))
        cnpj = _previous_value_for_label_context(lines, "CPF/CNPJ Beneficiário", "", ("Final",))
        original = _clean_receipt_amount(_bradesco_original_amount(lines))
        paid = _clean_receipt_amount(_line_value_near(lines, "Valor total", prefer_before=True)) or original
        if not paid_date or paid is None or not beneficiary or normalize_name(beneficiary) == "NAO INFORMADO":
            continue
        financial = _receipt_financial(original, paid, block)
        results.append(ParsedReceipt(
            data=paid_date,
            hora=paid_time,
            favorecido=beneficiary,
            valor=paid,
            tipo_operacao="BOLETO",
            texto_original=block,
            pagina_numero=page_number,
            origem_nome="BRADESCO BOLETO",
            financeiros=financial,
            beneficiario=beneficiary,
            nome_fantasia=fantasy,
            pagador=_line_value_near(lines, "Nome do Pagador", prefer_before=True),
            cnpj_beneficiario=cnpj,
            numero_documento=document,
        ))
    return results


def extract_receipts(text: str, page_number: int) -> list[ParsedReceipt]:
    getnet_sales = _extract_getnet_sales_receipts(text, page_number)
    if getnet_sales:
        return getnet_sales
    santander = _extract_santander_receipts(text, page_number)
    if santander:
        return santander
    caixa = _extract_caixa_receipts(text, page_number)
    if caixa:
        return caixa
    bradesco = _extract_bradesco_receipts(text, page_number)
    if bradesco:
        return bradesco
    page_blocks = split_banco_do_brasil_receipt_blocks(text)
    if len(page_blocks) > 1:
        results = []
        for block in page_blocks:
            results.extend(extract_receipt_block(block, page_number))
        return results
    return extract_receipt_block(text, page_number)


def extract_loan_receipts(text: str, page_number: int) -> list[ParsedReceipt]:
    normalized = normalize_name(text)
    if not any(term in normalized for term in ("EMPRESTIMO", "FINANCIAMENTO", "CONTRATO", "OPERACAO")):
        return []

    scheduled = _extract_banco_do_brasil_loan_schedule(text, page_number)
    if scheduled:
        return scheduled

    def label_text(*labels: str) -> str:
        for label in labels:
            value = receipt_label_value(text, label, value_prefix=r"[^\n]")
            if value:
                return " ".join(value.split())
        return ""

    def label_amount(*labels: str) -> Decimal | None:
        for label in labels:
            value = receipt_label_value(text, label, value_prefix=r"(?:R?\$?\s*[\d.,-])")
            if value:
                amount = parse_brl(value)
                if amount is not None:
                    return abs(amount)
        return None

    contract = label_text(
        "NR. CONTRATO",
        "NR CONTRATO",
        "Nº CONTRATO",
        "NUMERO DO CONTRATO",
        "NÚMERO DO CONTRATO",
        "CONTRATO",
        "NR. OPERACAO",
        "NR OPERAÇÃO",
        "Nº OPERAÇÃO",
        "OPERACAO",
        "OPERAÇÃO",
    )
    if not contract:
        match = re.search(r"\b\d{2,3}[.\s]\d{3}[.\s]\d{3}(?:[.\s]\d{3}[.\s]\d{3})?\b", text)
        contract = " ".join(match.group(0).replace(" ", ".").split()) if match else ""

    parsed_date, parsed_time = parse_date_time(
        "\n".join(
            value
            for value in [
                label_text("DATA DO PAGAMENTO", "DATA DO DÉBITO", "DATA DO DEBITO", "DATA DA PARCELA", "DATA"),
                text,
            ]
            if value
        )
    )
    principal = label_amount("VALOR PRINCIPAL", "PRINCIPAL", "CAPITAL", "AMORTIZACAO", "AMORTIZAÇÃO")
    interest = label_amount("JUROS", "VALOR JUROS", "JUROS DO PERIODO", "JUROS DO PERÍODO")
    iof = label_amount("IOF", "VALOR IOF")
    charges = label_amount("ENCARGOS", "TARIFAS", "TARIFA")
    paid = label_amount("VALOR TOTAL", "TOTAL", "VALOR DA PARCELA", "VALOR DO DÉBITO", "VALOR DO DEBITO", "VALOR PAGO", "VALOR COBRADO")
    component_total = sum((value or Decimal("0.00") for value in (principal, interest, iof, charges)), Decimal("0.00"))
    if paid is None and component_total > 0:
        paid = component_total
    if paid is None or paid <= 0:
        return []

    details = {}
    for key, value in (("contrato", contract), ("principal", principal), ("juros", interest), ("iof", iof), ("encargos", charges), ("valor_pago", paid)):
        if value:
            details[key] = str(value)
    financial = FinancialValues(
        valor_original=principal or paid,
        valor_juros=interest,
        valor_encargos=(iof or Decimal("0.00")) + (charges or Decimal("0.00")) or None,
        valor_pago=paid,
        detalhes=details,
        composicao_divergente=bool(component_total > 0 and abs(component_total - paid) > Decimal("0.01")),
    )
    beneficiary = " ".join(part for part in ["Empréstimo/Financiamento", contract] if part).strip()
    return [
        ParsedReceipt(
            data=parsed_date,
            hora=parsed_time,
            favorecido=beneficiary or "Empréstimo/Financiamento",
            valor=paid,
            tipo_operacao="EMPRESTIMO",
            texto_original=text,
            pagina_numero=page_number,
            origem_nome="CONTRATO",
            financeiros=financial,
            beneficiario=beneficiary or "Empréstimo/Financiamento",
            numero_documento=contract,
        )
    ]


def extract_loan_spreadsheet(path: str | Path) -> tuple[list[str], list[ParsedReceipt]]:
    path = Path(path)
    pages: list[str] = []
    receipts: list[ParsedReceipt] = []

    def as_decimal(value) -> Decimal | None:
        if value is None or value == "":
            return None
        if isinstance(value, Decimal):
            return value.quantize(Decimal("0.01"))
        if isinstance(value, (int, float)):
            return Decimal(str(value)).quantize(Decimal("0.01"))
        text = str(value).strip()
        if not text:
            return None
        if "," in text:
            return parse_brl(text)
        cleaned = re.sub(r"[^0-9.-]", "", text)
        if not cleaned:
            return None
        try:
            return Decimal(cleaned).quantize(Decimal("0.01"))
        except (InvalidOperation, ValueError):
            return None

    def as_date(value) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if value:
            parsed, _ = parse_date_time(str(value))
            return parsed
        return None

    if path.suffix.lower() == ".csv":
        raw = path.read_bytes()
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("latin-1")
        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t,")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"
        sources = [("CSV", [(list(row), [False] * len(row)) for row in csv.reader(text.splitlines(), dialect)])]
    else:
        workbook = load_workbook(path, data_only=True)
        sources = []
        for worksheet in workbook.worksheets:
            rows = []
            for row in worksheet.iter_rows():
                values = [cell.value for cell in row]
                highlighted = []
                for cell in row:
                    color = cell.font.color.rgb if cell.font and cell.font.color and cell.font.color.type == "rgb" else ""
                    highlighted.append(str(color).upper() in {"FFFF0000", "FFC00000"})
                rows.append((values, highlighted))
            sources.append((worksheet.title, rows))

    for page_number, (source_name, source_rows) in enumerate(sources, 1):
        lines: list[str] = []
        highlighted_lines: list[str] = []
        lenders: list[str] = []
        contract = ""
        title = ""
        credit_line = ""
        header_indexes: dict[str, int] = {}
        grouped: dict[tuple[date, str], dict[str, Decimal | list[str] | str | set[str]]] = {}
        for row, highlighted in source_rows:
            values = ["" if value is None else value for value in row]
            line = " | ".join(str(value) for value in values if value != "")
            if line:
                lines.append(line)
                if any(highlighted[: len(values)]):
                    highlighted_lines.append(line)
                if not title and "CONTROLE" in normalize_name(line):
                    title = line
            if not contract:
                match = re.search(r"\b\d{2,3}[.\s]\d{3}[.\s]\d{3}\b", line)
                if match:
                    contract = re.sub(r"\s+", "", match.group(0).strip(" ."))
            if "Linha de Crédito" in line or "Linha de Credito" in line:
                credit_line = " ".join(str(value) for value in values[1:] if value).strip()
            normalized_line = normalize_name(line)
            if header_indexes and "PARCELA" not in normalized_line and not line.startswith("202"):
                pass
            elif not header_indexes and line and "FINANCIADOR" not in normalized_line and "CONTROLE" not in normalized_line and "LINHA CREDITO" not in normalized_line and len(lenders) < 4:
                if any(word in normalized_line for word in ("AYMORE", "SANTANDER", "BANCO", "CREDITO FINANC")):
                    lenders.append(line)
            normalized_cells = [normalize_name(str(value)) for value in values]
            if "VENCIMENTO" in normalized_cells and "TIPO" in normalized_cells:
                header_indexes = {name: index for index, name in enumerate(normalized_cells)}
                continue
            if not header_indexes:
                continue
            row_date = as_date(values[header_indexes.get("VENCIMENTO", 0)])
            component = normalize_name(str(values[header_indexes.get("TIPO", 1)]))
            if not row_date or component not in {"JUROS", "CAPITAL", "AMORTIZACAO", "PRINCIPAL"}:
                continue
            amount_index = next((header_indexes[key] for key in ("VALOR PREVISTO R", "VALOR R", "PREVISTO R") if key in header_indexes), None)
            principal_index = next((header_indexes[key] for key in ("CAPITAL DA PARCELA R", "VALOR CAPITAL R", "CAPITAL R") if key in header_indexes), None)
            status_index = header_indexes.get("SITUACAO")
            debit_index = header_indexes.get("DEBITO")
            credit_index = header_indexes.get("CREDITO")
            history_index = header_indexes.get("HISTORICO")
            amount = as_decimal(values[amount_index]) if amount_index is not None and amount_index < len(values) else None
            principal_amount = as_decimal(values[principal_index]) if principal_index is not None and principal_index < len(values) else None
            if amount is None and principal_amount is None:
                continue
            status = str(values[status_index]).strip() if status_index is not None and status_index < len(values) else ""
            normalized_status = normalize_name(status)
            parcel_match = re.search(r"\bPARCELA\s+(\d+)\s+(\d+)\b", normalized_status)
            parcel = f"{parcel_match.group(1)}/{parcel_match.group(2)}" if parcel_match else ""
            bank = next((name for name in ("Santander", "Banco do Brasil", "BB", "Aymoré") if name.upper() in status.upper()), "")
            group_key = (row_date, parcel or row_date.isoformat())
            bucket = grouped.setdefault(group_key, {"principal": Decimal("0.00"), "juros": Decimal("0.00"), "encargos": Decimal("0.00"), "lines": [], "red_lines": [], "parcela": parcel, "situacoes": set(), "debitos": set(), "creditos": set(), "historicos": set(), "bancos": set()})
            if component == "JUROS":
                bucket["juros"] = Decimal(bucket["juros"]) + (amount or Decimal("0.00"))
            elif component in {"CAPITAL", "AMORTIZACAO", "PRINCIPAL"}:
                bucket["principal"] = Decimal(bucket["principal"]) + (principal_amount or amount or Decimal("0.00"))
            else:
                bucket["encargos"] = Decimal(bucket["encargos"]) + (amount or Decimal("0.00"))
            bucket["lines"].append(line)
            if any(highlighted[: len(values)]):
                bucket["red_lines"].append(line)
            if status:
                bucket["situacoes"].add(status)
            if bank:
                bucket["bancos"].add(bank)
            for index, field in ((debit_index, "debitos"), (credit_index, "creditos"), (history_index, "historicos")):
                if index is not None and index < len(values) and str(values[index]).strip():
                    bucket[field].add(str(values[index]).strip())

        pages.append("\n".join(lines))
        origin_name = "PLANILHA_FINANCIAMENTO" if any("GNATUS" in normalize_name(line) or "AYMORE" in normalize_name(line) for line in lines[:10]) else "PLANILHA_BB"
        origin_detail = "planilha_financiamento" if origin_name == "PLANILHA_FINANCIAMENTO" else "planilha_emprestimo_bb"
        for row_date, group_id in sorted(grouped):
            values = grouped[(row_date, group_id)]
            principal = Decimal(values["principal"])
            interest = Decimal(values["juros"])
            charges = Decimal(values["encargos"])
            paid = principal + interest + charges
            if paid <= 0:
                continue
            parcel = str(values["parcela"] or "")
            situations = sorted(values["situacoes"])
            banks = sorted(values["bancos"])
            details = {
                "contrato": contract,
                "principal": str(principal),
                "juros": str(interest),
                "encargos": str(charges),
                "valor_pago": str(paid),
                "origem": origin_detail,
                "parcela": parcel,
                "financiamento": title,
                "financiadores": lenders,
                "linha_credito": credit_line,
                "situacoes": situations,
                "bancos": banks,
                "debitos": sorted(values["debitos"]),
                "creditos": sorted(values["creditos"]),
                "historicos": sorted(values["historicos"]),
                "linhas_destacadas": values["red_lines"] or highlighted_lines,
            }
            financial = FinancialValues(
                valor_original=principal if principal > 0 else paid,
                valor_juros=interest if interest > 0 else None,
                valor_encargos=charges if charges > 0 else None,
                valor_pago=paid,
                detalhes=details,
                composicao_divergente=False,
            )
            descriptor = " ".join(part for part in [contract, parcel and f"Parcela {parcel}", "/".join(banks)] if part).strip()
            beneficiary = " ".join(part for part in ["Empréstimo/Financiamento", descriptor] if part).strip()
            receipts.append(
                ParsedReceipt(
                    data=row_date,
                    hora=None,
                    favorecido=beneficiary or "Empréstimo/Financiamento",
                    valor=paid,
                    tipo_operacao="EMPRESTIMO",
                    texto_original="\n".join(values["lines"]),
                    pagina_numero=page_number,
                    origem_nome=origin_name,
                    financeiros=financial,
                    beneficiario=beneficiary or "Empréstimo/Financiamento",
                    numero_documento=contract or parcel,
                )
            )
    return pages, receipts


def _extract_banco_do_brasil_loan_schedule(text: str, page_number: int) -> list[ParsedReceipt]:
    normalized = normalize_name(text)
    if not ("CRONOGRAMA" in normalized and "REPOSICAO" in normalized and "EXIGIVEL" in normalized):
        return []
    contract = ""
    contract_match = re.search(r"\bOpera[cç][aã]o\s*:?\s*([0-9.\s]{6,})", text, re.I)
    if contract_match:
        contract = re.sub(r"\s+", "", contract_match.group(1).strip(" ."))
    if not contract:
        fallback = re.search(r"\b\d{2,3}[.\s]\d{3}[.\s]\d{3}\b", text)
        contract = re.sub(r"\s+", "", fallback.group(0).strip(" .")) if fallback else ""

    amount = r"\d{1,3}(?:[.\s:;]\s*\d{3})*[,;:]\s*\d{2}"

    def loan_amounts(value: str) -> list[Decimal]:
        amounts: list[Decimal] = []
        for raw in re.findall(amount, value):
            normalized = re.sub(r"(?<=\d)[;:]\s*(?=\d{2}\b)", ",", raw)
            normalized = re.sub(r"(?<=\d)[;:]\s*(?=\d{3})", ".", normalized)
            normalized = re.sub(r"(?<=\d)\s+(?=\d{2}\b)", "", normalized)
            normalized = re.sub(r"\s+", "", normalized)
            parsed = parse_brl(normalized)
            if parsed is not None:
                amounts.append(parsed)
        return amounts

    row_pattern = re.compile(
        rf"^\s*(?P<date>\d{{2}}[/.]\d{{2}}[/.]\d{{4}})\s+"
        rf"(?P<component>JUROS|CAPITAL|AMORTIZA[ÇC][AÃ]O|PRINCIPAL)\b"
        rf"(?P<tail>.*?(?:{amount}).*)$",
        re.I | re.M,
    )
    grouped: dict[date, dict[str, Decimal | str]] = {}
    raw_lines: dict[date, list[str]] = defaultdict(list)
    for match in row_pattern.finditer(text):
        parsed_date, _ = parse_date_time(match.group("date"))
        if not parsed_date:
            continue
        values = loan_amounts(match.group("tail"))
        if not values:
            continue
        realized = values[1] if len(values) > 1 else values[0]
        component = normalize_name(match.group("component"))
        bucket = grouped.setdefault(parsed_date, {"principal": Decimal("0.00"), "juros": Decimal("0.00"), "encargos": Decimal("0.00")})
        if component == "JUROS":
            bucket["juros"] = Decimal(bucket["juros"]) + realized
        elif component in {"CAPITAL", "AMORTIZACAO", "PRINCIPAL"}:
            bucket["principal"] = Decimal(bucket["principal"]) + realized
        else:
            bucket["encargos"] = Decimal(bucket["encargos"]) + realized
        raw_lines[parsed_date].append(match.group(0).strip())

    results = []
    for parsed_date in sorted(grouped):
        values = grouped[parsed_date]
        principal = Decimal(values["principal"])
        interest = Decimal(values["juros"])
        charges = Decimal(values["encargos"])
        paid = principal + interest + charges
        if paid <= 0:
            continue
        details = {
            "contrato": contract,
            "principal": str(principal),
            "juros": str(interest),
            "encargos": str(charges),
            "valor_pago": str(paid),
            "origem": "cronograma_bb",
        }
        financial = FinancialValues(
            valor_original=principal if principal > 0 else paid,
            valor_juros=interest if interest > 0 else None,
            valor_encargos=charges if charges > 0 else None,
            valor_pago=paid,
            detalhes=details,
            composicao_divergente=False,
        )
        beneficiary = " ".join(part for part in ["Empréstimo/Financiamento", contract] if part).strip()
        results.append(
            ParsedReceipt(
                data=parsed_date,
                hora=None,
                favorecido=beneficiary or "Empréstimo/Financiamento",
                valor=paid,
                tipo_operacao="EMPRESTIMO",
                texto_original="\n".join(raw_lines[parsed_date]),
                pagina_numero=page_number,
                origem_nome="CRONOGRAMA_BB",
                financeiros=financial,
                beneficiario=beneficiary or "Empréstimo/Financiamento",
                numero_documento=contract,
            )
        )
    return results


def _extract_tefe_nfse_invoice(text: str, page_number: int) -> ParsedInvoice | None:
    normalized = normalize_name(text)
    if "PM TEFE" not in normalized or "NOTA FISCAL SERVICOS ELETRONICA" not in normalized:
        return None
    lines = [" ".join(line.split()) for line in text.splitlines() if line.strip()]

    def standalone_document(line: str) -> str:
        match = re.search(r"\b(?:\d{3}\.\d{3}\.\d{3}-\d{2}|\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})\b", line)
        return match.group(0) if match else ""

    def likely_name(line: str) -> bool:
        normalized_line = normalize_name(line)
        blocked = {
            "NOME RAZAO SOCIAL",
            "RG INSCRICAO ESTADUAL",
            "INSCRICAO MUNICIPAL",
            "CPF CNPJ DOCUMENTO",
            "CPF CNPJ",
            "CADASTRO",
            "LOGRADOURO",
            "COMPLEMENTO",
            "BAIRRO",
            "CEP COD POSTAL",
            "CIDADE PAIS",
            "COD IBGE",
            "TELEFONE",
            "E MAIL",
        }
        return bool(normalized_line and normalized_line not in blocked and not standalone_document(line) and not re.search(r"\d", line))

    def extract_party(window: list[str], cnpj_only: bool = False) -> tuple[str, str]:
        document_pattern = r"\b\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\b" if cnpj_only else r"\b(?:\d{3}\.\d{3}\.\d{3}-\d{2}|\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})\b"
        for position, candidate in enumerate(window):
            doc_match = re.search(document_pattern, candidate)
            if not doc_match:
                continue
            document = doc_match.group(0)
            name = candidate[doc_match.end():].strip()
            name = re.sub(r"^(?:\d+\s+){1,3}", "", name).strip()
            if not name:
                label_index = next((index for index in range(position - 1, -1, -1) if normalize_name(window[index]) == "NOME RAZAO SOCIAL"), None)
                if label_index is not None:
                    name = next((item for item in window[label_index + 1:position] if likely_name(item)), "")
            if not name:
                name = next((item for item in reversed(window[:position]) if likely_name(item)), "")
            return name, document
        return "", ""

    number = ""
    for index, line in enumerate(lines):
        if "NUMERO NFS E" in normalize_name(line):
            window = lines[max(0, index - 3):index + 4]
            number = next((candidate for candidate in window if re.fullmatch(r"\d{3,}", candidate)), "")
            break
    if not number:
        match = re.search(r"N[uú]mero da NFS-e\s*\n\s*(\d{3,})", text, re.I)
        number = match.group(1) if match else ""

    emission_text = next(
        (
            " ".join(lines[max(0, index - 4):index + 4])
            for index, line in enumerate(lines)
            if "DATA" in normalize_name(line) and "EMISSAO" in normalize_name(line)
        ),
        text,
    )
    parsed_date, _ = parse_date_time(emission_text)

    tomador_name = ""
    tomador_document = ""
    for index, line in enumerate(lines):
        if normalize_name(line) == "TOMADOR SERVICOS":
            tomador_name, tomador_document = extract_party(lines[index + 1:index + 10])
            if not tomador_document:
                tomador_name, tomador_document = extract_party(lines[max(0, index - 30):index])
            break

    prestador_name = ""
    prestador_document = ""
    for index, line in enumerate(lines):
        if normalize_name(line) == "PRESTADOR SERVICOS":
            prestador_name, prestador_document = extract_party(lines[index + 1:index + 10], cnpj_only=True)
            if not prestador_document:
                prestador_name, prestador_document = extract_party(lines[max(0, index - 25):index], cnpj_only=True)
            break

    invoice_total = None
    total_match = re.search(r"Valor\s+L[ií]quido\s+da\s+NFS-e\s*:\s*R\$\s*([\d.]+,\d{2})", text, re.I)
    if total_match:
        invoice_total = parse_brl(total_match.group(1))
    if invoice_total is None:
        total_match = re.search(r"Valor\s+Total\s+dos\s+Servi[cç]os.*?\n\s*R\$\s*([\d.]+,\d{2})", text, re.I | re.S)
        if total_match:
            invoice_total = parse_brl(total_match.group(1))

    service = ""
    service_match = re.search(r"\b1\.0\s+\S+\s+(.+?)\s+(?:\d+(?:[.,]\d+)?)\s+R\$\s*[\d.]+,\d{2}", text, re.I)
    if service_match:
        service = " ".join(service_match.group(1).split())

    payment = {}
    compact_text = " ".join(lines)
    faturas_match = re.search(r"FATURAS:\s*(.+?)(?:\s+Val\.\s*Aprox\.|\s+NFS-e\s+COMPOSTA|\Z)", compact_text, re.I)
    payment_matches = re.findall(
        r"(?:^|\s+-\s+)(.+?)\s+Venc:\s*(\d{2}/\d{2}/\d{4})\s+R\$\s*([\d.]+,\d{2})(?:\s+Doc:\s*([^\s]+))?",
        faturas_match.group(1) if faturas_match else "",
        re.I,
    )
    if payment_matches:
        payments = []
        for payment_type, payment_due, payment_value, payment_doc in payment_matches:
            parsed_payment_date, _ = parse_date_time(payment_due)
            payments.append({
                "forma_pagamento": " ".join(payment_type.split()),
                "data_vencimento": parsed_payment_date.isoformat() if parsed_payment_date else "",
                "data_pagamento": parsed_payment_date.isoformat() if parsed_payment_date else "",
                "valor": str(parse_brl(payment_value) or ""),
                "documento": payment_doc or "",
            })
        first_payment = payments[0]
        payment_date = parse_date_time(payment_matches[0][1])[0]
        payment_type_original = first_payment["forma_pagamento"] if len(payments) == 1 else "Pagamento dividido"
        payment_total = sum((Decimal(item["valor"] or "0.00") for item in payments), Decimal("0.00"))
        payment = {
            "forma_pagamento": payment_type_original,
            "tipo_pagamento_original": payment_type_original,
            "data_vencimento": payment_date.isoformat() if payment_date else "",
            "data_pagamento": payment_date.isoformat() if payment_date else "",
            "valor_pagamento": str(payment_total if len(payments) > 1 else parse_brl(payment_matches[0][2]) or ""),
            "documento_pagamento": first_payment["documento"],
            "pagamentos": payments,
        }

    counterparty = tomador_name or prestador_name
    document = tomador_document or prestador_document
    if not counterparty and not number and invoice_total is None:
        return None
    return ParsedInvoice(
        parsed_date,
        counterparty,
        document,
        number,
        invoice_total,
        text,
        page_number,
        {
            "layout": "tefe_nfse",
            "tomador": tomador_name,
            "cpf_cnpj_tomador": tomador_document,
            "prestador": prestador_name,
            "cnpj_prestador": prestador_document,
            "servico": service,
            **payment,
        },
    )


def extract_invoices(text: str, page_number: int) -> list[ParsedInvoice]:
    normalized = normalize_name(text)
    if not any(term in normalized for term in ("NOTA FISCAL", "NF E", "NFS E", "DANFE")):
        return []
    tefe_invoice = _extract_tefe_nfse_invoice(text, page_number)
    if tefe_invoice:
        return [tefe_invoice]

    def label_text(*labels: str) -> str:
        for label in labels:
            value = receipt_label_value(text, label, value_prefix=r"[^\n]")
            if value:
                return " ".join(value.split())
        return ""

    def label_amount(*labels: str) -> Decimal | None:
        for label in labels:
            value = receipt_label_value(text, label, value_prefix=r"(?:R?\$?\s*[\d.,-])")
            if value:
                amount = parse_brl(value)
                if amount is not None:
                    return abs(amount)
        return None

    number = label_text(
        "NÚMERO DA NOTA",
        "NUMERO DA NOTA",
        "Nº DA NOTA",
        "N° DA NOTA",
        "NOTA FISCAL Nº",
        "NOTA FISCAL N",
        "NF-E Nº",
        "NFS-E Nº",
        "NÚMERO",
        "NUMERO",
    )
    if not number:
        match = re.search(r"\b(?:NF-?E|NFS-?E|NOTA FISCAL)\s*(?:N[º°.]?|NUMERO)?\s*[:\-]?\s*(\d{3,})", text, re.I)
        number = match.group(1) if match else ""

    cnpj_match = re.search(r"\b\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\b", text)
    cpf_cnpj = cnpj_match.group(0) if cnpj_match else ""
    supplier = label_text(
        "PRESTADOR",
        "RAZÃO SOCIAL",
        "RAZAO SOCIAL",
        "FORNECEDOR",
        "EMITENTE",
        "NOME/RAZÃO SOCIAL",
        "NOME/RAZAO SOCIAL",
    )
    if not supplier:
        lines = [" ".join(line.split()) for line in text.splitlines() if line.strip()]
        supplier = next((line for line in lines if cpf_cnpj and cpf_cnpj not in line and len(line) > 4 and not re.search(r"\d{2}/\d{2}/\d{4}", line)), "")

    parsed_date, _ = parse_date_time(
        "\n".join(
            value
            for value in [
                label_text("DATA DE EMISSÃO", "DATA DE EMISSAO", "EMISSÃO", "EMISSAO"),
                text,
            ]
            if value
        )
    )
    due_date, _ = parse_date_time(label_text("DATA DE VENCIMENTO", "VENCIMENTO", "VENC:"))
    payment_date, _ = parse_date_time(label_text("DATA DE PAGAMENTO", "DATA DO PAGAMENTO", "PAGAMENTO EM", "PAGO EM"))
    payment_original = label_text("FORMA DE PAGAMENTO", "MEIO DE PAGAMENTO", "TIPO DE PAGAMENTO", "PAGAMENTO")
    total = label_amount(
        "VALOR TOTAL DA NOTA",
        "VALOR TOTAL",
        "VALOR DOS SERVIÇOS",
        "VALOR DOS SERVICOS",
        "VALOR DO SERVIÇO",
        "VALOR DO SERVICO",
    )
    if not supplier and not number and total is None:
        return []
    data = {
        "data_vencimento": due_date.isoformat() if due_date else "",
        "data_pagamento": payment_date.isoformat() if payment_date else "",
        "forma_pagamento": payment_original,
        "tipo_pagamento_original": payment_original,
    }
    return [ParsedInvoice(parsed_date, supplier, cpf_cnpj, number, total, text, page_number, data)]


def split_banco_do_brasil_receipt_blocks(text: str) -> list[str]:
    pattern = r"(?m)^(?=(?:SISBB\s+-\s+SISTEMA|(?:\d{2}[/.]\d{2}[/.]\d{4})\s+-\s+BANCO\s+DO\s+BRASIL\b))"
    blocks = [block.strip() for block in re.split(pattern, text) if block.strip()]
    return blocks if len(blocks) > 1 else [text]


def receipt_operation_from_text(text: str) -> str:
    normalized = normalize_name(text)
    if "PIX" in normalized or "QR CODE" in normalized:
        return "PIX"
    if "TED" in normalized:
        return "TED"
    if "TRANSFERENCIA" in normalized:
        return "TRANSFERÊNCIA"
    if "DEBITO AUTOMATICO" in normalized:
        return "DÉBITO AUTOMÁTICO"
    if "RECEB" in normalized:
        return "RECEBIMENTO"
    return "PAGAMENTO"


GETNET_CARD_RE = re.compile(r"^(?:MASTERCARD|VISA|ELO|AMEX|HIPERCARD)(?:\s+\S+)*\s+(?:CR[ÉE]DITO|CREDITO|D[ÉE]BITO|DEBITO)$", re.I)


def _extract_getnet_sales_receipts(text: str, page_number: int) -> list[ParsedReceipt]:
    normalized = normalize_name(text)
    inline_records = _extract_getnet_inline_sales_receipts(text, page_number)
    if inline_records:
        return inline_records
    if "O QUE VENDI" not in normalized or "CONSOLIDADO POR DATA" not in normalized:
        return []

    lines = [" ".join(line.split()) for line in text.splitlines() if line.strip()]
    results: list[ParsedReceipt] = []
    index = 0
    while index < len(lines):
        card = lines[index]
        if not GETNET_CARD_RE.fullmatch(card):
            index += 1
            continue
        row = lines[index:index + 7]
        if len(row) < 7:
            break
        parsed_date, _ = parse_date_time(row[1])
        gross = parse_brl(row[3])
        net = parse_brl(row[4])
        fee = parse_brl(row[5])
        quantity = row[6]
        if (
            not parsed_date
            or gross is None
            or net is None
            or fee is None
            or not re.fullmatch(r"\d+", quantity)
        ):
            index += 1
            continue
        if gross == Decimal("0.00") and net == Decimal("0.00") and fee == Decimal("0.00"):
            index += 7
            continue
        results.append(_getnet_sales_receipt(parsed_date, card, row[2], gross, abs(fee), net, quantity, "\n".join(row), page_number))
        index += 7
    return results


def _extract_getnet_inline_sales_receipts(text: str, page_number: int) -> list[ParsedReceipt]:
    amount = r"R\$\s*\d{1,3}(?:\.\d{3})*,\d{2}"
    card = r"(?:MASTERCARD|VISA|ELO|AMEX|HIPERCARD)(?:\s+\S+)*?\s+(?:CR[ÉE]DITO|CREDITO|D[ÉE]BITO|DEBITO)"
    pattern = re.compile(
        rf"^\s*(?P<establishment>\d{{6,}})\s+"
        rf"(?P<card>{card})\s+"
        rf"(?P<date>\d{{2}}/\d{{2}}/\d{{4}})\s+"
        rf"(?P<quantity>\d+)\s+"
        rf"(?P<gross>{amount})\s+"
        rf"(?P<fee>-?\s*{amount})\s+"
        rf"(?P<net>{amount})\s*$",
        re.I | re.M,
    )
    results: list[ParsedReceipt] = []
    for match in pattern.finditer(text):
        parsed_date, _ = parse_date_time(match.group("date"))
        gross = parse_brl(match.group("gross"))
        fee = parse_brl(match.group("fee"))
        net = parse_brl(match.group("net"))
        if not parsed_date or gross is None or fee is None or net is None:
            continue
        if gross == Decimal("0.00") and net == Decimal("0.00") and fee == Decimal("0.00"):
            continue
        results.append(_getnet_sales_receipt(
            parsed_date,
            " ".join(match.group("card").split()),
            match.group("establishment"),
            gross,
            abs(fee),
            net,
            match.group("quantity"),
            match.group(0).strip(),
            page_number,
        ))
    return results


def _getnet_sales_receipt(parsed_date: date, card: str, establishment: str, gross: Decimal, fee: Decimal, net: Decimal, quantity: str, raw_text: str, page_number: int) -> ParsedReceipt:
    details = {
        "cartao": card,
        "estabelecimento": establishment,
        "quantidade_vendas": quantity,
        "valor_original": str(gross),
        "valor_tarifa": str(fee),
        "valor_pago": str(net),
    }
    financial = FinancialValues(
        valor_original=gross,
        valor_tarifa=fee,
        valor_pago=net,
        detalhes=details,
        composicao_divergente=abs(gross - fee - net) > Decimal("0.01"),
    )
    return ParsedReceipt(
        data=parsed_date,
        hora=None,
        favorecido=f"GETNET - {card}",
        valor=net,
        tipo_operacao="GETNET VENDAS",
        texto_original=raw_text,
        pagina_numero=page_number,
        origem_nome="GETNET VENDAS",
        financeiros=financial,
        beneficiario=f"GETNET - {card}",
        numero_documento=establishment,
    )


def _date_from_day_month(value: str, period: tuple[int, int] | None) -> date | None:
    match = re.fullmatch(r"(\d{2})/(\d{2})", value.strip())
    if not match or not period:
        return None
    return date(period[1], int(match.group(2)), int(match.group(1)))


def _text_section_lines(text: str, start: str, *ends: str) -> list[str]:
    match = re.search(re.escape(start), text, re.I)
    if not match:
        return []
    section = text[match.end():]
    end_positions = [found.start() for end in ends if (found := re.search(re.escape(end), section, re.I))]
    if end_positions:
        section = section[: min(end_positions)]
    return [" ".join(line.split()) for line in section.splitlines() if line.strip()]


def _santander_receipt_financial(amount: Decimal) -> FinancialValues:
    return FinancialValues(valor_original=amount, valor_pago=amount, detalhes={"valor_original": str(amount), "valor_pago": str(amount)})


def _santander_receipt(data: date, favorecido: str, amount: Decimal, operation: str, text: str, page_number: int, origin: str, document: str = "", hora: str | None = None) -> ParsedReceipt:
    financial = _santander_receipt_financial(amount)
    name = " ".join(favorecido.split())
    return ParsedReceipt(data, hora, name, amount, operation, text.strip(), page_number, origin, financial, name, numero_documento=document)


def _extract_santander_receipts(text: str, page_number: int) -> list[ParsedReceipt]:
    internet_banking = _extract_santander_internet_banking_receipt(text, page_number)
    if internet_banking:
        return internet_banking
    normalized = normalize_name(text)
    if not any(marker in normalized for marker in ("DEBITO AUTOMATICO EM CONTA CORRENTE", "COMPROVANTES PAGAMENTO", "TRANSFERENCIAS ENTRE CONTAS DOCS TEDS PIXS ENVIADOS")):
        return []
    period = _santander_statement_period(text)
    records = []
    records.extend(_extract_santander_automatic_debit_receipts(text, page_number, period))
    records.extend(_extract_santander_consumption_receipts(text, page_number, period))
    records.extend(_extract_santander_transfer_receipts(text, page_number, period))
    return records


def _line_value_after(lines: list[str], label: str) -> str:
    wanted = normalize_name(label)
    for index, line in enumerate(lines):
        normalized = normalize_name(line.replace(":", " "))
        if normalized == wanted and index + 1 < len(lines):
            return lines[index + 1].strip()
        if normalized.startswith(wanted + " "):
            return line.split(":", 1)[1].strip() if ":" in line else line[len(label):].strip()
    return ""


def _santander_beneficiary_section_lines(lines: list[str]) -> list[str]:
    start = -1
    for index, line in enumerate(lines):
        normalized = normalize_name(line)
        if "DADOS" in normalized and "ORIGINAL" in normalized and "PAGADOR" not in normalized:
            start = index + 1
            break
    if start < 0:
        return []
    end = len(lines)
    for index in range(start, len(lines)):
        normalized = normalize_name(lines[index])
        if ("DADOS" in normalized and "PAGADOR" in normalized) or normalized == "DADOS PAGAMENTO":
            end = index
            break
    return lines[start:end]


def _collect_multiline_label_value(lines: list[str], label: str, stop_labels: set[str]) -> str:
    wanted = normalize_name(label)
    for index, line in enumerate(lines):
        normalized = normalize_name(line.replace(":", " "))
        if normalized != wanted and not normalized.startswith(wanted + " "):
            continue
        values = []
        if ":" in line and line.split(":", 1)[1].strip():
            values.append(line.split(":", 1)[1].strip())
        cursor = index + 1
        while cursor < len(lines):
            current = lines[cursor].strip()
            current_normalized = normalize_name(current.replace(":", " "))
            if current_normalized in stop_labels or current_normalized.startswith("DADOS "):
                break
            values.append(current)
            cursor += 1
        return " ".join(values).strip()
    return ""


def _first_cnpj(lines: list[str]) -> str:
    match = re.search(r"\b\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\b", "\n".join(lines))
    return match.group(0) if match else ""


def _extract_santander_internet_banking_receipt(text: str, page_number: int) -> list[ParsedReceipt]:
    normalized = normalize_name(text)
    if not ("INTERNET BANKING EMPRESARIAL" in normalized and "DADOS PAGAMENTO" in normalized and "VALOR TOTAL PAGO" in normalized):
        return []
    lines = [" ".join(line.split()) for line in text.splitlines() if line.strip()]
    beneficiary_lines = _santander_beneficiary_section_lines(lines)
    stop_labels = {"CNPJ", "NOME FANTASIA", "DADOS PAGADOR ORIGINAL", "DADOS PAGADOR EFETIVO", "DADOS PAGAMENTO"}
    name = _collect_multiline_label_value(beneficiary_lines, "Razão Social", stop_labels)
    if not name:
        name = _collect_multiline_label_value(beneficiary_lines, "Nome Fantasia", stop_labels)
    paid_date, paid_time = parse_date_time(_line_value_after(lines, "Data da Transação"))
    paid_amount = parse_brl(_line_value_after(lines, "Valor total pago"))
    if not name or not paid_date or paid_amount is None:
        return []
    original_amount = parse_brl(_line_value_after(lines, "Valor Nominal"))
    charges = parse_brl(_line_value_after(lines, "Encargos"))
    document = _line_value_after(lines, "Número de Autenticação da Instituição Financeira Favorecida") or _line_value_after(lines, "Nosso Número")
    financial = FinancialValues(
        valor_original=original_amount or paid_amount,
        valor_encargos=charges,
        valor_pago=paid_amount,
        detalhes={"valor_pago": str(paid_amount)},
    )
    if financial.valor_original is not None:
        financial.detalhes["valor_original"] = str(financial.valor_original)
    if charges is not None:
        financial.detalhes["valor_encargos"] = str(charges)
    if financial.valor_original is not None and charges is not None:
        financial.composicao_divergente = abs(financial.valor_original + charges - paid_amount) > Decimal("0.01")
    cnpj = _first_cnpj(beneficiary_lines)
    return [ParsedReceipt(paid_date, paid_time, name, paid_amount, "PAGAMENTO", text.strip(), page_number, "SANTANDER RAZÃO SOCIAL", financial, name, cnpj_beneficiario=cnpj, numero_documento=document)]


def _extract_santander_automatic_debit_receipts(text: str, page_number: int, period: tuple[int, int] | None) -> list[ParsedReceipt]:
    lines = _text_section_lines(text, "Débito Automático em Conta Corrente", "Comprovantes de Pagamento", "Transferências entre Contas", "Créditos Contratados")
    results = []
    index = 0
    while index < len(lines):
        paid_date = _date_from_day_month(lines[index], period)
        if not paid_date:
            index += 1
            continue
        row = []
        index += 1
        while index < len(lines) and not re.fullmatch(r"\d{2}/\d{2}", lines[index]):
            row.append(lines[index])
            index += 1
        value_index = next((i for i, line in enumerate(row) if _santander_parse_movement_value(line) is not None), -1)
        if value_index <= 0:
            continue
        amount = _santander_parse_movement_value(row[value_index])
        if amount is None:
            continue
        document = next((line for line in row[:value_index] if re.fullmatch(r"\d{4,}", line)), "")
        document_index = row.index(document) if document else value_index
        description = " ".join(line for line in row[:document_index] if not _santander_is_receipt_header_line(line))
        if description:
            results.append(_santander_receipt(paid_date, description, amount, "DÉBITO AUTOMÁTICO", "\n".join([paid_date.strftime("%d/%m"), *row]), page_number, "SANTANDER DÉBITO AUTOMÁTICO", document))
    return results


def _extract_santander_consumption_receipts(text: str, page_number: int, period: tuple[int, int] | None) -> list[ParsedReceipt]:
    lines = _text_section_lines(text, "Contas de Consumo", "Transferências entre Contas", "Créditos Contratados")
    results = []
    index = 0
    while index < len(lines):
        paid_date = _date_from_day_month(lines[index], period)
        if not paid_date:
            index += 1
            continue
        row = []
        index += 1
        while index < len(lines) and not re.fullmatch(r"\d{2}/\d{2}", lines[index]):
            row.append(lines[index])
            index += 1
        value_index = next((i for i, line in enumerate(row) if _santander_parse_movement_value(line) is not None), -1)
        if value_index < 0:
            continue
        amount = _santander_parse_movement_value(row[value_index])
        if amount is None:
            continue
        before_value = [line for line in row[:value_index] if not _santander_is_receipt_header_line(line)]
        while before_value and normalize_name(before_value[0]) in {"INTERNET", "BANKING", "INTERNET BANKING"}:
            before_value.pop(0)
        if before_value and before_value[-1] == "-":
            before_value.pop()
        name = " ".join(before_value)
        document = next((line for line in reversed(row[value_index + 1:]) if re.fullmatch(r"[\d-]{12,}", line)), "")
        if name:
            results.append(_santander_receipt(paid_date, name, amount, "PAGAMENTO", "\n".join([paid_date.strftime("%d/%m"), *row]), page_number, "SANTANDER CONTAS DE CONSUMO", document))
    return results


def _extract_santander_transfer_receipts(text: str, page_number: int, period: tuple[int, int] | None) -> list[ParsedReceipt]:
    lines = _text_section_lines(text, "Transferências entre Contas, DOCs, TEDs e PIXs Enviados", "Não estão contempladas", "*Identificador", "Créditos Contratados")
    results = []
    row = []

    def flush() -> None:
        nonlocal row
        if not row:
            return
        joined = " ".join(row)
        match = re.search(
            r"^(?P<date>\d{2}/\d{2})(?:\s+INTERNET(?:\s+BANKING)?|\s+BANKING)*\s+(?P<op>TRANSF\.?\s+CONTAS|TED|DOC|PIX)\s+(?P<name>.+?)\s+(?P<bank>\d{4})\s+(?P<agency>\d{4})\s+(?P<account>\d{6,})\s+(?P<value>\d{1,3}(?:\.\d{3})*,\d{2})$",
            joined,
            re.I,
        )
        if match:
            paid_date = _date_from_day_month(match.group("date"), period)
            amount = _santander_parse_movement_value(match.group("value"))
            if paid_date and amount is not None:
                raw_operation = normalize_name(match.group("op"))
                operation = "TED" if raw_operation == "TED" else "PIX" if raw_operation == "PIX" else "TRANSFERÊNCIA"
                document = f"{match.group('bank')} {match.group('agency')} {match.group('account')}"
                results.append(_santander_receipt(paid_date, match.group("name"), amount, operation, "\n".join(row), page_number, f"SANTANDER {operation}", document))
        row = []

    for line in lines:
        if _santander_is_receipt_header_line(line):
            continue
        if re.match(r"^\d{2}/\d{2}\b", line):
            flush()
        row.append(line)
    flush()
    return results


def _santander_is_receipt_header_line(line: str) -> bool:
    return normalize_name(line) in {
        "DATA", "DESCRICAO", "N IDENTIFICACAO", "VALOR R", "REALIZADO", "MOTIVO",
        "LIMITE PARA", "DEBITO R", "DATA PAGAMENTO CANAL", "NOME EMPRESA",
        "DATA VENCIMENTO", "VALOR", "R", "CODIGO BARRAS", "AUTENTICACAO BANCARIA",
        "DATA CANAL", "TIPO", "FAVORECIDO", "BANCO AGENCIA", "CONTA",
    }


def extract_receipt_block(text: str, page_number: int) -> list[ParsedReceipt]:
    """A record is emitted only from a self-contained block with required labels."""
    blocks = re.split(r"(?=^\s*VALOR\s*:)", text, flags=re.I | re.M)
    results = []
    for block in blocks:
        value_match = re.search(r"^\s*VALOR\s*:\s*([^\n]+)", block, re.I | re.M)
        name_info = find_receipt_name_with_origin(block)
        name = name_info[0] if name_info else None
        date_match = re.search(r"^\s*(?:DATA DO PAGAMENTO|DATA PAGAMENTO)\s*:?\s*([^\n]+)", block, re.I | re.M) or re.search(r"^\s*(?:DATA|DEBITO EM|DÉBITO EM)\s*:\s*([^\n]+)", block, re.I | re.M)
        if not (value_match and name and date_match):
            continue
        parsed_date, parsed_time = parse_date_time(date_match.group(1))
        financial = extract_financial_values(block)
        amount = financial.valor_pago
        if not parsed_date or amount is None or not name:
            continue
        operation = receipt_operation_from_text(block)
        if operation == "PAGAMENTO":
            operation = receipt_operation_from_text(text)
        participants = receipt_participants(block); results.append(ParsedReceipt(parsed_date, parsed_time, name, amount, operation, block.strip(), page_number, name_info[1], financial, participants.get("beneficiario", name), participants.get("nome_fantasia", ""), participants.get("beneficiario_final", ""), participants.get("pagador", ""), participants.get("cnpj_beneficiario", ""), participants.get("cnpj_beneficiario_final", ""), receipt_document_number(block)))
    if results:
        return results

    banco_do_brasil = _extract_banco_do_brasil_receipt(text, page_number)
    if banco_do_brasil:
        return banco_do_brasil

    # Payment receipts use a distinct layout, usually with one receipt per PDF page.
    value_match = receipt_label_value(text, "VALOR DO DOCUMENTO", "VALOR COBRADO", "VALOR PAGO", "VALOR TOTAL", "VALOR RECEBIDO", "VALOR TRANSFERIDO", "VALOR DA TRANSFERÊNCIA", "VALOR DA TRANSFERENCIA", "VALOR DO DEBITO", "VALOR DO DÉBITO", "VALOR", value_prefix=r"(?:R?\$?\s*[\d.,-])")
    name_info = find_receipt_name_with_origin(text)
    name = name_info[0] if name_info else None
    if not name:
        name = receipt_label_value(text, "RECEBIDO DE", "RECEBIMENTO DE")
        name_info = (name, "RECEBIDO DE") if name else None
    date_value = receipt_label_value(text, "DATA DO PAGAMENTO", "DATA PAGAMENTO", "DATA DO DEBITO", "DATA DO DÉBITO", "DATA", "DATA DO RECEBIMENTO", "DEBITO EM", "DÉBITO EM", value_prefix=r"\d")
    date_match = re.search(DATE_RE, date_value or "")
    if value_match and name and date_match:
        parsed_date, parsed_time = parse_date_time(date_value or "")
        financial = extract_financial_values(text)
        amount = financial.valor_pago
        if parsed_date and amount is not None and name and name.upper() not in {"PIX", "TED", "DOCUMENTO"}:
            operation = receipt_operation_from_text(text)
            participants = receipt_participants(text); return [ParsedReceipt(parsed_date, parsed_time, name, amount, operation, text.strip(), page_number, name_info[1], financial, participants.get("beneficiario", name), participants.get("nome_fantasia", ""), participants.get("beneficiario_final", ""), participants.get("pagador", ""), participants.get("cnpj_beneficiario", ""), participants.get("cnpj_beneficiario_final", ""), receipt_document_number(text))]
    return results


def _extract_banco_do_brasil_receipt(text: str, page_number: int) -> list[ParsedReceipt]:
    """Handles BB payment and account-transfer receipts with values on following lines."""
    participants = receipt_participants(text)
    payment_name = participants.get("beneficiario") or participants.get("beneficiario_final") or participants.get("nome_fantasia")
    payment_origin = "BENEFICIARIO"
    if not payment_name:
        name_info = find_receipt_name_with_origin(text)
        payment_name = name_info[0] if name_info else None
        payment_origin = name_info[1] if name_info else payment_origin
    payment_date = receipt_label_value(text, "DATA DO PAGAMENTO", "DATA PAGAMENTO", "DATA DO DEBITO", "DATA DO DÉBITO", "PAGAMENTO EFETUADO EM", value_prefix=r"\d")
    payment_value = receipt_label_value(text, "VALOR DO DOCUMENTO", "VALOR COBRADO", "VALOR PAGO", "VALOR TOTAL", "VALOR RECEBIDO", "VALOR DO DEBITO", "VALOR DO DÉBITO", value_prefix=r"(?:R?\$?\s*[\d.,-])")
    if payment_name and payment_date and payment_value:
        parsed_date, _ = parse_date_time(payment_date)
        financial = extract_financial_values(text)
        amount = financial.valor_pago
        if parsed_date and amount is not None:
            return [ParsedReceipt(parsed_date, None, payment_name, amount, receipt_operation_from_text(text), text.strip(), page_number, payment_origin, financial, participants.get("beneficiario", payment_name), participants.get("nome_fantasia", ""), participants.get("beneficiario_final", ""), participants.get("pagador", ""), participants.get("cnpj_beneficiario", ""), participants.get("cnpj_beneficiario_final", ""), receipt_document_number(text))]

    transfer_name = re.search(r"TRANSFERIDO PARA\s*:\s*(?:\n|\s)+(?:CLIENTE|NOME|FAVORECIDO)\s*:?\s*([^\n]+)", text, re.I)
    transfer_date = receipt_label_value(text, "DATA DA TRANSFERÊNCIA", "DATA DA TRANSFERENCIA", "DATA TRANSFERÊNCIA", "DATA TRANSFERENCIA", "DATA", value_prefix=r"\d")
    transfer_value = receipt_label_value(text, "VALOR TOTAL", "VALOR TRANSFERIDO", "VALOR DA TRANSFERÊNCIA", "VALOR DA TRANSFERENCIA", "VALOR", value_prefix=r"(?:R?\$?\s*[\d.,-])")
    if transfer_name and transfer_date and transfer_value:
        parsed_date, _ = parse_date_time(transfer_date)
        financial = extract_financial_values(text)
        amount = financial.valor_pago
        if parsed_date and amount is not None:
            return [ParsedReceipt(parsed_date, None, transfer_name.group(1).strip(), amount, "TRANSFERÊNCIA", text.strip(), page_number, "TRANSFERIDO PARA", financial, numero_documento=receipt_document_number(text))]
    return []


def extract_statement(text: str, page_number: int, bank: str = "") -> list[ParsedStatement]:
    if bank == "Banco do Brasil":
        bb_records = _extract_banco_do_brasil_statement(text, page_number)
        if bb_records:
            return bb_records
    if bank == "Santander":
        santander_records = _extract_santander_statement(text, page_number)
        if santander_records:
            return santander_records
    if bank == "BASA":
        basa_records = _extract_basa_statement(text, page_number)
        if basa_records:
            return basa_records
    if bank == "Caixa":
        caixa_records = _extract_caixa_statement(text, page_number)
        if caixa_records:
            return caixa_records
    if bank == "Bradesco":
        bradesco_records, _ = _extract_bradesco_statement_page(text, page_number)
        if bradesco_records:
            return bradesco_records

    results = []
    for line in text.splitlines():
        columns = [item.strip() for item in line.split("|")]
        if len(columns) < 3:
            continue
        parsed_date, _ = parse_date_time(columns[0])
        amount = parse_brl(columns[-1])
        if not parsed_date or amount is None:
            continue
        history = " ".join(columns[1:-1])
        _, hour = parse_date_time(history)
        name = re.sub(r".*?\d{2}/\d{2}\s+(?:\d{2}:\d{2}(?::\d{2})?\s+)?", "", history).strip()
        nature = "Débito" if re.search(r"enviado|debito|débito|pagamento", history, re.I) else "Crédito"
        results.append(ParsedStatement(parsed_date, hour, history, name, amount, nature, line, page_number))
    return results


def extract_statement_pages(pages: list[str], bank: str = "") -> list[ParsedStatement]:
    if bank == "Santander":
        total_pages = len(pages)
        combined = "\n".join(
            f"Pagina:{number}/{total_pages}\n{page_text}"
            for number, page_text in enumerate(pages, 1)
        )
        return extract_statement(combined, 1, bank)
    if bank == "Bradesco":
        return _extract_bradesco_statement_pages(pages)

    records: list[ParsedStatement] = []
    for number, page_text in enumerate(pages, 1):
        records.extend(extract_statement(page_text, number, bank))
    return records


def extract_santander_pdfplumber_statement(path: str | Path) -> PdfStatementExtraction | None:
    try:
        import pdfplumber
    except ImportError:
        logger.warning("pdfplumber não instalado; usando extração textual para Santander")
        return None

    pages_text: list[str] = []
    pages_words: list[list[dict[str, object]]] = []
    try:
        with pdfplumber.open(path) as document:
            for page in document.pages:
                pages_text.append(page.extract_text(x_tolerance=1, y_tolerance=3, layout=True) or "")
                pages_words.append(page.extract_words(x_tolerance=1, y_tolerance=3, keep_blank_chars=False, use_text_flow=False))
    except Exception as error:
        logger.warning("Falha ao ler Santander com pdfplumber: %s", error)
        return None

    if not any(_is_santander_current_account_statement_page(text) for text in pages_text):
        return None

    records = extract_santander_words_statement(pages_words, pages_text)
    _validate_santander_expected_statement(records, pages_text)
    return PdfStatementExtraction(pages_text, records)


def extract_basa_pdfplumber_pages(path: str | Path) -> list[str]:
    try:
        import pdfplumber
    except ImportError:
        logger.warning("pdfplumber não instalado; usando PyMuPDF para extrato BASA")
        return []

    try:
        with pdfplumber.open(path) as document:
            pages = [
                page.extract_text(x_tolerance=1, y_tolerance=3, layout=True) or ""
                for page in document.pages
            ]
    except Exception as error:
        logger.warning("Falha ao ler BASA com pdfplumber: %s", error)
        return []

    if any("VALOR LANCTO" in page and "D/C" in page for page in pages):
        return pages
    return []


def extract_getnet_pdfplumber_pages(path: str | Path) -> list[str]:
    try:
        import pdfplumber
    except ImportError:
        logger.warning("pdfplumber não instalado; usando PyMuPDF para relatório Getnet")
        return []

    try:
        with pdfplumber.open(path) as document:
            pages = [
                page.extract_text(x_tolerance=1, y_tolerance=3, layout=True) or ""
                for page in document.pages
            ]
    except Exception as error:
        logger.warning("Falha ao ler Getnet com pdfplumber: %s", error)
        return []

    if any("O que vendi" in page or _extract_getnet_inline_sales_receipts(page, index) for index, page in enumerate(pages, 1)):
        return pages
    return []


def extract_santander_words_statement(
    pages_words: list[list[dict[str, object]]],
    pages_text: list[str],
) -> list[ParsedStatement]:
    period = _santander_statement_period("\n".join(pages_text))
    if not period:
        return []

    records: list[ParsedStatement] = []
    current: dict[str, object] | None = None
    current_date: date | None = None
    in_movement_table = False
    columns: dict[str, float] | None = None
    pending_columns: dict[str, float] = {}
    final_balance_pattern = _santander_balance_line_pattern(period)
    initial_balance_pattern = _santander_previous_balance_line_pattern(period)

    def flush() -> None:
        nonlocal current
        if not current:
            return
        history = _santander_clean_description(str(current["historico"]))
        history = _dedupe_santander_description(history)
        if history and not _santander_should_skip_history(history):
            raw_value = str(current["raw_value"])
            records.append(ParsedStatement(
                data=current["data"],  # type: ignore[arg-type]
                hora=None,
                historico=history,
                nome=_santander_statement_name(history),
                valor=current["valor"],  # type: ignore[arg-type]
                natureza=str(current["natureza"]),
                texto_original=str(current["texto_original"]),
                pagina_numero=int(current["pagina_numero"]),
                numero_documento=str(current.get("numero_documento") or ""),
            ))
        current = None

    for page_index, (page_words, page_text) in enumerate(zip(pages_words, pages_text), 1):
        page_lines = _pdf_words_to_lines(page_words)
        page_text_norm = normalize_name(page_text)
        if not in_movement_table and "CONTA CORRENTE" not in page_text_norm and "MOVIMENTACAO" not in page_text_norm:
            continue
        for line in page_lines:
            text = _word_line_text(line)
            normalized = normalize_name(text)
            if final_balance_pattern.search(normalized):
                flush()
                return records
            if initial_balance_pattern.search(normalized) and current_date is None:
                current_date = date(period[1], period[0], 1)
                continue
            if _santander_informational_section_starts(normalized):
                if in_movement_table and columns:
                    flush()
                    in_movement_table = False
                    pending_columns.clear()
                    break
                continue
            if normalized == "MOVIMENTACAO":
                in_movement_table = True
                continue
            if normalized == "CONTA CORRENTE":
                in_movement_table = True
                continue
            found_columns = _santander_columns_from_words(line, pending_columns)
            if found_columns:
                columns = found_columns
                pending_columns.clear()
                in_movement_table = True
                continue
            if not in_movement_table or not columns or _santander_should_skip_word_line(normalized):
                continue

            parsed_date = _santander_word_line_date(line, period[1], columns)
            if parsed_date:
                current_date = parsed_date

            movement_value = _santander_word_line_value(line, columns)
            if movement_value:
                if not current_date:
                    continue
                flush()
                history_words = _santander_description_words(line, columns, movement_value["x0"])
                history = _santander_clean_description(" ".join(history_words))
                document = _santander_document_from_words(line, columns)
                if not history:
                    continue
                current = {
                    "data": current_date,
                    "historico": history,
                    "valor": movement_value["amount"],
                    "raw_value": movement_value["raw_value"],
                    "natureza": movement_value["natureza"],
                    "texto_original": text,
                    "pagina_numero": page_index,
                    "numero_documento": document,
                }
                continue

            if current and not parsed_date:
                continuation = _santander_continuation_words(line, columns)
                if continuation:
                    addition = _santander_clean_description(" ".join(continuation))
                    if addition:
                        current["historico"] = f"{current['historico']} {addition}"
                        current["texto_original"] = f"{current['texto_original']}\n{text}"

    flush()
    return records


def _santander_balance_line_pattern(period: tuple[int, int]) -> re.Pattern[str]:
    month, year = period
    last_day = calendar.monthrange(year, month)[1]
    return re.compile(rf"\bSALDO EM {last_day:02d} {month:02d}\b")


def _santander_previous_balance_line_pattern(period: tuple[int, int]) -> re.Pattern[str]:
    month, year = period
    previous_month = month - 1 or 12
    previous_year = year - 1 if month == 1 else year
    previous_last_day = calendar.monthrange(previous_year, previous_month)[1]
    return re.compile(rf"\bSALDO EM {previous_last_day:02d} {previous_month:02d}\b")


def _pdf_words_to_lines(words: list[dict[str, object]]) -> list[list[dict[str, object]]]:
    lines: list[list[dict[str, object]]] = []
    for word in sorted(words, key=lambda item: (float(item.get("top", 0)), float(item.get("x0", 0)))):
        y = (float(word.get("top", 0)) + float(word.get("bottom", word.get("top", 0)))) / 2
        for line in lines:
            line_y = sum((float(item.get("top", 0)) + float(item.get("bottom", item.get("top", 0)))) / 2 for item in line) / len(line)
            if abs(line_y - y) <= 3:
                line.append(word)
                break
        else:
            lines.append([word])
    for line in lines:
        line.sort(key=lambda item: float(item.get("x0", 0)))
    return lines


def _word_line_text(line: list[dict[str, object]]) -> str:
    return " ".join(str(word.get("text", "")).strip() for word in line if str(word.get("text", "")).strip())


def _santander_columns_from_words(line: list[dict[str, object]], pending: dict[str, float] | None = None) -> dict[str, float] | None:
    words = [(normalize_name(str(word.get("text", ""))), float(word.get("x0", 0)), float(word.get("x1", 0))) for word in line]
    positions: dict[str, float] = dict(pending or {})
    matched_header_word = False
    for text, x0, x1 in words:
        center = (x0 + x1) / 2
        if text == "DATA":
            positions["date"] = x0
            matched_header_word = True
        elif text.startswith("DESCRICAO"):
            positions["description"] = x0
            matched_header_word = True
        elif text in {"N", "NO", "NUMERO", "DOCUMENTO"}:
            positions.setdefault("document", x0)
            matched_header_word = True
        elif text == "MOVIMENTOS":
            matched_header_word = True
        elif text.startswith("CREDITOS"):
            positions["credit"] = center
            matched_header_word = True
        elif text.startswith("DEBITOS"):
            positions["debit"] = center
            matched_header_word = True
        elif text == "SALDO":
            positions["balance"] = center
            matched_header_word = True
    if {"description", "credit", "debit"}.issubset(positions):
        positions.setdefault("date", 0.0)
        positions.setdefault("document", (positions["description"] + positions["credit"]) / 2)
        positions.setdefault("balance", positions["debit"] + (positions["debit"] - positions["credit"]))
        return positions
    if pending is not None:
        if matched_header_word:
            pending.clear()
            pending.update(positions)
        elif pending:
            pending.clear()
    return None


def _santander_word_line_date(line: list[dict[str, object]], year: int, columns: dict[str, float]) -> date | None:
    for word in line:
        text = str(word.get("text", ""))
        x0 = float(word.get("x0", 0))
        if x0 <= columns["description"] and re.fullmatch(r"\d{2}/\d{2}(?:/\d{2,4})?", text):
            parsed_year = year
            parts = text.split("/")
            if len(parts) == 3:
                parsed_year = int(parts[2])
                parsed_year = 2000 + parsed_year if parsed_year < 100 else parsed_year
            return date(parsed_year, int(parts[1]), int(parts[0]))
    return None


def _santander_word_line_value(line: list[dict[str, object]], columns: dict[str, float]) -> dict[str, object] | None:
    candidates = []
    for word in line:
        text = str(word.get("text", "")).strip()
        if not re.fullmatch(r"-?\d{1,3}(?:\.\d{3})*,\d{2}-?", text):
            continue
        x0 = float(word.get("x0", 0))
        x1 = float(word.get("x1", x0))
        center = (x0 + x1) / 2
        credit_distance = abs(center - columns["credit"])
        debit_distance = abs(center - columns["debit"])
        balance_distance = abs(center - columns["balance"])
        column = "Crédito" if credit_distance <= debit_distance else "Débito"
        movement_distance = min(credit_distance, debit_distance)
        if balance_distance < movement_distance:
            continue
        amount = _santander_parse_movement_value(text)
        if amount is None or amount == Decimal("0.00"):
            continue
        candidates.append({
            "x0": x0,
            "raw_value": text,
            "amount": amount,
            "natureza": "Débito" if text.endswith("-") else column,
            "distance": movement_distance,
        })
    if not candidates:
        return None
    return sorted(candidates, key=lambda item: float(item["distance"]))[0]


def _santander_description_words(line: list[dict[str, object]], columns: dict[str, float], value_x0: object) -> list[str]:
    limit = float(value_x0)
    description_start = columns["description"] - 1
    words = []
    for word in line:
        text = str(word.get("text", "")).strip()
        x0 = float(word.get("x0", 0))
        if not text or text == "-" or re.fullmatch(r"\d{2}/\d{2}(?:/\d{2,4})?", text):
            continue
        if x0 >= limit or x0 >= columns["document"]:
            continue
        if x0 >= description_start:
            words.append(text)
    return words


def _santander_continuation_words(line: list[dict[str, object]], columns: dict[str, float]) -> list[str]:
    description_start = columns["description"] - 1
    return [
        str(word.get("text", "")).strip()
        for word in line
        if str(word.get("text", "")).strip()
        and str(word.get("text", "")).strip() != "-"
        and float(word.get("x0", 0)) >= description_start
        and float(word.get("x0", 0)) < columns["document"]
    ]


def _santander_document_from_words(line: list[dict[str, object]], columns: dict[str, float]) -> str:
    document_words = [
        str(word.get("text", "")).strip()
        for word in line
        if float(word.get("x0", 0)) >= columns["document"]
        and float(word.get("x0", 0)) < min(columns["credit"], columns["debit"])
        and str(word.get("text", "")).strip() != "-"
        and re.fullmatch(r"[\d./-]+", str(word.get("text", "")).strip())
    ]
    return " ".join(document_words)


def _santander_should_skip_word_line(normalized_line: str) -> bool:
    return not normalized_line or bool(re.search(
        r"^(?:DATA|DESCRICAO|N DOCUMENTO|MOVIMENTOS|CREDITOS|DEBITOS|SALDO|SALDO EM 31 12|RESUMO|NOME|AGENCIA|EXTRATO(?:\b| ).*|PAGINA(?:\b| ).*|BALP(?:\b| ).*)$",
        normalized_line,
        re.I,
    ))


def _santander_informational_section_starts(normalized_line: str) -> bool:
    return bool(re.search(
        r"\b(?:DEBITO AUTOMATICO|COMPROVANTES DE PAGAMENTO|TRANSFERENCIAS ENTRE CONTAS|DOCS|TEDS E PIXS ENVIADOS|CONTAMAX EMPRESARIAL|INVESTIMENTOS|SALDOS POR PERIODO)\b",
        normalized_line,
        re.I,
    ))


def _dedupe_santander_description(history: str) -> str:
    parts = [part for part in history.split() if part != "-"]
    deduped = []
    for part in parts:
        if deduped and normalize_name(deduped[-1]) == normalize_name(part):
            continue
        deduped.append(part)
    for size in range(len(deduped) // 2, 1, -1):
        previous = deduped[-2 * size : -size]
        repeated = deduped[-size:]
        if normalize_name(" ".join(previous)) == normalize_name(" ".join(repeated)):
            deduped = deduped[:-size]
            break
    return " ".join(deduped)


def _validate_santander_expected_statement(records: list[ParsedStatement], pages_text: list[str]) -> None:
    text = "\n".join(pages_text)
    credit = sum((record.valor or Decimal()) for record in records if record.natureza == "Crédito")
    debit = sum((record.valor or Decimal()) for record in records if record.natureza == "Débito")
    summary_credit = _santander_summary_amount(text, r"Total de Cr[eé]ditos")
    summary_debit = _santander_summary_amount(text, r"Total de D[eé]bitos")
    summary_errors = []
    if summary_credit is not None and credit != summary_credit:
        summary_errors.append(f"créditos {credit} != {summary_credit}")
    if summary_debit is not None and debit != summary_debit:
        summary_errors.append(f"débitos {debit} != {summary_debit}")
    if summary_errors:
        raise ValueError("Falha de validação do extrato Santander: " + "; ".join(summary_errors))

    if "janeiro/2024" not in text.lower() or "47.224,73" not in text:
        return
    errors = []
    if len(records) != SANTANDER_EXPECTED_RECORDS:
        errors.append(f"quantidade {len(records)} != {SANTANDER_EXPECTED_RECORDS}")
    if credit != SANTANDER_EXPECTED_CREDIT:
        errors.append(f"créditos {credit} != {SANTANDER_EXPECTED_CREDIT}")
    if debit != SANTANDER_EXPECTED_DEBIT:
        errors.append(f"débitos {debit} != {SANTANDER_EXPECTED_DEBIT}")
    initial_balance = _santander_summary_amount(text, r"Saldo de Conta Corrente em 31/12")
    if initial_balance is not None and initial_balance != SANTANDER_EXPECTED_INITIAL_BALANCE:
        errors.append(f"saldo inicial {initial_balance} != {SANTANDER_EXPECTED_INITIAL_BALANCE}")
    final_balance = _santander_summary_amount(text, r"Saldo de Conta Corrente em 31/01")
    if final_balance is not None and final_balance != SANTANDER_EXPECTED_FINAL_BALANCE:
        errors.append(f"saldo final {final_balance} != {SANTANDER_EXPECTED_FINAL_BALANCE}")
    if errors:
        raise ValueError("Falha de validação do extrato Santander: " + "; ".join(errors))


def _santander_summary_amount(text: str, label: str) -> Decimal | None:
    match = re.search(label + r".{0,120}?(-?[\d.]+,\d{2}-?)", text, re.I | re.S)
    return _santander_parse_movement_value(match.group(1)) if match else None


def _extract_santander_statement(text: str, page_number: int) -> list[ParsedStatement]:
    """Extracts Santander current-account statements without reusing bank-specific BB rules."""
    page_parts = _split_santander_text_pages(text)
    if len(page_parts) > 1:
        records: list[ParsedStatement] = []
        current_date: date | None = None
        period = _santander_statement_period(text)
        for fallback_index, (part_page_number, part_text) in enumerate(page_parts):
            part_records, current_date = _extract_santander_statement_page(
                part_text,
                part_page_number or page_number + fallback_index,
                period,
                current_date,
            )
            records.extend(part_records)
        return records

    records, _ = _extract_santander_statement_page(text, page_number, None, None)
    return records


def _extract_santander_statement_page(
    text: str,
    page_number: int,
    known_period: tuple[int, int] | None,
    initial_date: date | None,
) -> tuple[list[ParsedStatement], date | None]:
    period = _santander_statement_period(text)
    if not period:
        period = known_period
    if not period:
        return [], initial_date

    row_records, last_date = _extract_santander_structured_rows(text, page_number, period, initial_date)
    if row_records:
        return row_records, last_date

    row_records = _extract_santander_inline_rows(text, page_number, period[1])
    if row_records:
        return row_records, row_records[-1].data or initial_date

    row_records = _extract_santander_single_date_column(text, page_number, period)
    return row_records, row_records[-1].data if row_records else initial_date


def _split_santander_text_pages(text: str) -> list[tuple[int | None, str]]:
    parts = [part.strip() for part in re.split(r"(?=Pagina:\d+/\d+)", text) if part.strip()]
    if len(parts) <= 1:
        return [(None, text)]
    result = []
    for part in parts:
        match = re.search(r"Pagina:(\d+)/\d+", part)
        result.append((int(match.group(1)) if match else None, part))
    return result


def _santander_statement_period(text: str) -> tuple[int, int] | None:
    match = re.search(r"\b([a-zç]+)\s*/\s*(20\d{2})\b", text, re.I)
    if not match:
        return None
    month = PT_MONTHS.get(match.group(1).lower())
    if not month:
        return None
    return month, int(match.group(2))


def _extract_santander_structured_rows(
    text: str,
    page_number: int,
    period: tuple[int, int],
    initial_date: date | None = None,
) -> tuple[list[ParsedStatement], date | None]:
    if not _is_santander_current_account_statement_page(text):
        return [], initial_date

    results: list[ParsedStatement] = []
    current: dict[str, object] | None = None
    current_date: date | None = initial_date

    def flush() -> None:
        nonlocal current
        if not current:
            return
        history = _santander_clean_description(str(current["historico"]))
        if not history or _santander_should_skip_history(history):
            current = None
            return
        raw_value = str(current["raw_value"])
        results.append(ParsedStatement(
            data=current["data"],  # type: ignore[arg-type]
            hora=None,
            historico=history,
            nome=_santander_statement_name(history),
            valor=current["valor"],  # type: ignore[arg-type]
            natureza=_santander_nature(history, raw_value),
            texto_original=str(current["texto_original"]),
            pagina_numero=page_number,
            numero_documento=str(current.get("numero_documento") or ""),
        ))
        current = None

    for raw_line in text.splitlines():
        line = " ".join(raw_line.split())
        if not line or _santander_should_skip_line(line):
            continue

        parsed = _santander_parse_structured_line(line, current_date, period[1])
        if parsed:
            flush()
            current_date = parsed["data"]
            current = parsed
            continue

        line_date = _santander_line_date(line, period[1])
        if line_date:
            current_date = line_date
            continue

        if current and not _santander_is_summary_or_section_line(line):
            current["historico"] = f"{current['historico']} {line}"
            current["texto_original"] = f"{current['texto_original']}\n{raw_line.strip()}"

    flush()
    return results, current_date


def _is_santander_current_account_statement_page(text: str) -> bool:
    normalized = normalize_name(text)
    if "EXTRATO CONSOLIDADO INTELIGENTE" not in normalized and "SANTANDER" not in normalized:
        return False
    if "MOVIMENTOS" not in normalized or "DESCRICAO" not in normalized:
        return False
    return not bool(re.search(r"\b(?:SALDOS POR PERIODO|COMPROVANTES DE PAGAMENTO|INVESTIMENTOS|MOVIMENTACAO MENSAL)\b", normalized))


def _santander_parse_structured_line(line: str, current_date: date | None, year: int) -> dict[str, object] | None:
    working = line
    line_date = _santander_line_date(working, year)
    if line_date:
        working = re.sub(r"^\d{2}/\d{2}(?:/\d{2,4})?\s*", "", working).strip()
    elif current_date:
        line_date = current_date
    else:
        return None

    value = _santander_first_movement_value(working)
    if not value:
        return None

    value_start, raw_value = value
    amount = _santander_parse_movement_value(raw_value)
    if amount is None or amount == Decimal("0.00"):
        return None

    before_value = working[:value_start].strip()
    if not before_value or _santander_is_summary_or_section_line(before_value):
        return None

    description, document = _santander_split_document(before_value, working, value_start)
    description = _santander_clean_description(description)
    if not description or not _santander_is_movement_history(description):
        return None

    return {
        "data": line_date,
        "historico": description,
        "valor": amount,
        "raw_value": raw_value,
        "texto_original": line,
        "numero_documento": document,
    }


def _santander_line_date(line: str, year: int) -> date | None:
    match = re.match(r"^\s*(\d{2})/(\d{2})(?:/(\d{2,4}))?\b", line)
    if not match:
        return None
    parsed_year = int(match.group(3)) if match.group(3) else year
    parsed_year = 2000 + parsed_year if parsed_year < 100 else parsed_year
    return date(parsed_year, int(match.group(2)), int(match.group(1)))


def _santander_first_movement_value(line: str) -> tuple[int, str] | None:
    for match in re.finditer(r"\d{4,6}(?P<amount>\d{1,3}(?:\.\d{3})*,\d{2}-?)", line):
        amount_text = match.group("amount")
        amount = _santander_parse_movement_value(amount_text)
        if amount and amount != Decimal("0.00"):
            return match.start("amount"), amount_text
    for match in re.finditer(r"(?<![\d.])-?\d{1,3}(?:\.\d{3})*,\d{2}-?", line):
        amount = _santander_parse_movement_value(match.group(0))
        if amount and amount != Decimal("0.00"):
            return match.start(), match.group(0)
    return None


def _santander_split_document(before_value: str, full_line: str, value_start: int) -> tuple[str, str]:
    tight_document = re.search(r"(?P<document>\d{4,})$", full_line[:value_start])
    if tight_document and before_value.endswith(tight_document.group("document")):
        document = tight_document.group("document")
        return before_value[: -len(document)].strip(), document

    spaced_document = re.match(r"(?P<description>.+?)\s+(?P<document>\d{4,})$", before_value)
    if spaced_document:
        return spaced_document.group("description").strip(), spaced_document.group("document")
    return before_value, ""


def _extract_santander_inline_rows(text: str, page_number: int, year: int) -> list[ParsedStatement]:
    results = []
    pattern = re.compile(r"(?m)^\s*(?P<day>\d{2})/(?P<month>\d{2})(?:/(?P<year>\d{2,4}))?\s*(?P<body>.+?)\s+(?P<value>-?[\d.]+,\d{2}-?)\s*$")
    for match in pattern.finditer(text):
        body = " ".join(match.group("body").split())
        if not _santander_is_movement_history(body):
            continue
        parsed_year = int(match.group("year")) if match.group("year") else year
        parsed_year = 2000 + parsed_year if parsed_year < 100 else parsed_year
        amount = parse_brl(match.group("value"))
        if amount is None:
            continue
        nature = _santander_nature(body, match.group("value"))
        if _santander_should_skip_history(body):
            continue
        results.append(ParsedStatement(
            data=date(parsed_year, int(match.group("month")), int(match.group("day"))),
            hora=None,
            historico=body,
            nome=_santander_statement_name(body),
            valor=abs(amount),
            natureza=nature,
            texto_original=match.group(0).strip(),
            pagina_numero=page_number,
        ))
    return results


def _extract_santander_single_date_column(text: str, page_number: int, period: tuple[int, int]) -> list[ParsedStatement]:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if "Descrição" not in lines or "Movimentos (R$)" not in lines:
        return []

    dates = [item for item in lines if re.fullmatch(r"\d{2}/\d{2}", item)]
    if len(dates) != 1:
        # Later Santander pages often print all dates, then all descriptions, then
        # all amounts. Plain text loses row alignment, so do not guess the dates.
        return []

    try:
        desc_start = lines.index(dates[0]) + 1
        movements_index = lines.index("Movimentos (R$)")
    except ValueError:
        return []

    raw_descriptions = [
        line for line in lines[desc_start:movements_index]
        if line not in {"Descrição", "Nº Documento"} and not line.startswith("SALDO EM")
    ]
    descriptions = _santander_group_descriptions(raw_descriptions)
    values = _santander_movement_values(lines[movements_index + 1:])
    if values and values[0][0] == Decimal("0.00"):
        values = values[1:]
    descriptions = _santander_align_descriptions_to_values(descriptions, values)

    if not descriptions or len(descriptions) != len(values):
        logger.info(
            "Extrato Santander ignorado por perda de alinhamento: pagina=%s descricoes=%s valores=%s",
            page_number,
            len(descriptions),
            len(values),
        )
        return []

    day, month = map(int, dates[0].split("/"))
    statement_date = date(period[1], month, day)
    results = []
    for history, (amount, raw_value) in zip(descriptions, values):
        if not _santander_is_movement_history(history):
            continue
        results.append(ParsedStatement(
            data=statement_date,
            hora=None,
            historico=history,
            nome=_santander_statement_name(history),
            valor=amount,
            natureza=_santander_nature(history, raw_value),
            texto_original=f"{dates[0]} {history} {raw_value}",
            pagina_numero=page_number,
        ))
    return results


def _santander_group_descriptions(lines: list[str]) -> list[str]:
    starts = (
        "TARIFA", "TED ", "TRANSFERENCIA", "TRANSF ", "PAGAMENTO", "DEBITO",
        "DÉBITO", "JUROS", "IOF", "APLICACAO", "APLICAÇÃO", "ANTECIPACAO",
        "ANTECIPAÇÃO", "RESGATE", "PGTO ",
    )
    groups: list[list[str]] = []
    for line in lines:
        normalized = line.upper()
        if normalized.startswith("TRANSFERENCIA ENTRE CONTA") and groups and "TED RECEBIDA" in groups[-1][0].upper():
            groups[-1].append(line)
        elif normalized.startswith(starts) or not groups:
            groups.append([line])
        else:
            groups[-1].append(line)
    return [" ".join(group) for group in groups]


def _santander_align_descriptions_to_values(descriptions: list[str], values: list[tuple[Decimal, str]]) -> list[str]:
    if len(descriptions) == len(values):
        return descriptions
    if len(descriptions) != len(values) + 1:
        return descriptions

    for index, (history, (_, raw_value)) in enumerate(zip(descriptions, values)):
        if _santander_alignment_score(history, raw_value) < 0:
            return descriptions[:index] + descriptions[index + 1:]

    baseline = sum(_santander_alignment_score(history, raw_value) for history, (_, raw_value) in zip(descriptions, values))
    best_score = baseline
    best_descriptions = descriptions
    for index in range(len(descriptions)):
        candidate = descriptions[:index] + descriptions[index + 1:]
        score = sum(_santander_alignment_score(history, raw_value) for history, (_, raw_value) in zip(candidate, values))
        if score > best_score:
            best_score = score
            best_descriptions = candidate
    return best_descriptions


def _santander_alignment_score(history: str, raw_value: str) -> int:
    expected = _santander_keyword_nature(history)
    if not expected:
        return 0
    raw_debit = raw_value.strip().endswith("-")
    if expected == "Débito":
        return 3 if raw_debit else -2
    return 3 if not raw_debit else -2


def _santander_movement_values(lines: list[str]) -> list[tuple[Decimal, str]]:
    values = []
    for line in lines:
        if line in {"Créditos", "Débitos", "DébitosSaldo (R$)", "Saldo (R$)", "-"}:
            continue
        if line.startswith(("Extrato_", "BALP_", "Pagina:")):
            break
        amount = _santander_parse_movement_value(line)
        if amount is not None:
            values.append((amount, line))
    return values


def _santander_parse_movement_value(value: str) -> Decimal | None:
    match = re.search(r"(?P<amount>\d{1,3}(?:\.\d{3})*,\d{2})", value.strip())
    if not match:
        return None
    amount = parse_brl(match.group("amount"))
    return abs(amount) if amount is not None else None


def _santander_is_movement_history(history: str) -> bool:
    return bool(re.search(
        r"\b(TARIFA|TED|TRANSFERENCIA|TRANSF|PAGAMENTO|PGTO|DEBITO|DÉBITO|JUROS|IOF|APLICACAO|APLICAÇÃO|ANTECIPACAO|ANTECIPAÇÃO|RESGATE)\b",
        history,
        re.I,
    ))


def _santander_should_skip_history(history: str) -> bool:
    if SANTANDER_IGNORE_AUTOMATIC_INVESTMENT_MOVEMENTS and SANTANDER_AUTOMATIC_INVESTMENT_RE.search(history):
        return True
    return bool(re.search(r"\b(?:SALDO EM|SALDO DE|SALDO DISPON[IÍ]VEL|TOTAL DE CR[EÉ]DITOS|TOTAL DE D[EÉ]BITOS)\b", history, re.I))


def _santander_should_skip_line(line: str) -> bool:
    return _santander_is_summary_or_section_line(line) or bool(re.search(
        r"(?:EXTRATO_PJ|BALP_|PAGINA:|CENTRAL DE ATENDIMENTO|OUVIDORIA|SAC|WWW\.SANTANDER|PUBLICIDADE)",
        line,
        re.I,
    ))


def _santander_is_summary_or_section_line(line: str) -> bool:
    return bool(re.search(
        r"^(?:Data|Descrição|N[º°] Documento|Movimentos \(R\$\)|Créditos|Débitos|Saldo \(R\$\)|Conta Corrente|Movimentação|Resumo|Nome|Agência|Saldo|Total|Depósitos|Pagamentos|Outros|Saldos por Período|Comprovantes de Pagamento|Transferências entre Contas|Investimentos)\b",
        line,
        re.I,
    ))


def _santander_clean_description(value: str) -> str:
    return " ".join(value.replace("Descrição", " ").replace("Nº Documento", " ").split())


def _santander_nature(history: str, raw_value: str) -> str:
    keyword_nature = _santander_keyword_nature(history)
    if keyword_nature:
        return keyword_nature
    return "Débito" if raw_value.strip().endswith("-") else "Crédito"


def _santander_keyword_nature(history: str) -> str | None:
    normalized = normalize_name(history)
    if re.search(r"\b(RECEBIDA|ANTECIPACAO GETNET|ANTECIPAÇÃO GETNET|RESGATE CONTAMAX|GETNET)\b", normalized, re.I):
        return "Crédito"
    if re.search(r"\b(TARIFA|APLICACAO CONTAMAX|APLICAÇÃO CONTAMAX|DEBITO|DÉBITO|JUROS|IOF|PAGAMENTO DARF|PAGAMENTO FGTS|PAGAMENTO DE BOLETO|PGTO CONTA|TED ENVIADA|TRANSF VALOR)\b", normalized, re.I):
        return "Débito"
    return None


def _santander_statement_name(history: str) -> str:
    cleaned = re.sub(
        r"\b(?:TARIFA|TED|RECEBIDA|ENVIADA|TRANSFERENCIA|TRANSF|PAGAMENTO|CARTAO|DE|DEBITO|CREDITO|APLICACAO|ANTECIPACAO|RESGATE|CONTAMAX|AUTOMATICO|JUROS|IOF|PGTO|CONTA|CANAIS|INTERNET|EM|OUTROS|BANCOS|BOLETO)\b",
        " ",
        history,
        flags=re.I,
    )
    name = " ".join(cleaned.split())
    name_tokens = set(normalize_name(name).split())
    history_tokens = set(normalize_name(history).split())
    return "" if name_tokens and name_tokens <= history_tokens else name


def _extract_basa_statement(text: str, page_number: int) -> list[ParsedStatement]:
    if "VALOR LANCTO" not in text or "D/C" not in text:
        return []

    amount_pattern = r"-?\d{1,3}(?:\.\d{3})*,\d{2}-?"
    row_pattern = re.compile(
        rf"^\s*(?P<data>\d{{2}}/\d{{2}}/\d{{4}})\s+"
        rf"(?P<documento>\d+)\s+"
        rf"(?P<historico>.+?)\s+"
        rf"(?P<valor>{amount_pattern})\s+"
        rf"(?P<natureza>[CD])\s+"
        rf"(?P<saldo>{amount_pattern})\s*$",
        re.M,
    )
    results: list[ParsedStatement] = []
    for match in row_pattern.finditer(text):
        parsed_date, _ = parse_date_time(match.group("data"))
        amount = parse_brl(match.group("valor"))
        if not parsed_date or amount is None:
            continue
        history = " ".join(match.group("historico").split())
        if _basa_should_skip_history(history):
            continue
        results.append(ParsedStatement(
            data=parsed_date,
            hora=None,
            historico=history,
            nome=_basa_statement_name(history),
            valor=abs(amount),
            natureza="Crédito" if match.group("natureza") == "C" else "Débito",
            texto_original=match.group(0).strip(),
            pagina_numero=page_number,
            numero_documento=match.group("documento").strip(),
        ))
    return results


def _statement_period(text: str) -> tuple[date, date] | None:
    url_period = re.search(r"hdnDataInicio=(\d{2}/\d{2}/\d{4}).*?hdnDataFinal=(\d{2}/\d{2}/\d{4})", text, re.I | re.S)
    if url_period:
        start, _ = parse_date_time(url_period.group(1))
        end, _ = parse_date_time(url_period.group(2))
        if start and end:
            return start, end
    explicit_period = re.search(r"(?m)^\s*(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s*$", text)
    if explicit_period:
        start, _ = parse_date_time(explicit_period.group(1))
        end, _ = parse_date_time(explicit_period.group(2))
        if start and end:
            return start, end
    between_period = re.search(r"Entre\s+(\d{2}/\d{2}/\d{4})\s+e\s+(\d{2}/\d{2}/\d{4})", text, re.I)
    if between_period:
        start, _ = parse_date_time(between_period.group(1))
        end, _ = parse_date_time(between_period.group(2))
        if start and end:
            return start, end
    dates = [parse_date_time(match.group(0))[0] for match in re.finditer(DATE_RE, text)]
    dates = [item for item in dates if item]
    if len(dates) < 2:
        return None
    return dates[0], dates[1]


def _in_statement_period(parsed_date: date | None, period: tuple[date, date] | None) -> bool:
    return bool(parsed_date and (not period or period[0] <= parsed_date <= period[1]))


def _extract_caixa_statement(text: str, page_number: int) -> list[ParsedStatement]:
    normalized = normalize_name(text)
    has_period_url = "hdnDataInicio=" in text and "hdnDataFinal=" in text
    if "GERENCIADOR CAIXA" not in normalized or ("EXTRATO POR PERIODO" not in normalized and not has_period_url):
        return []
    lines = _lines(text)
    period = _statement_period(text)
    results: list[ParsedStatement] = []
    index = 0
    while index < len(lines) - 3:
        parsed_date, _ = parse_date_time(lines[index])
        if not parsed_date:
            index += 1
            continue
        document = lines[index + 1].strip()
        history = lines[index + 2].strip()
        raw_value = lines[index + 3].strip()
        if not re.fullmatch(r"\d+", document) or not amount_text(raw_value):
            index += 1
            continue
        if _basa_should_skip_history(history) or not _in_statement_period(parsed_date, period):
            index += 4
            continue
        amount = parse_brl(raw_value)
        if amount is None:
            index += 1
            continue
        nature = "Crédito" if raw_value.upper().endswith("C") else "Débito"
        raw = "\n".join(lines[index:index + 5])
        results.append(ParsedStatement(
            data=parsed_date,
            hora=None,
            historico=history,
            nome="",
            valor=abs(amount),
            natureza=nature,
            texto_original=raw,
            pagina_numero=page_number,
            numero_documento=document,
        ))
        index += 5 if index + 4 < len(lines) and amount_text(lines[index + 4]) else 4
    return results


def _extract_bradesco_statement_pages(pages: list[str]) -> list[ParsedStatement]:
    records: list[ParsedStatement] = []
    current_date: date | None = None
    period = _statement_period("\n".join(pages))
    for page_number, text in enumerate(pages, 1):
        page_records, current_date = _extract_bradesco_statement_page(text, page_number, current_date, period)
        records.extend(page_records)
    return records


def _extract_bradesco_statement_page(text: str, page_number: int, initial_date: date | None = None, period: tuple[date, date] | None = None) -> tuple[list[ParsedStatement], date | None]:
    normalized = normalize_name(text)
    if "EXTRATO MENSAL POR PERIODO" not in normalized and "EXTRATO DE AG" not in normalized:
        return [], initial_date
    lines = _lines(text)
    current_date = initial_date
    results: list[ParsedStatement] = []
    index = 0
    while index < len(lines):
        parsed_date, _ = parse_date_time(lines[index])
        if parsed_date and re.fullmatch(DATE_RE, lines[index]):
            current_date = parsed_date
            index += 1
            continue
        if not current_date:
            index += 1
            continue
        if _basa_should_skip_history(lines[index]) or _bradesco_should_skip_line(lines[index]):
            index += 1
            continue
        if amount_text(lines[index]):
            index += 1
            continue
        doc_index = -1
        for candidate in range(index + 1, min(index + 5, len(lines) - 1)):
            candidate_date, _ = parse_date_time(lines[candidate])
            if candidate_date and re.fullmatch(DATE_RE, lines[candidate]):
                doc_index = -2
                break
            if re.fullmatch(r"\d+", lines[candidate]) and amount_text(lines[candidate + 1]):
                doc_index = candidate
                break
        if doc_index == -2:
            index = candidate
            continue
        if doc_index < 0:
            index += 1
            continue
        description = " ".join(lines[index:doc_index]).strip()
        raw_value = lines[doc_index + 1].strip()
        amount = parse_brl(raw_value)
        if not description or amount is None or _basa_should_skip_history(description) or not _in_statement_period(current_date, period):
            index = doc_index + 3
            continue
        nature = "Débito" if raw_value.startswith("-") else "Crédito"
        name = _bradesco_statement_name(description)
        results.append(ParsedStatement(
            data=current_date,
            hora=None,
            historico=description,
            nome=name,
            valor=abs(amount),
            natureza=nature,
            texto_original="\n".join(lines[index:doc_index + 3]),
            pagina_numero=page_number,
            numero_documento=lines[doc_index],
        ))
        index = doc_index + 3
    return results, current_date


def _bradesco_should_skip_line(line: str) -> bool:
    normalized = normalize_name(line)
    return normalized in {
        "DATA",
        "LANCAMENTO",
        "DCTO",
        "CREDITO R",
        "DEBITO R",
        "SALDO R",
        "TOTAL",
        "FOLHA",
        "EXTRATO MENSAL POR PERIODO",
    } or normalized.startswith(("AGENCIA CONTA", "TOTAL DISPONIVEL", "NOME USUARIO", "DATA OPERACAO", "SAC SERVICO", "OUVIDORIA"))


def _bradesco_statement_name(history: str) -> str:
    parts = history.split()
    if len(parts) > 1 and parts[0].upper().rstrip(":") in {"REM", "DES", "CONTR"}:
        return " ".join(parts[1:]).strip()
    if "\n" in history:
        return history.splitlines()[-1].strip()
    return history


def _basa_should_skip_history(history: str) -> bool:
    return bool(re.search(r"\b(?:SALDO|TOTAL DE|LIMITE|VENCTO|TIPO CONTA)\b", normalize_name(history), re.I))


def _basa_statement_name(history: str) -> str:
    return re.sub(r"^\d+\s*-\s*", "", history).strip()


def _extract_banco_do_brasil_statement(text: str, page_number: int) -> list[ParsedStatement]:
    """Extracts the vertical text layout produced by Banco do Brasil statements."""
    pattern = re.compile(
        r"(?m)^(?P<data>\d{2}/\d{2}/\d{4})\n"
        r"\d{4}\n"
        r"(?P<historico>[^\n]+)\n"
        r"(?P<documento>[^\n]+?)(?:\n|[ \t]+)"
        r"(?P<valor>[\d.]+,\d{2})[ \t]+(?P<natureza>[CD])(?:[ \t]+[^\n]+)?"
        r"(?P<detalhe>(?:\n(?!\d{2}/\d{2}/\d{4}\n)[^\n]+)*)"
    )
    results = []
    for match in pattern.finditer(text):
        parsed_date, _ = parse_date_time(match.group("data"))
        amount = parse_brl(match.group("valor"))
        # BB codes identify the operation and counterparty. Keep them intact so
        # the extracted statement remains auditable and can be matched precisely.
        document = " ".join(match.group("documento").split())
        full_document = document if re.search(r"\s", document) or _banco_do_brasil_contract_history(match.group("historico")) else ""
        history = " ".join(part for part in [" ".join(match.group("historico").split()), full_document] if part)
        if re.search(r"\b(?:saldo anterior|saldo do dia|saldo final|limite|valor total devido|cheque especial)\b", history, re.I):
            continue
        detail = match.group("detalhe").strip()
        _, hour = parse_date_time(detail)
        detail = re.sub(r"^\d{2}/\d{2}(?:\s+\d{2}:\d{2}(?::\d{2})?)?\s+", "", detail)
        name = detail.strip()
        origin_date = re.search(r"^\s*(\d{2}/\d{2})(?:\s|$)", detail)
        data_origem = origin_date.group(1) if origin_date and re.search(r"transfer[êe]ncia", history, re.I) else ""
        if data_origem and name == data_origem:
            name = ""
        results.append(ParsedStatement(
            data=parsed_date,
            hora=hour,
            historico=history,
            nome=name,
            valor=amount,
            natureza="Crédito" if match.group("natureza") == "C" else "Débito",
            texto_original=match.group(0).strip(),
            pagina_numero=page_number,
            data_origem=data_origem,
            numero_documento=document,
        ))
    return results


def _banco_do_brasil_contract_history(history: str) -> bool:
    normalized = normalize_name(history)
    return bool(re.search(r"\b(?:CAP GIRO|AMORTIZACAO|EMPRESTIMO|FINANCIAMENTO|PARCELA|CONTRATO|OPERACAO|CREDITO RURAL|COMERCIAL)\b", normalized))


def deduplicate_statement_records(records: list[ParsedStatement]) -> list[ParsedStatement]:
    """Keeps a repeated transaction at a PDF page boundary only once."""
    seen_pages: dict[tuple[object, ...], int] = {}
    unique = []
    for record in records:
        key = (normalize_name(record.texto_original), record.data, record.valor, record.natureza)
        previous_page = seen_pages.get(key)
        if previous_page is not None and previous_page != record.pagina_numero:
            logger.info("Extrato BB duplicado entre páginas ignorado: data=%s historico=%s valor=%s natureza=%s", record.data, record.historico, record.valor, record.natureza)
            continue
        seen_pages[key] = record.pagina_numero
        unique.append(record)
    return unique
